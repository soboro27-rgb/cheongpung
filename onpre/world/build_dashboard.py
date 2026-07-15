#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
매출처 마케팅 대시보드 빌더
실행: python build_dashboard.py [25년파일.xlsx] [26년파일.xlsx]
파일 생략 시 데모 데이터로 실행
"""
import sys, json, re, warnings, random, math
from pathlib import Path
from datetime import datetime
from collections import defaultdict, Counter

warnings.filterwarnings('ignore')

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False
    print("[주의] pandas/openpyxl 미설치. 데모 모드로 실행합니다.")
    print("   pip install pandas openpyxl")

OUTPUT_FILE = Path(__file__).parent / 'sales_dashboard.html'

COL_KR = {
    '주문자명':'orderer_name',
    '주문상태':'order_status','주문묶음번호':'bundle_no','판매사이트':'site',
    '품목코드':'item_code','품명':'item_name','규격':'spec','상품명':'product_name',
    '수량':'qty','단가':'unit_price','판매가':'sale_price','부가세':'vat',
    '합계금액':'total','수수료':'fee','수수료율(%)':'fee_rate','정산금액':'settlement',
    '재고원가':'cost','할인가':'discount','마진':'margin','세액포함 여부':'tax_incl',
    '영업담당':'sales_rep','주문일자':'order_date','쇼핑몰 주문번호':'shop_order_no',
    '수령자명':'recipient_name','수령자전화번호':'recipient_phone',
    '수령자핸드폰':'recipient_mobile',
    '우편번호':'postal_code',
    '주소1':'addr1','주소2':'addr2','대분류':'cat1','중분류':'cat2','소분류':'cat3',
    '배송방법':'delivery','운송장번호1':'tracking1','운송장번호2':'tracking2',
    '수주상태':'order_recv_status','작지상태':'work_status','작업지시번호':'work_no',
    '작업지시량':'work_qty','실적수량':'perf_qty','검사수량':'insp_qty',
    '출고수량':'ship_qty','반품수량':'return_qty',
}

# 플레이오토 컬럼 → 내부 스키마 매핑
COL_PLAYAUTO = {
    '쇼핑몰':           'site',
    '주문자명':          'orderer_name',
    '수령자명':          'recipient_name',
    '우편번호':          'postal_code',
    '주소':             'addr1',
    '수령자전화번호':     'recipient_phone',
    '수령자휴대폰번호':   'recipient_mobile',
    '주문자휴대폰번호':   'orderer_mobile',
    '주문자전화번호':     'orderer_phone',
    '금액':             'total',
    '묶음번호':          'bundle_no',
    '주문일':            'order_date',
    '온라인상품명':       'item_name',
    '판매자관리코드':     'item_code',
    '주문수량':          'qty',
    '주문상태':          'order_status',
    '배송방법':          'delivery',
    '운송장번호':        'tracking1',
    'SKU상품명':        'product_name',
    '실결제금액':        'settlement',
    '쇼핑몰주문번호':    'shop_order_no',
}

CORP_RE = re.compile(
    r'[\(\（]주[\)\）]|^주식회사\s*|\s*주식회사$|㈜|\s*\(주\)|\s*코퍼레이션|\s*Corp\.?|\s*Inc\.?',
    re.IGNORECASE
)

REGIONS = {
    '100':'서울 중구','110':'서울 종로','120':'서울 강남','130':'서울 강서',
    '140':'인천 남동','150':'인천 서구','160':'경기 수원','170':'경기 성남',
    '180':'경기 고양','190':'경기 용인','200':'춘천','210':'강릉','220':'원주',
    '300':'대전 서구','310':'청주 흥덕','320':'충주','400':'부산 해운대',
    '410':'부산 부산진','420':'울산 중구','430':'창원 의창',
    '500':'광주 서구','510':'전주 완산','520':'목포',
    '600':'대구 수성','610':'대구 달서','620':'구미 신평','700':'제주시',
}

CHANNELS = ('쿠팡', '네이버/스마트스토어', 'G마켓/옥션/11번가', '번개장터', '렌탈', '기업직납', '자사몰/직판', '일반소비자', '딜러/파트너')
# 채널 그룹 — 탭 필터용 (상위 3개)
CHANNEL_GROUPS = {
    '오픈마켓': ('쿠팡', '네이버/스마트스토어', 'G마켓/옥션/11번가', '번개장터'),
    'B2B':      ('렌탈', '기업직납', '자사몰/직판', '딜러/파트너'),
    'B2C':      ('일반소비자',),
}
_SELF_MALL = ('리뉴올', '컴퓨존', '굿데이몰', '팝니다닷컴', '중고전자몰', '초이스샵', '딜러샵', '클링', '리뉴올몰')
_CORP_PREFIX = re.compile(r'^\s*(?:주식회사|[\(\（]주[\)\）]|㈜|\(유\)|\(사\)|\(재\)|\(재단\))')

def classify_channel(site) -> str:
    s = str(site) if site and not (isinstance(site, float) and math.isnan(site)) else ''
    if not s: return '딜러/파트너'
    if '일반소비자' in s:                               return '일반소비자'
    if '쿠팡' in s or '로켓그로스' in s:               return '쿠팡'
    if '스토어팜' in s or '네이버페이' in s:           return '네이버/스마트스토어'
    if '지마켓' in s or '옥션' in s or '11번가' in s:  return 'G마켓/옥션/11번가'
    if '번개장터' in s:                                  return '번개장터'
    if '렌탈' in s:                                      return '렌탈'
    if any(k in s for k in _SELF_MALL):                  return '자사몰/직판'
    if _CORP_PREFIX.match(s):                             return '기업직납'
    return '딜러/파트너'

# ─────────────────────────── 업종 분류 ──────────────────────────────────────

_RESI_RE = re.compile(
    r'아파트|아이파크|자이|래미안|푸르지오|힐스테이트|롯데캐슬|e편한세상|더샵|위브|두산|'
    r'현대홈타운|벽산|대림|코아루|부영|대우|한일유|사랑으로|'
    r'\d+동\s*\d+호|\d+단지|빌라|원룸|하이빌|맨션|파크뷰|스카이뷰'
)

_INDUSTRY_RULES = [
    ('교육',     '#f59e0b', ('학원','학교','대학교','대학원','교육원','공학관','어린이집','유치원','어학원','캠퍼스','기숙사','독서실','교습소')),
    ('의료/복지', '#ef4444', ('병원','요양원','요양센터','요양병원','의원','클리닉','약국','한의원','치과','안과','복지관','복지센터','데이케어','육아원','재가센터')),
    ('금융/보험', '#0ea5e9', ('은행','신협','새마을금고','농협','저축은행','금융','보험','증권','자산운용','캐피탈','대부','투자','손해보험','생명보험','화재보험','카드사','카드','신용금고')),
    ('외식/음식', '#f97316', ('치킨','피자','카페','커피','플레이스','식당','레스토랑','베이커리','분식','포차','주점','노래연습','노래방','코인노래','뷔페','쌀국수','냉면','삼겹','구이','고깃집','순대')),
    ('제조/산업', '#3b82f6', ('산업','공장','제조','플라스틱','금속','기계','소재','철강','화학','섬유','가구','인쇄','포장','물류창고','성형')),
    ('종교',     '#8b5cf6', ('교회','성당','사찰','절','성전','교당','수도원')),
    ('공공/기관', '#14b8a6', ('구청','시청','도청','경찰','소방','군청','행정','주민센터','우체국')),
    ('판매/유통', '#06b6d4', ('마트','슈퍼','편의점','백화점','쇼핑','매장','대리점','판매점','상회','상사','도매','유통','스토어')),
]

def classify_industry(addr: str, company_hint: str) -> str:
    text = f"{addr or ''} {company_hint or ''}".strip()
    if not text:
        return '미분류'
    for ind_name, _, keywords in _INDUSTRY_RULES:
        if any(kw in text for kw in keywords):
            return ind_name
    if _RESI_RE.search(text) and not company_hint:
        return '개인/주거'
    if company_hint:
        return '기타 사업체'
    if _RESI_RE.search(text):
        return '개인/주거'
    return '미분류'

IND_COLORS = {name: color for name, color, _ in _INDUSTRY_RULES}
IND_COLORS.update({'개인/주거': '#9ca3af', '기타 사업체': '#6b7280', '미분류': '#d1d5db'})

# ─────────────────────────── 유틸 ───────────────────────────────────────────

def norm_name(s):
    if not s or (isinstance(s, float) and math.isnan(s)): return ''
    s = CORP_RE.sub('', str(s)).strip()
    return re.sub(r'\s+', ' ', s)

def norm_phone(v):
    if not v or (isinstance(v, float) and math.isnan(v)): return ''
    return re.sub(r'\D', '', str(v))

def _s(v):
    return '' if not v or str(v) in ('nan', 'None', 'NaN') else str(v).strip()

def cust_key(name, phone):
    return f"{norm_name(name)}|{norm_phone(phone)}"

# ─────────────────────────── 기업명 추정 ─────────────────────────────────────

_CORP_IN_ADDR = re.compile(
    r'(?:주식회사|㈜|[\(\（]주[\)\）])\s*([가-힣A-Za-z0-9·&]{2,})|'
    r'([가-힣A-Za-z0-9·&]{2,})\s*(?:주식회사|㈜)'
)
_AFTER_FLOOR = re.compile(
    r'(?:\d+층|지하\s*\d+층?|[A-Z]\d*층?|B\d+)\s+'
    r'([가-힣A-Za-z0-9·&\(\)]{2,}(?:\s+[가-힣A-Za-z0-9·&]{2,})*)'
)
_BLDG_ANCHOR = re.compile(
    r'([가-힣A-Za-z0-9·&]{2,}(?:\s+[가-힣A-Za-z0-9·&]{2,})?)\s*'
    r'(?:빌딩|타워|센터|플라자|오피스|사옥|본사|지사|공장|연구소|캠퍼스)'
)
_FLOOR_ROOM = re.compile(
    r'\d+층|\d+호|\d+-\d+호?|\d+동\s*\d+|\d+F\b|B\d+\b|지하\d+', re.IGNORECASE
)
_APT_FILTER = re.compile(
    r'아파트|아이파크|자이|래미안|푸르지오|힐스테이트|롯데캐슬|e편한세상|더샵|'
    r'위브|두산|현대홈타운|한강|벽산|대림|코아루|현대아이|경남아너스|동부센트레빌'
)
_ORG_SUFFIX = (
    '요양원','요양센터','요양병원','병원','의원','안과','치과','한의원','약국',
    '어린이집','유치원','학원','어학원','교회','성당','사찰','절',
    '복지관','복지센터','재가센터','데이케어',
    'PC','컴퓨터','전산','정보','통신','전자','기술','테크',
    '산업','상사','물산','무역','유통','제조','공장',
    '사무소','관리사무소','경비실',
    '대리점','판매점','매장','상회','가게',
    '연구소','연구원','재단','협회','조합','노조',
)

def _is_org_name(name: str) -> bool:
    if not name or len(name) < 2: return False
    if re.match(r'^[가-힣]{2,4}$', name): return False   # 일반 한국 이름
    for suf in _ORG_SUFFIX:
        if suf in name: return True
    if re.search(r'[A-Za-z0-9]', name) and len(name) >= 3: return True
    if re.match(r'^[가-힣]{5,}$', name): return True      # 5자+ 한글 = 기관명 가능성
    return False

def infer_company(name: str, addr1, addr2='') -> str:
    """주소·이름에서 기업명 추정. 없으면 '' 반환."""
    name = str(name).strip() if name else ''
    a1 = '' if not addr1 or str(addr1) in ('nan','None','') else str(addr1).strip()
    a2 = '' if not addr2 or str(addr2) in ('nan','None','') else str(addr2).strip()
    full = f"{a1} {a2}".strip()

    # 이름이 전화번호 형식이면 기관 판별 건너뜀
    if re.match(r'^[\d\-\+\s]{7,}$', name):
        name = ''

    # 1. 수령자명 자체가 기관/법인명
    if _is_org_name(name):
        return norm_name(name) or name

    # 2. 주소 내 명시적 법인 표기 (주)기업명
    m = _CORP_IN_ADDR.search(full)
    if m:
        found = (m.group(1) or m.group(2) or '').strip()
        if len(found) >= 2: return found

    # 3. addr2가 단순 기업명 (층/호 없음, 아파트 아님)
    if a2 and not _FLOOR_ROOM.search(a2) and not _APT_FILTER.search(a2):
        cleaned = re.sub(r'^[가-힣]+(동|층|호|관)\s*', '', a2).strip()
        cleaned = re.sub(r'\s+', ' ', cleaned).strip()
        if len(cleaned) >= 2 and not re.match(r'^\d', cleaned):
            return cleaned

    # 4. 층 번호 뒤에 오는 텍스트
    m = _AFTER_FLOOR.search(full)
    if m:
        candidate = re.sub(r'[\(\)\[\]]', '', m.group(1)).strip()
        candidate = _APT_FILTER.sub('', candidate).strip()
        # 호수/방번호만 남은 경우 제외
        if len(candidate) >= 2 and not _FLOOR_ROOM.match(candidate) and not re.match(r'^\d', candidate):
            return candidate

    # 5. 빌딩/센터/사옥 앞 이름
    m = _BLDG_ANCHOR.search(full)
    if m:
        found = m.group(1).strip()
        if len(found) >= 2 and not re.match(r'^\d', found):
            return found

    return ''

# ─────────────────────────── 파일 로드 ──────────────────────────────────────

def _is_playauto(df) -> bool:
    """플레이오토 형식 감지: '쇼핑몰' 컬럼 있고 '판매사이트' 없을 때"""
    return '쇼핑몰' in df.columns and '판매사이트' not in df.columns and '묶음번호' in df.columns

def load_file(path):
    path = Path(path)
    if path.suffix.lower() == '.csv':
        # BOM 있는 UTF-8 및 CP949 모두 대응
        for enc in ('utf-8-sig', 'cp949', 'utf-8'):
            try:
                df = pd.read_csv(path, dtype=str, encoding=enc)
                break
            except (UnicodeDecodeError, Exception):
                continue
        else:
            df = pd.read_csv(path, dtype=str, encoding='cp949', errors='replace')
    else:
        df = pd.read_excel(path, dtype=str)

    if _is_playauto(df):
        df.rename(columns={k: v for k, v in COL_PLAYAUTO.items() if k in df.columns}, inplace=True)
        # ERP에 있는 컬럼 없으면 빈값으로 채움
        for col in ['cat1','cat2','cat3','sales_rep','margin','cost','vat','addr2',
                    'unit_price','sale_price','fee','fee_rate','discount','spec']:
            if col not in df.columns:
                df[col] = ''
        df['_source'] = 'playauto'
    else:
        df.rename(columns={k: v for k, v in COL_KR.items() if k in df.columns}, inplace=True)
        df['_source'] = 'erp'

    for col in ['total','qty','unit_price','sale_price','settlement','margin','cost']:
        if col in df.columns:
            df[col] = pd.to_numeric(
                df[col].astype(str).str.replace(',','').str.strip(), errors='coerce'
            ).fillna(0)
    if 'order_date' in df.columns:
        df['order_date'] = pd.to_datetime(df['order_date'], errors='coerce')
    return df

# ─────────────────────────── 데모 데이터 ────────────────────────────────────

COMPANY_PREFIXES = [
    '대한','한국','서울','부산','인천','대구','광주','한울','미래','동양',
    '삼성','현대','기아','롯데','신한','우리','하나','국민','중앙','통일',
    '영진','신흥','태양','금성','동성','성원','세진','이원','진명','한빛',
]
COMPANY_SUFFIXES = [
    '전자','통신','솔루션','시스템','IT','테크','정보','네트워크','컴퓨터',
    '기업','상사','유통','물산','개발','엔지니어링','산업','서비스','그룹',
]
CORP_FORMS = ['(주) ','','주식회사 ','']
CATEGORIES = {
    '컴퓨터/주변기기': ['데스크탑','노트북','모니터','키보드/마우스','저장장치'],
    '네트워크/서버':   ['스위치','라우터','서버','NAS','케이블'],
    '소모품/인쇄':     ['토너','잉크','용지','라벨','OA소모품'],
    '사무가구/설비':   ['의자','책상','서랍장','화이트보드','파티션'],
    '보안/솔루션':     ['CCTV','출입통제','소프트웨어','백신','방화벽'],
}
REPS = ['김영수','이지영','박민준','최서연','정호준','윤하나']

random.seed(42)

def _gen_company():
    return random.choice(CORP_FORMS) + random.choice(COMPANY_PREFIXES) + random.choice(COMPANY_SUFFIXES)

def _gen_phone():
    return f"02-{random.randint(1000,9999)}-{random.randint(1000,9999)}"

def build_demo_data():
    all_cats = list(CATEGORIES.keys())
    customers = []

    def add(n, segment, has25, has26):
        for _ in range(n):
            customers.append({
                'name': _gen_company(), 'phone': _gen_phone(),
                'rep': random.choice(REPS),
                'region': random.choice(list(REGIONS.keys())),
                'cats': random.sample(all_cats, k=random.randint(2 if has25 and has26 else 1, 5)),
                'segment': segment, 'has_25': has25, 'has_26': has26,
            })

    add(150, '충성',    True,  True)
    add(47,  '이탈',    True,  False)
    add(68,  '크로스셀', random.choice([True,False]), True)   # 크로스셀은 1-2 카테고리
    add(52,  '업셀',    random.choice([True,False]), True)

    # 크로스셀 거래처 카테고리를 1-2개로 덮어쓰기
    for c in customers:
        if c['segment'] == '크로스셀':
            c['cats'] = random.sample(all_cats, k=random.randint(1, 2))

    orders_25, orders_26 = [], []
    bundle_seq = [0]

    def make_order(c, yr):
        bundle_seq[0] += 1
        cat1 = random.choice(c['cats'])
        return {
            'cust_key': cust_key(c['name'], c['phone']),
            'name': c['name'], 'phone': c['phone'],
            'rep': c['rep'], 'region': c['region'],
            'cat1': cat1, 'cat2': random.choice(CATEGORIES[cat1]),
            'month': random.randint(1, 12),
            'bundle_no': f"{yr}-{bundle_seq[0]:06d}",
        }

    for c in customers:
        if c['has_25']:
            n = random.randint(3, 18) if c['segment'] in ('충성','이탈') else random.randint(1, 8)
            for _ in range(n):
                o = make_order(c, 25)
                o['total'] = random.randint(300_000, 6_000_000)
                orders_25.append(o)
        if c['has_26']:
            n = random.randint(8, 28) if c['segment'] == '업셀' else random.randint(2, 15)
            for _ in range(n):
                o = make_order(c, 26)
                o['total'] = random.randint(50_000, 300_000) if c['segment'] == '업셀' else random.randint(200_000, 5_000_000)
                orders_26.append(o)

    return customers, orders_25, orders_26


def analyze_demo(customers, orders_25, orders_26):
    all_cats = list(CATEGORIES.keys())

    monthly_25 = defaultdict(float)
    monthly_26 = defaultdict(float)
    for o in orders_25: monthly_25[o['month']] += o['total']
    for o in orders_26: monthly_26[o['month']] += o['total']

    regional = defaultdict(float)
    for o in orders_25 + orders_26: regional[o['region']] += o['total']
    top_regions = sorted(regional.items(), key=lambda x: -x[1])[:10]

    cust_total_25 = defaultdict(float)
    cust_orders_25 = defaultdict(set)
    cust_cats_25   = defaultdict(set)
    for o in orders_25:
        cust_total_25[o['cust_key']] += o['total']
        cust_orders_25[o['cust_key']].add(o['bundle_no'])
        cust_cats_25[o['cust_key']].add(o['cat1'])

    cust_total_26 = defaultdict(float)
    cust_orders_26 = defaultdict(set)
    for o in orders_26:
        cust_total_26[o['cust_key']] += o['total']
        cust_orders_26[o['cust_key']].add(o['bundle_no'])

    # 이탈 TOP
    churn_list = sorted(
        [c for c in customers if c['segment'] == '이탈'],
        key=lambda c: -cust_total_25[cust_key(c['name'], c['phone'])]
    )
    churn_top = []
    for c in churn_list[:20]:
        ck = cust_key(c['name'], c['phone'])
        churn_top.append({
            'name': c['name'], 'rep': c['rep'],
            'region': REGIONS.get(c['region'], c['region']),
            'total_25': round(cust_total_25[ck]),
            'order_count_25': len(cust_orders_25[ck]),
            'cats': list(cust_cats_25[ck]),
            'last_order': f"25년 {random.randint(7,12)}월",
        })

    # 크로스셀 매트릭스
    cat_opp = defaultdict(lambda: defaultdict(int))
    for c in customers:
        if c['segment'] == '크로스셀':
            has = set(c['cats'])
            for m in [x for x in all_cats if x not in has]:
                for h in c['cats']:
                    cat_opp[h][m] += 1
    crosssell_matrix = [
        {'current_cat': h, 'opportunities': [{'cat':k,'count':v} for k,v in sorted(opps.items(),key=lambda x:-x[1])[:3]]}
        for h, opps in cat_opp.items()
    ]

    # 업셀
    global_avg = 800_000
    upsell_top = []
    for c in customers:
        if c['segment'] != '업셀': continue
        ck = cust_key(c['name'], c['phone'])
        freq = len(cust_orders_26[ck])
        avg = cust_total_26[ck] / freq if freq else 0
        upsell_top.append({
            'name': c['name'], 'rep': c['rep'],
            'order_count': freq,
            'avg_order': round(avg),
            'global_avg': global_avg,
            'gap_pct': round((global_avg - avg) / global_avg * 100) if avg else 100,
        })
    upsell_top.sort(key=lambda x: -x['order_count'])

    # 영업담당
    rep_d = defaultdict(lambda: {'loyal':0,'churn':0,'crosssell':0,'upsell':0,'revenue':0.0})
    for c in customers:
        ck = cust_key(c['name'], c['phone'])
        r = c['rep']; s = c['segment']
        rep_d[r]['revenue'] += cust_total_26.get(ck, cust_total_25.get(ck, 0))
        rep_d[r][{'충성':'loyal','이탈':'churn','크로스셀':'crosssell','업셀':'upsell'}.get(s,'loyal')] += 1

    seg_counts = defaultdict(int)
    for c in customers: seg_counts[c['segment']] += 1

    return {
        'kpi': {'total': len(customers), 'loyal': seg_counts['충성'],
                'churn': seg_counts['이탈'], 'crosssell': seg_counts['크로스셀'],
                'upsell': seg_counts['업셀']},
        'monthly_25': {str(m): round(monthly_25.get(m, 0)) for m in range(1, 13)},
        'monthly_26': {str(m): round(monthly_26.get(m, 0)) for m in range(1, 13)},
        'regional': [{'region': REGIONS.get(r, r), 'total': round(t)} for r, t in top_regions],
        'churn_top': churn_top,
        'crosssell_matrix': crosssell_matrix,
        'upsell_top': upsell_top[:20],
        'rep_stats': sorted([{'name':k,**v} for k,v in rep_d.items()], key=lambda x:-x['revenue']),
        'all_cats': all_cats,
        'mode': 'demo',
    }

# ─────────────────────────── 실 데이터 분석 ─────────────────────────────────

def analyze_real(dfs: dict):
    """dfs: {연도(int): DataFrame}  예) {2024: df24, 2025: df25, 2026: df26}"""
    years = sorted(dfs.keys())
    cur_yr = years[-1]          # 가장 최근 연도 = 현재 기준
    hist_yrs = years[:-1]       # 이전 연도들
    df_cur = dfs[cur_yr]

    # cust_key 생성 — 주문자명 우선, 없으면 수령자명 fallback
    # 전화: 수령자 핸드폰 → 주문자 핸드폰 → 수령자 전화 순
    def assign_key(df):
        phone   = df.get('recipient_phone',  pd.Series(dtype=str)).fillna('')
        mobile  = df.get('recipient_mobile', pd.Series(dtype=str)).fillna('')
        o_mobile= df.get('orderer_mobile',   pd.Series(dtype=str)).fillna('')
        best_phone = mobile.where(mobile.str.strip() != '', o_mobile)
        best_phone = best_phone.where(best_phone.str.strip() != '', phone)
        orderer   = df.get('orderer_name',   pd.Series(dtype=str)).fillna('')
        recipient = df.get('recipient_name', pd.Series(dtype=str)).fillna('')
        primary_name = orderer.where(orderer.str.strip() != '', recipient)
        df['cust_key'] = [
            cust_key(n, p)
            for n, p in zip(primary_name, best_phone)
        ]
        df['display_name'] = primary_name
    for df in dfs.values():
        assign_key(df)
        df['channel'] = df['site'].apply(classify_channel) if 'site' in df.columns else 'B2B'

    all_cats = sorted(set(
        v for df in dfs.values()
        for v in df.get('cat1', pd.Series(dtype=str)).dropna().unique()
    ))

    keys_cur  = set(df_cur['cust_key'])
    keys_hist = set(k for yr in hist_yrs for k in dfs[yr]['cust_key'])

    # 거래처 기본 정보 (과거→현재 순으로 덮어쓰기)
    cust_info = {}
    for yr in years:
        for _, row in dfs[yr].drop_duplicates('cust_key').iterrows():
            ck = row['cust_key']
            if ck not in cust_info:
                # 주문자명 우선, 없으면 수령자명
                raw_name = _s(row.get('display_name', '')) or _s(row.get('orderer_name', '')) or _s(row.get('recipient_name', ''))
                company_hint = infer_company(
                    raw_name,
                    row.get('addr1', ''),
                    row.get('addr2', ''),
                )
                cust_info[ck] = {
                    'name': norm_name(raw_name),
                    'rep':  str(row.get('sales_rep', '')),
                    'region': str(row.get('postal_code', ''))[:3],
                    'addr': _s(row.get('addr1', '')),
                    'phone':  _s(row.get('recipient_phone', '')),
                    'mobile': _s(row.get('recipient_mobile', '')),
                    'cats': set(),
                    'company_hint': company_hint,
                }

    for df in dfs.values():
        if 'cat1' in df.columns:
            for ck, grp in df.groupby('cust_key'):
                if ck in cust_info:
                    cust_info[ck]['cats'].update(grp['cat1'].dropna())

    # 현재연도 기준 집계
    order_col_cur = 'bundle_no' if 'bundle_no' in df_cur.columns else 'cust_key'
    cust_total_cur = df_cur.groupby('cust_key')['total'].sum() if 'total' in df_cur.columns else pd.Series(dtype=float)
    cust_freq_cur  = df_cur.groupby('cust_key')[order_col_cur].nunique()
    avg_order_cur  = (cust_total_cur / cust_freq_cur.clip(lower=1)).fillna(0)
    global_avg     = float(avg_order_cur.mean()) if len(avg_order_cur) else 1.0
    avg_freq       = float(cust_freq_cur.mean()) if len(cust_freq_cur) else 1.0

    # 직전연도 집계 (이탈 우선순위용)
    prev_yr = hist_yrs[-1] if hist_yrs else cur_yr
    df_prev = dfs[prev_yr]
    order_col_prev = 'bundle_no' if 'bundle_no' in df_prev.columns else 'cust_key'
    cust_total_prev = df_prev.groupby('cust_key')['total'].sum() if 'total' in df_prev.columns else pd.Series(dtype=float)
    cust_orders_prev_cnt = df_prev.groupby('cust_key')[order_col_prev].nunique()

    def get_segment(ck):
        in_cur  = ck in keys_cur
        in_hist = ck in keys_hist
        cats = cust_info.get(ck, {}).get('cats', set())
        freq = float(cust_freq_cur.get(ck, 0))
        avg  = float(avg_order_cur.get(ck, 0))
        if in_hist and not in_cur: return '이탈'
        if len(cats) <= 2:         return '크로스셀'
        if freq >= avg_freq and avg <= global_avg * 0.7: return '업셀'
        return '충성'

    for ck in cust_info:
        cust_info[ck]['segment'] = get_segment(ck)
        cust_info[ck]['cats'] = list(cust_info[ck]['cats'])

    # 거래처별 주력 채널 (주문 수 기준 최빈값)
    cust_ch_counts = defaultdict(lambda: defaultdict(int))
    for df in dfs.values():
        for ck_val, grp in df.groupby('cust_key'):
            for ch, cnt in grp['channel'].value_counts().items():
                cust_ch_counts[ck_val][ch] += int(cnt)
    for ck in cust_info:
        counts = cust_ch_counts.get(ck, {})
        cust_info[ck]['channel'] = max(counts, key=counts.get) if counts else '딜러/파트너'

    seg_counts = Counter(c['segment'] for c in cust_info.values())

    # 업종 분류 (churn_top 빌드 전에 먼저 계산)
    for ck, c in cust_info.items():
        c['industry'] = classify_industry(c.get('addr', ''), c.get('company_hint', ''))

    # 연도별 월별 매출
    def monthly_agg(df):
        if 'order_date' not in df.columns or 'total' not in df.columns:
            return {}
        s = df.dropna(subset=['order_date']).groupby(df['order_date'].dt.month)['total'].sum()
        return {str(int(k)): round(float(v)) for k, v in s.items()}

    monthly_by_year = {str(yr): monthly_agg(df) for yr, df in dfs.items()}

    # 지역별
    all_df = pd.concat(list(dfs.values()))
    if 'postal_code' in all_df.columns and 'total' in all_df.columns:
        all_df['_reg'] = all_df['postal_code'].astype(str).str[:3]
        reg_s = all_df.groupby('_reg')['total'].sum().sort_values(ascending=False).head(10)
        regional = [{'region': REGIONS.get(r, r+'지역'), 'total': round(float(t))} for r, t in reg_s.items()]
    else:
        regional = []

    # 이탈 TOP (직전연도 매출 기준 정렬)
    churn_sorted = sorted(
        [(ck, c) for ck, c in cust_info.items() if c['segment'] == '이탈'],
        key=lambda x: -float(cust_total_prev.get(x[0], 0))
    )
    def _churn_entry(ck, c):
        return {
            'name': c['name'],
            'company_hint': c.get('company_hint', ''),
            'display_name': c['name'] + (f" ({c['company_hint']})" if c.get('company_hint') else ''),
            'rep': c['rep'],
            'region': REGIONS.get(c['region'], c['region']+'지역'),
            'addr': c.get('addr', ''),
            'total_prev': round(float(cust_total_prev.get(ck, 0))),
            'order_count_prev': int(cust_orders_prev_cnt.get(ck, 0)),
            'cats': c['cats'][:4],
            'last_order': f'{str(prev_yr)[2:]}년',
            'prev_yr_label': f'{str(prev_yr)[2:]}년',
            'channel': c.get('channel', '딜러/파트너'),
            'industry': c.get('industry', '미분류'),
            'phone':  c.get('phone', ''),
            'mobile': c.get('mobile', ''),
        }

    # 엑셀 export용 전체 이탈 목록 (메모리만 보관, JSON 미포함)
    churn_all_export = [_churn_entry(ck, c) for ck, c in churn_sorted]

    # 대시보드용: 업종별 상위 100개씩 + 전체 상위 100개 (JSON 경량화)
    _ind_buckets = defaultdict(list)
    for ck, c in churn_sorted:
        _ind_buckets[c.get('industry', '미분류')].append((ck, c))
    churn_top_set = {}  # ck → entry (중복 방지)
    for entries in _ind_buckets.values():
        for ck, c in entries[:100]:
            if ck not in churn_top_set:
                churn_top_set[ck] = _churn_entry(ck, c)
    # 전체 상위 100개도 포함
    for ck, c in churn_sorted[:100]:
        if ck not in churn_top_set:
            churn_top_set[ck] = _churn_entry(ck, c)
    churn_top = sorted(churn_top_set.values(), key=lambda x: -x['total_prev'])

    # 업종별 전체 카운트 (필터 뱃지용)
    churn_industry_counts = {ind: len(lst) for ind, lst in _ind_buckets.items()}
    churn_total_count = len(churn_sorted)

    # 크로스셀 매트릭스 (Market Basket Analysis — affinity lift 기반)
    cat_buyers = defaultdict(set)
    for ck, c in cust_info.items():
        for cat in c['cats']:
            cat_buyers[cat].add(ck)
    cat_co = defaultdict(lambda: defaultdict(int))
    for ck, c in cust_info.items():
        cats = list(set(c['cats']))
        for i in range(len(cats)):
            for j in range(i + 1, len(cats)):
                cat_co[cats[i]][cats[j]] += 1
                cat_co[cats[j]][cats[i]] += 1
    crosssell_matrix = []
    for cat_a in sorted(cat_buyers.keys(), key=lambda x: -len(cat_buyers[x])):
        n_a = len(cat_buyers[cat_a])
        if n_a == 0:
            continue
        recs = []
        for cat_b, co_count in cat_co[cat_a].items():
            n_b = len(cat_buyers[cat_b])
            affinity = round(co_count / min(n_a, n_b) * 100, 1)
            recs.append({'cat': cat_b, 'co_count': co_count, 'affinity': affinity})
        recs.sort(key=lambda x: -x['affinity'])
        crosssell_matrix.append({
            'current_cat': cat_a,
            'buyer_count': n_a,
            'opportunities': recs[:3],
        })

    # 업셀
    upsell_top = []
    for ck, c in cust_info.items():
        if c['segment'] != '업셀': continue
        avg  = float(avg_order_cur.get(ck, 0))
        freq = int(cust_freq_cur.get(ck, 0))
        upsell_top.append({
            'name': c['name'], 'rep': c['rep'],
            'order_count': freq, 'avg_order': round(avg),
            'global_avg': round(global_avg),
            'gap_pct': round((global_avg - avg) / global_avg * 100) if avg and global_avg else 0,
            'channel': c.get('channel', '딜러/파트너'),
        })
    upsell_top.sort(key=lambda x: -x['order_count'])

    # 영업담당
    rep_d = defaultdict(lambda: {'loyal':0,'churn':0,'crosssell':0,'upsell':0,'revenue':0.0})
    for ck, c in cust_info.items():
        r = c.get('rep',''); s = c['segment']
        rep_d[r]['revenue'] += float(cust_total_cur.get(ck, cust_total_prev.get(ck, 0)))
        rep_d[r][{'충성':'loyal','이탈':'churn','크로스셀':'crosssell','업셀':'upsell'}.get(s,'loyal')] += 1

    # ── 채널별 KPI ────────────────────────────────────────────────────────
    channel_kpi = {}
    for ch in CHANNELS:
        ck_ch = [ck for ck, c in cust_info.items() if c.get('channel') == ch]
        ch_segs = Counter(cust_info[ck]['segment'] for ck in ck_ch)
        ch_rev  = sum(float(cust_total_cur.get(ck, cust_total_prev.get(ck, 0))) for ck in ck_ch)
        channel_kpi[ch] = {
            'total': len(ck_ch), 'revenue': round(ch_rev),
            'loyal': ch_segs.get('충성', 0), 'churn': ch_segs.get('이탈', 0),
            'crosssell': ch_segs.get('크로스셀', 0), 'upsell': ch_segs.get('업셀', 0),
        }

    # ── 채널별 월별 매출 ───────────────────────────────────────────────────
    monthly_by_channel = {}
    for ch in CHANNELS:
        ch_monthly = {}
        for yr, df in dfs.items():
            df_ch = df[df['channel'] == ch] if 'channel' in df.columns else df.iloc[:0]
            if 'order_date' not in df_ch.columns or 'total' not in df_ch.columns:
                ch_monthly[str(yr)] = {}
            else:
                s = df_ch.dropna(subset=['order_date']).groupby(df_ch['order_date'].dt.month)['total'].sum()
                ch_monthly[str(yr)] = {str(int(k)): round(float(v)) for k, v in s.items()}
        monthly_by_channel[ch] = ch_monthly

    # ── RFM 분석 ──────────────────────────────────────────────────────────
    rfm_data = _compute_rfm(df_cur, cust_info, keys_cur, cust_freq_cur, cust_total_cur)

    # ── ABC 분석 ──────────────────────────────────────────────────────────
    abc_data = _compute_abc(cust_info, cust_total_cur)

    # ── 품목 구매패턴 ──────────────────────────────────────────────────────
    pattern_data = _compute_patterns(dfs, cust_info, all_cats)

    # ── 엑셀 export용 데이터 ──────────────────────────────────────────────
    cs_cat_recs = {r['current_cat']: r['opportunities'] for r in crosssell_matrix}
    crosssell_customers = []
    for ck, c in cust_info.items():
        if c['segment'] != '크로스셀': continue
        rev = float(cust_total_cur.get(ck, cust_total_prev.get(ck, 0)))
        recs = []
        for cat in c['cats']:
            for opp in cs_cat_recs.get(cat, [])[:2]:
                if opp['cat'] not in c['cats'] and opp['cat'] not in recs:
                    recs.append(opp['cat'])
        crosssell_customers.append({
            'name': c['name'], 'rep': c['rep'],
            'channel': c.get('channel', '딜러/파트너'),
            'region': REGIONS.get(c['region'], c['region']+'지역'),
            'company_hint': c.get('company_hint', ''),
            'revenue': round(rev),
            'cats': c['cats'],
            'recommended_cats': recs[:3],
        })
    crosssell_customers.sort(key=lambda x: -x['revenue'])
    crosssell_customers = crosssell_customers[:100]

    vip_customers = []
    for ck, c in cust_info.items():
        if c['segment'] != '충성': continue
        rev = float(cust_total_cur.get(ck, 0))
        freq = int(cust_freq_cur.get(ck, 0))
        vip_customers.append({
            'name': c['name'], 'rep': c['rep'],
            'channel': c.get('channel', '딜러/파트너'),
            'region': REGIONS.get(c['region'], c['region']+'지역'),
            'company_hint': c.get('company_hint', ''),
            'revenue': round(rev),
            'order_count': freq,
            'cats': c['cats'],
        })
    vip_customers.sort(key=lambda x: -x['revenue'])
    vip_customers = vip_customers[:100]

    # ── 구매빈도 상위 / 매출 상위 (전체 세그먼트) ──────────────────────────────
    # 고객별 연도별 주문이력 집계 (최근 30건 제한)
    _order_hist: dict = defaultdict(list)
    for yr, df in sorted(dfs.items()):
        if 'order_date' not in df.columns: continue
        _cols = ['cust_key','order_date','total','cat1','cat2','site','channel','bundle_no']
        _sub  = df[[c for c in _cols if c in df.columns]].copy()
        _sub  = _sub.dropna(subset=['order_date'])
        for _, r in _sub.iterrows():
            _order_hist[r['cust_key']].append({
                'date':    r['order_date'].strftime('%Y-%m-%d') if pd.notna(r['order_date']) else '',
                'yr':      int(yr),
                'cat1':    str(r.get('cat1', '') or ''),
                'cat2':    str(r.get('cat2', '') or ''),
                'total':   round(float(r.get('total', 0) or 0)),
                'site':    _s(r.get('site', '')),
                'channel': str(r.get('channel', '') or classify_channel(r.get('site', ''))),
            })
    # 날짜 내림차순 정렬 후 최근 30건만 보관
    for ck in _order_hist:
        _order_hist[ck].sort(key=lambda x: x['date'], reverse=True)
        _order_hist[ck] = _order_hist[ck][:30]

    def _build_entry(ck, c):
        rev  = float(cust_total_cur.get(ck, cust_total_prev.get(ck, 0)))
        freq = int(cust_freq_cur.get(ck, 0))
        hist = _order_hist.get(ck, [])
        first_date = hist[-1]['date'] if hist else ''
        last_date  = hist[0]['date']  if hist else ''
        return {
            'name': c['name'],
            'company_hint': c.get('company_hint', ''),
            'rep': c['rep'],
            'channel': c.get('channel', '딜러/파트너'),
            'region': REGIONS.get(c['region'], c['region']+'지역'),
            'addr': c.get('addr', ''),
            'mobile': c.get('mobile', ''),
            'segment': c['segment'],
            'industry': c.get('industry', '미분류'),
            'revenue': round(rev),
            'order_count': freq,
            'cats': c['cats'][:3],
            'first_order': first_date,
            'last_order':  last_date,
            'orders': hist,
        }

    all_ranked     = [_build_entry(ck, c) for ck, c in cust_info.items()]
    top_by_freq    = sorted(all_ranked, key=lambda x: (-x['order_count'], -x['revenue']))[:50]
    top_by_revenue = sorted(all_ranked, key=lambda x: (-x['revenue'], -x['order_count']))[:50]

    # ── 업종별 분석 (classify_industry는 위에서 이미 실행됨) ─────────────────
    ind_rev = defaultdict(float)
    ind_cnt = defaultdict(int)
    for ck, c in cust_info.items():
        ind = c['industry']
        ind_cnt[ind] += 1
        ind_rev[ind] += float(cust_total_cur.get(ck, cust_total_prev.get(ck, 0)))

    industry_data = sorted([
        {'name': k, 'count': ind_cnt[k], 'revenue': round(ind_rev[k]), 'color': IND_COLORS.get(k, '#888')}
        for k in ind_cnt
    ], key=lambda x: -x['revenue'])

    return {
        'kpi': {'total': len(cust_info), 'loyal': seg_counts['충성'],
                'churn': seg_counts['이탈'], 'crosssell': seg_counts['크로스셀'],
                'upsell': seg_counts['업셀']},
        'monthly_by_year': monthly_by_year,
        'years': [str(y) for y in years],
        'cur_yr': str(cur_yr),
        'prev_yr': str(prev_yr),
        'regional': regional,
        'churn_top': churn_top,
        'churn_industry_counts': churn_industry_counts,
        'churn_total_count': churn_total_count,
        '_churn_all_export': churn_all_export,
        'crosssell_matrix': crosssell_matrix,
        'upsell_top': upsell_top[:20],
        'rep_stats': sorted([{'name':k,**v} for k,v in rep_d.items() if k], key=lambda x:-x['revenue']),
        'all_cats': all_cats,
        'channel_kpi': channel_kpi,
        'monthly_by_channel': monthly_by_channel,
        'channel_groups': {k: list(v) for k, v in CHANNEL_GROUPS.items()},
        'crosssell_customers': crosssell_customers,
        'vip_customers': vip_customers,
        'top_by_freq': top_by_freq,
        'top_by_revenue': top_by_revenue,
        'mode': 'real',
        'rfm': rfm_data,
        'abc': abc_data,
        'purchase_pattern': pattern_data,
        'industry': industry_data,
    }

# ─────────────────────────── RFM / ABC / 패턴 보조 함수 ─────────────────────

def _export_customer2(data, out_path):
    """거래처_2.csv: 기업 추정명 포함 전체 거래처 목록 (Excel UTF-8-BOM)"""
    try:
        rows = []
        for c in data.get('churn_top', []):
            rows.append({
                '수령자명': c['name'],
                '추정기업명': c.get('company_hint', ''),
                '표시명': c.get('display_name', c['name']),
                '영업담당': c['rep'],
                '지역': c['region'],
                '세그먼트': '이탈',
                f"{c.get('prev_yr_label','전년')} 매출": c['total_prev'],
                '주문건수': c['order_count_prev'],
                '구매카테고리': ', '.join(c.get('cats', [])),
            })
        # RFM top도 포함
        for r in data.get('rfm', {}).get('top', []):
            rows.append({
                '수령자명': r.get('name',''),
                '추정기업명': '',
                '표시명': r.get('name',''),
                '영업담당': r.get('rep',''),
                '지역': '',
                '세그먼트': r.get('segment',''),
                '최근년 매출': r.get('monetary', 0),
                '주문건수': r.get('frequency', 0),
                '구매카테고리': '',
            })
        if rows:
            pd.DataFrame(rows).to_csv(out_path, index=False, encoding='utf-8-sig')
            print(f"[내보내기] 거래처_2.csv: {out_path}")
    except Exception as e:
        print(f"[주의] 거래처_2.csv 생성 실패: {e}")


def _export_action_list(data, out_path):
    """영업 액션 리스트 — 4시트 Excel"""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        print("[주의] openpyxl 미설치. 액션 리스트 생성 건너뜀.")
        return
    try:
        wb = openpyxl.Workbook()
        H_FILL = PatternFill(start_color='1D4ED8', end_color='1D4ED8', fill_type='solid')
        H_FONT = Font(color='FFFFFF', bold=True, size=11)
        H_ALIGN = Alignment(horizontal='center', vertical='center')

        def style_header(ws):
            ws.row_dimensions[1].height = 22
            for cell in ws[1]:
                cell.fill = H_FILL
                cell.font = H_FONT
                cell.alignment = H_ALIGN
            ws.freeze_panes = 'A2'

        def fmt_rev(v):
            return f"{round(v/10000):,}만원" if v >= 10000 else str(round(v))

        prev_yr = data.get('prev_yr', '25')
        cur_yr  = data.get('cur_yr',  '26')

        # ── 시트1: 재활성화 우선순위 전체 (업종 포함) ────────────────────────
        churn_export = data.get('_churn_all_export') or data.get('churn_top', [])
        ws1 = wb.active
        ws1.title = '재활성화_우선순위'
        ws1.append(['순위', '거래처명', '추정기업명', '업종', '채널', '영업담당', '지역',
                    '주소', '핸드폰',
                    f'{prev_yr}년 매출', '주문건수', '구매카테고리', '권장 액션'])
        for i, c in enumerate(churn_export, 1):
            ws1.append([
                i, c['name'], c.get('company_hint', ''),
                c.get('industry', '미분류'),
                c.get('channel', ''), c['rep'], c['region'],
                c.get('addr', ''), c.get('mobile', ''),
                fmt_rev(c.get('total_prev', 0)), c.get('order_count_prev', 0),
                ', '.join(c.get('cats', [])), '전화/이메일 재활성화 제안',
            ])
        style_header(ws1)
        ws1.column_dimensions['B'].width = 20
        ws1.column_dimensions['H'].width = 14

        # ── 시트1-B: 업종별 재활성화 요약 ────────────────────────────────────
        ws1b = wb.create_sheet('업종별_재활성화')
        ws1b.append(['업종', '이탈 거래처수', f'{prev_yr}년 총매출', '평균 매출', '구성 채널(최빈)'])
        ind_grp = defaultdict(list)
        for c in churn_export:
            ind_grp[c.get('industry', '미분류')].append(c)
        for ind, lst in sorted(ind_grp.items(), key=lambda x: -sum(c['total_prev'] for c in x[1])):
            tot = sum(c['total_prev'] for c in lst)
            avg = round(tot / len(lst)) if lst else 0
            top_ch = Counter(c.get('channel','') for c in lst).most_common(1)
            ws1b.append([ind, len(lst), fmt_rev(tot), fmt_rev(avg), top_ch[0][0] if top_ch else ''])
        style_header(ws1b)
        ws1b.column_dimensions['A'].width = 16
        ws1b.column_dimensions['C'].width = 16

        # ── 시트2: 크로스셀 기회 ──────────────────────────────────────────
        ws2 = wb.create_sheet('크로스셀_기회')
        ws2.append(['거래처명', '추정기업명', '채널', '영업담당', '지역',
                    '매출', '현재 구매 카테고리', '추천 카테고리', '권장 액션'])
        for c in data.get('crosssell_customers', []):
            ws2.append([
                c['name'], c.get('company_hint', ''),
                c.get('channel', ''), c['rep'], c['region'],
                fmt_rev(c['revenue']),
                ', '.join(c.get('cats', [])),
                ', '.join(c.get('recommended_cats', [])) or '-',
                '추천 카테고리 상품 제안',
            ])
        style_header(ws2)
        ws2.column_dimensions['A'].width = 20

        # ── 시트3: 업셀 후보 ──────────────────────────────────────────────
        ws3 = wb.create_sheet('업셀_후보')
        ws3.append(['거래처명', '채널', '영업담당', '주문빈도', '평균 객단가',
                    '전체 평균 객단가', '차이율', '권장 액션'])
        for c in data.get('upsell_top', []):
            ws3.append([
                c['name'], c.get('channel', ''), c['rep'],
                c['order_count'], fmt_rev(c['avg_order']),
                fmt_rev(c['global_avg']),
                f"{c['gap_pct']}% 낮음",
                '고단가 제품 라인업 제안',
            ])
        style_header(ws3)
        ws3.column_dimensions['A'].width = 20

        # ── 시트4: VIP 관리 ───────────────────────────────────────────────
        ws4 = wb.create_sheet('VIP_관리')
        ws4.append(['거래처명', '추정기업명', '채널', '영업담당', '지역',
                    f'{cur_yr}년 매출', '주문건수', '구매카테고리', '권장 액션'])
        for c in data.get('vip_customers', []):
            ws4.append([
                c['name'], c.get('company_hint', ''),
                c.get('channel', ''), c['rep'], c['region'],
                fmt_rev(c['revenue']), c['order_count'],
                ', '.join(c.get('cats', [])),
                '관계 유지 및 연간 계약 검토',
            ])
        style_header(ws4)
        ws4.column_dimensions['A'].width = 20

        wb.save(out_path)
        print(f"[내보내기] 액션 리스트: {out_path}")
    except Exception as e:
        print(f"[주의] 액션 리스트 생성 실패: {e}")


def _compute_rfm(df_cur, cust_info, keys_cur, cust_freq_cur, cust_total_cur):
    try:
        ref_date = df_cur['order_date'].max() if 'order_date' in df_cur.columns else pd.Timestamp.now()
        if pd.isna(ref_date): ref_date = pd.Timestamp.now()

        last_order = df_cur.groupby('cust_key')['order_date'].max() if 'order_date' in df_cur.columns else pd.Series(dtype='datetime64[ns]')
        recency_days = ((ref_date - last_order).dt.days).reindex(list(keys_cur)).fillna(999)

        rfm = pd.DataFrame({
            'cust_key': list(keys_cur),
            'R_raw': [float(recency_days.get(k, 999)) for k in keys_cur],
            'F_raw': [int(cust_freq_cur.get(k, 0)) for k in keys_cur],
            'M_raw': [float(cust_total_cur.get(k, 0)) for k in keys_cur],
        })

        def qscore(s, ascending):
            labels = [5,4,3,2,1] if ascending else [1,2,3,4,5]
            try:
                return pd.qcut(s.rank(method='first'), 5, labels=labels).astype(int)
            except Exception:
                return pd.Series([3]*len(s), index=s.index)

        rfm['R'] = qscore(rfm['R_raw'], ascending=True)   # 낮을수록 최근 → 5점
        rfm['F'] = qscore(rfm['F_raw'], ascending=False)
        rfm['M'] = qscore(rfm['M_raw'], ascending=False)
        rfm['score'] = rfm['R'] + rfm['F'] + rfm['M']

        def seg(r):
            R, F, M = r['R'], r['F'], r['M']
            if R >= 4 and F >= 4 and M >= 4: return '최우수'
            if F >= 4 and M >= 4:            return '충성고객'
            if R >= 4 and F <= 2:            return '신규/잠재'
            if R <= 2 and (F >= 4 or M >= 4): return '이탈위험'
            if R <= 2 and F <= 2:            return '휴면'
            return '관찰필요'

        rfm['segment'] = rfm.apply(seg, axis=1)
        name_map = {ck: c['name'] for ck, c in cust_info.items()}
        rep_map  = {ck: c['rep']  for ck, c in cust_info.items()}
        rfm['name'] = rfm['cust_key'].map(name_map)
        rfm['rep']  = rfm['cust_key'].map(rep_map)

        seg_counts = rfm['segment'].value_counts().to_dict()
        top20 = rfm.nlargest(20, 'score')
        top_list = [{
            'name': r['name'], 'rep': r['rep'],
            'R': int(r['R']), 'F': int(r['F']), 'M': int(r['M']),
            'score': int(r['score']),
            'segment': r['segment'],
            'recency': int(r['R_raw']),
            'frequency': int(r['F_raw']),
            'monetary': round(r['M_raw']),
        } for _, r in top20.iterrows()]

        return {
            'seg_counts': seg_counts,
            'top': top_list,
            'ref_date': ref_date.strftime('%Y-%m-%d'),
            'total': len(rfm),
        }
    except Exception as e:
        return {'seg_counts': {}, 'top': [], 'ref_date': '', 'total': 0, 'error': str(e)}


def _compute_abc(cust_info, cust_total_cur):
    try:
        records = [{'cust_key': k, 'revenue': float(cust_total_cur.get(k, 0))}
                   for k in cust_info]
        abc_df = pd.DataFrame(records).sort_values('revenue', ascending=False).reset_index(drop=True)
        total_rev = abc_df['revenue'].sum()
        if total_rev == 0:
            return {'summary': {}, 'top_a': [], 'pareto': {}}

        abc_df['cum_pct'] = abc_df['revenue'].cumsum() / total_rev * 100
        abc_df['grade'] = 'C'
        abc_df.loc[abc_df['cum_pct'] <= 80, 'grade'] = 'A'
        abc_df.loc[(abc_df['cum_pct'] > 80) & (abc_df['cum_pct'] <= 95), 'grade'] = 'B'

        name_map = {ck: c['name'] for ck, c in cust_info.items()}
        rep_map  = {ck: c['rep']  for ck, c in cust_info.items()}
        abc_df['name'] = abc_df['cust_key'].map(name_map)
        abc_df['rep']  = abc_df['cust_key'].map(rep_map)

        summary = {}
        for g in ['A','B','C']:
            grp = abc_df[abc_df['grade'] == g]
            summary[g] = {
                'count': int(len(grp)),
                'revenue': round(float(grp['revenue'].sum())),
                'pct_cust': round(len(grp)/len(abc_df)*100, 1),
                'pct_rev':  round(float(grp['revenue'].sum())/total_rev*100, 1),
            }

        top_a = abc_df[abc_df['grade']=='A'].head(20)[['name','rep','revenue']].to_dict('records')
        top_a = [{'name': r['name'], 'rep': r['rep'], 'revenue': round(float(r['revenue']))} for r in top_a]

        n = min(50, len(abc_df))
        pareto = {
            'labels':   abc_df.head(n)['name'].tolist(),
            'revenue':  [round(float(v)/1e6, 1) for v in abc_df.head(n)['revenue']],
            'cum_pct':  [round(v, 1) for v in abc_df.head(n)['cum_pct']],
            'grades':   abc_df.head(n)['grade'].tolist(),
        }

        return {'summary': summary, 'top_a': top_a, 'pareto': pareto}
    except Exception as e:
        return {'summary': {}, 'top_a': [], 'pareto': {}, 'error': str(e)}


def _compute_patterns(dfs, cust_info, all_cats):
    try:
        # 카테고리 공동구매 친화도
        cat_co   = defaultdict(int)
        cat_solo = defaultdict(int)
        for c in cust_info.values():
            cats = list(set(c['cats']))
            for cat in cats: cat_solo[cat] += 1
            for i in range(len(cats)):
                for j in range(i+1, len(cats)):
                    cat_co[tuple(sorted([cats[i], cats[j]]))] += 1

        affinity = []
        for (a, b), co in cat_co.items():
            base = min(cat_solo.get(a, 1), cat_solo.get(b, 1))
            affinity.append({'cat_a': a, 'cat_b': b, 'co_count': co,
                             'affinity': round(co/base*100, 1)})
        affinity.sort(key=lambda x: -x['co_count'])

        # 카테고리별 중분류 TOP5
        all_df = pd.concat(list(dfs.values()))
        cat2_pattern = {}
        if 'cat1' in all_df.columns and 'cat2' in all_df.columns and 'total' in all_df.columns:
            for cat1_val, grp in all_df.groupby('cat1'):
                top5 = grp.groupby('cat2')['total'].sum().sort_values(ascending=False).head(5)
                cat2_pattern[str(cat1_val)] = [
                    {'name': str(k), 'total': round(float(v))} for k, v in top5.items()
                ]

        # 연도별 카테고리 매출 비교 (신규 품목 발굴용)
        yr_cat_rev = {}
        for yr, df in dfs.items():
            if 'cat1' in df.columns and 'total' in df.columns:
                s = df.groupby('cat1')['total'].sum().sort_values(ascending=False).head(10)
                yr_cat_rev[str(yr)] = [{'cat': str(k), 'total': round(float(v))} for k, v in s.items()]

        return {
            'affinity': affinity[:20],
            'cat2_pattern': cat2_pattern,
            'yr_cat_rev': yr_cat_rev,
        }
    except Exception as e:
        return {'affinity': [], 'cat2_pattern': {}, 'yr_cat_rev': {}, 'error': str(e)}

# ─────────────────────────── HTML 템플릿 ────────────────────────────────────

HTML = r"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>매출처 마케팅 대시보드</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<script src="https://cdn.sheetjs.com/xlsx-0.20.3/package/dist/xlsx.full.min.js"></script>
<style>
:root{
  --bg:#f5f6f8;--card:#fff;--primary:#2563eb;--danger:#dc2626;
  --warning:#d97706;--success:#16a34a;--purple:#7c3aed;
  --text:#111827;--muted:#6b7280;--border:#e5e7eb;
}
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Apple SD Gothic Neo',sans-serif;background:var(--bg);color:var(--text);font-size:14px}
.header{background:var(--card);border-bottom:1px solid var(--border);padding:16px 24px;display:flex;align-items:center;justify-content:space-between}
.header h1{font-size:18px;font-weight:700}
.header .meta{color:var(--muted);font-size:12px}
.demo-banner{background:#fef3c7;border-bottom:1px solid #fde68a;padding:8px 24px;font-size:12px;color:#92400e;display:none}
/* 채널 탭 */
.ch-tabs{background:var(--card);border-bottom:2px solid var(--border);padding:0 24px;display:flex;gap:0;flex-wrap:wrap}
.ch-tab{padding:10px 14px;border:none;background:none;cursor:pointer;font-size:13px;font-weight:600;color:var(--muted);border-bottom:3px solid transparent;margin-bottom:-2px;transition:color .15s;white-space:nowrap}
.ch-tab:hover{color:var(--text)}
.ch-tab.active{color:var(--primary);border-bottom-color:var(--primary)}
.ch-tab[data-ch="쿠팡"].active{color:#e8470a;border-bottom-color:#e8470a}
.ch-tab[data-ch="네이버/스마트스토어"].active{color:#03c75a;border-bottom-color:#03c75a}
.ch-tab[data-ch="G마켓/옥션/11번가"].active{color:#7c3aed;border-bottom-color:#7c3aed}
.ch-tab[data-ch="번개장터"].active{color:#f59e0b;border-bottom-color:#f59e0b}
.ch-tab[data-ch="렌탈"].active{color:#0f766e;border-bottom-color:#0f766e}
.ch-tab[data-ch="기업직납"].active{color:#1d4ed8;border-bottom-color:#1d4ed8}
.ch-tab[data-ch="자사몰/직판"].active{color:#be185d;border-bottom-color:#be185d}
.ch-tab[data-ch="딜러/파트너"].active{color:#92400e;border-bottom-color:#92400e}
.ch-tab[data-ch="일반소비자"].active{color:#d97706;border-bottom-color:#d97706}
.ch-tab-sep{width:1px;background:var(--border);margin:8px 4px;align-self:stretch}
.container{max-width:1400px;margin:0 auto;padding:24px}
.section{margin-bottom:28px}
.sec-title{font-size:14px;font-weight:700;margin-bottom:12px;padding-left:8px;border-left:3px solid var(--primary)}
/* KPI */
.kpi-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:14px}
.kpi-card{background:var(--card);border:1px solid var(--border);padding:20px}
.kpi-label{font-size:12px;color:var(--muted);margin-bottom:6px}
.kpi-value{font-size:36px;font-weight:700;line-height:1}
.kpi-card.loyal .kpi-value{color:var(--success)}
.kpi-card.churn .kpi-value{color:var(--danger)}
.kpi-card.crosssell .kpi-value{color:var(--primary)}
.kpi-card.upsell .kpi-value{color:var(--purple)}
/* Seg bar */
.seg-wrap{background:var(--card);border:1px solid var(--border);padding:20px}
.seg-bar{display:flex;height:36px;overflow:hidden;border-radius:2px}
.seg-item{display:flex;align-items:center;justify-content:center;font-size:11px;font-weight:600;color:#fff}
.seg-item.loyal{background:var(--success)}
.seg-item.churn{background:var(--danger)}
.seg-item.crosssell{background:var(--primary)}
.seg-item.upsell{background:var(--purple)}
.seg-legend{display:flex;gap:20px;margin-top:12px;flex-wrap:wrap}
.seg-legend-item{display:flex;align-items:center;gap:6px;font-size:12px}
.seg-dot{width:10px;height:10px;border-radius:2px;flex-shrink:0}
/* Grid */
.two-col{display:grid;grid-template-columns:1fr 1fr;gap:16px}
/* Card */
.card{background:var(--card);border:1px solid var(--border);padding:20px}
.chart-wrap{position:relative;height:260px}
/* Table */
table{width:100%;border-collapse:collapse}
th{text-align:left;font-size:11px;font-weight:600;color:var(--muted);border-bottom:1px solid var(--border);padding:6px 8px;text-transform:uppercase}
td{padding:8px;border-bottom:1px solid var(--border);font-size:13px;vertical-align:middle}
tr:last-child td{border-bottom:none}
tr.click{cursor:pointer}
tr.click:hover td{background:#f9fafb}
.amount{font-variant-numeric:tabular-nums}
/* Cross-sell */
.matrix td.has{background:#dbeafe;color:var(--primary);font-weight:600}
.matrix td.opp{background:#fef9c3;color:#854d0e;font-size:12px}
.matrix td.empty{color:var(--border);text-align:center}
/* Gap bar */
.gap-bar{height:5px;background:#e5e7eb;border-radius:3px;margin-top:3px}
.gap-fill{height:5px;background:var(--purple);border-radius:3px}
/* Modal */
.overlay{display:none;position:fixed;inset:0;background:rgba(0,0,0,.4);z-index:200;align-items:center;justify-content:center}
.overlay.open{display:flex}
.modal{background:var(--card);width:460px;max-width:90vw;padding:24px;max-height:90vh;overflow-y:auto}
.modal.wide{width:680px}
.modal h3{font-size:16px;font-weight:700;margin-bottom:16px;padding-right:24px}
.modal-section{font-size:11px;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:.5px;margin:16px 0 6px;padding-bottom:4px;border-bottom:2px solid var(--border)}
.hist-table{width:100%;border-collapse:collapse;font-size:12px;margin-top:4px}
.hist-table th{text-align:left;font-size:10px;font-weight:600;color:var(--muted);padding:4px 6px;border-bottom:1px solid var(--border)}
.hist-table td{padding:5px 6px;border-bottom:1px solid var(--border)}
.hist-table tr:last-child td{border-bottom:none}
.hist-table tr:hover td{background:#f9fafb}
.modal-close{float:right;cursor:pointer;font-size:22px;line-height:1;color:var(--muted);margin-top:-2px}
.mrow{display:flex;justify-content:space-between;padding:8px 0;border-bottom:1px solid var(--border);font-size:13px}
.mrow:last-child{border-bottom:none}
.mrow .lbl{color:var(--muted)}
.action-badge{background:#fee2e2;color:var(--danger);padding:2px 8px;font-size:11px;font-weight:600}
/* Column filter bar */
.cf-label{font-size:11px;font-weight:600;color:var(--muted)}
.cf-sel{padding:5px 8px;border:1px solid var(--border);background:var(--card);font-size:12px;color:var(--text);cursor:pointer;border-radius:3px;min-width:110px}
.cf-sel:focus{outline:none;border-color:var(--primary)}
/* Industry filter chips */
.ind-chip{display:inline-flex;align-items:center;gap:5px;padding:4px 10px;border-radius:20px;border:1px solid var(--border);font-size:11px;font-weight:600;cursor:pointer;transition:all .15s;background:var(--card);color:var(--muted)}
.ind-chip:hover{border-color:currentColor;opacity:.85}
.ind-chip.active{color:#fff;border-color:transparent}
.ind-chip .ind-dot{width:7px;height:7px;border-radius:50%;flex-shrink:0}
/* Responsive */
@media(max-width:960px){.kpi-grid{grid-template-columns:repeat(2,1fr)}.two-col{grid-template-columns:1fr}}
@media(max-width:480px){.kpi-grid{grid-template-columns:1fr 1fr}}
</style>
</head>
<body>
<div class="header">
  <h1>매출처 마케팅 대시보드</h1>
  <span class="meta" id="build-meta"></span>
</div>
<div class="demo-banner" id="demo-banner">⚠ 데모 데이터입니다. 실제 엑셀 파일을 지정하고 다시 빌드하세요.</div>
<div class="ch-tabs">
  <button class="ch-tab active" data-ch="all" onclick="setChannel('all')">전체</button>
  <div class="ch-tab-sep"></div>
  <button class="ch-tab" data-ch="쿠팡" onclick="setChannel('쿠팡')">쿠팡</button>
  <button class="ch-tab" data-ch="네이버/스마트스토어" onclick="setChannel('네이버/스마트스토어')">네이버</button>
  <button class="ch-tab" data-ch="G마켓/옥션/11번가" onclick="setChannel('G마켓/옥션/11번가')">G마켓/옥션/11번가</button>
  <button class="ch-tab" data-ch="번개장터" onclick="setChannel('번개장터')">번개장터</button>
  <div class="ch-tab-sep"></div>
  <button class="ch-tab" data-ch="렌탈" onclick="setChannel('렌탈')">렌탈</button>
  <button class="ch-tab" data-ch="기업직납" onclick="setChannel('기업직납')">기업직납</button>
  <button class="ch-tab" data-ch="자사몰/직판" onclick="setChannel('자사몰/직판')">자사몰/직판</button>
  <button class="ch-tab" data-ch="딜러/파트너" onclick="setChannel('딜러/파트너')">딜러/파트너</button>
  <div class="ch-tab-sep"></div>
  <button class="ch-tab" data-ch="일반소비자" onclick="setChannel('일반소비자')">일반소비자</button>
</div>
<div class="container">

  <!-- KPI -->
  <div class="section">
    <div class="sec-title">핵심 지표</div>
    <div class="kpi-grid">
      <div class="kpi-card"><div class="kpi-label">전체 거래처</div><div class="kpi-value" id="k-total">-</div></div>
      <div class="kpi-card churn"><div class="kpi-label">이탈 (재활성화 대상)</div><div class="kpi-value" id="k-churn">-</div></div>
      <div class="kpi-card crosssell"><div class="kpi-label">크로스셀 대상</div><div class="kpi-value" id="k-crosssell">-</div></div>
      <div class="kpi-card upsell"><div class="kpi-label">업셀 대상</div><div class="kpi-value" id="k-upsell">-</div></div>
    </div>
  </div>

  <!-- Segment bar -->
  <div class="section">
    <div class="sec-title">거래처 세그먼트 분포</div>
    <div class="seg-wrap">
      <div class="seg-bar" id="seg-bar"></div>
      <div class="seg-legend">
        <div class="seg-legend-item"><div class="seg-dot" style="background:var(--success)"></div>충성</div>
        <div class="seg-legend-item"><div class="seg-dot" style="background:var(--danger)"></div>이탈</div>
        <div class="seg-legend-item"><div class="seg-dot" style="background:var(--primary)"></div>크로스셀</div>
        <div class="seg-legend-item"><div class="seg-dot" style="background:var(--purple)"></div>업셀</div>
      </div>
    </div>
  </div>

  <!-- Monthly chart -->
  <div class="section">
    <div class="sec-title" id="monthly-title">월별 매출 추이</div>
    <div class="card"><div class="chart-wrap"><canvas id="monthlyChart"></canvas></div></div>
  </div>

  <!-- Reactivation (full width) -->
  <div class="section">
    <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:10px">
      <div class="sec-title" style="margin-bottom:0">재활성화 대상 추출 <span id="churn-count-badge" style="font-size:11px;color:var(--muted);font-weight:400"></span></div>
      <button id="churn-excel-btn" onclick="exportChurnCsv()" style="padding:7px 14px;background:#16a34a;color:#fff;border:none;cursor:pointer;font-size:12px;font-weight:600;border-radius:3px;white-space:nowrap">📥 현재 필터 엑셀 저장</button>
    </div>
    <div id="ind-filter-wrap" style="display:flex;flex-wrap:wrap;gap:6px;margin-bottom:8px"></div>
    <div style="display:flex;flex-wrap:wrap;gap:8px;align-items:center;margin-bottom:10px">
      <label class="cf-label">담당자</label>
      <select id="cf-rep"    class="cf-sel" onchange="setCfFilter()"><option value="all">전체</option></select>
      <label class="cf-label">지역</label>
      <select id="cf-region" class="cf-sel" onchange="setCfFilter()"><option value="all">전체</option></select>
      <label class="cf-label">채널</label>
      <select id="cf-ch"     class="cf-sel" onchange="setCfFilter()"><option value="all">전체</option></select>
      <div style="width:1px;height:22px;background:var(--border);margin:0 4px"></div>
      <label class="cf-label" style="color:var(--primary)">시트 분류</label>
      <select id="cf-group"  class="cf-sel" style="border-color:var(--primary)">
        <option value="none">분류 없음 (1시트)</option>
        <option value="industry">업종별</option>
        <option value="rep">담당자별</option>
        <option value="region">지역별</option>
        <option value="channel">채널별</option>
      </select>
      <button onclick="resetAllFilters()" style="padding:5px 12px;border:1px solid var(--border);background:var(--card);cursor:pointer;font-size:12px;color:var(--muted);border-radius:3px">초기화</button>
    </div>
    <div class="card" style="padding:0">
      <div id="churn-scroll" style="overflow:auto;max-height:900px;cursor:grab">
        <table id="churn-tbl" style="min-width:900px">
          <thead style="position:sticky;top:0;z-index:2;background:var(--card)">
            <tr><th>#</th><th>거래처명</th><th>추정기업</th><th>업종</th><th>영업담당</th><th>지역</th><th>주소</th><th>핸드폰</th><th id="churn-yr-col">전년 매출</th><th>주문</th><th>카테고리</th></tr>
          </thead>
          <tbody></tbody>
        </table>
      </div>
    </div>
  </div>

  <!-- Top by Freq / Top by Revenue -->
  <div class="section" style="margin-bottom:8px">
    <div id="top-ind-filter-wrap" style="display:flex;flex-wrap:wrap;gap:6px"></div>
  </div>
  <div class="two-col section">
    <div>
      <div class="sec-title">구매 빈도 상위 50</div>
      <div class="card" style="overflow-x:auto">
        <table id="top-freq-tbl">
          <thead><tr><th>#</th><th>거래처명</th><th>담당자</th><th>채널</th><th>주문건수</th><th>매출</th><th>세그먼트</th></tr></thead>
          <tbody></tbody>
        </table>
      </div>
    </div>
    <div>
      <div class="sec-title">매출 상위 50</div>
      <div class="card" style="overflow-x:auto">
        <table id="top-rev-tbl">
          <thead><tr><th>#</th><th>거래처명</th><th>담당자</th><th>채널</th><th>매출</th><th>주문건수</th><th>세그먼트</th></tr></thead>
          <tbody></tbody>
        </table>
      </div>
    </div>
  </div>

  <!-- Cross-sell -->
  <div class="section">
    <div class="sec-title">크로스셀 기회 매트릭스</div>
    <div class="card" style="overflow-x:auto">
      <table class="matrix" id="cs-tbl">
        <thead><tr><th>현재 구매 카테고리</th><th>구매 거래처 수</th><th>추천 카테고리 1</th><th>추천 카테고리 2</th><th>추천 카테고리 3</th></tr></thead>
        <tbody></tbody>
      </table>
    </div>
  </div>

  <!-- Rep + Regional -->
  <div class="two-col section">
    <div>
      <div class="sec-title">영업담당자별 거래처 현황</div>
      <div class="card" style="overflow-x:auto">
        <table id="rep-tbl">
          <thead><tr><th>담당자</th><th>충성</th><th>이탈</th><th>크로스셀</th><th>업셀</th><th id="rep-yr-col">최근년 매출</th></tr></thead>
          <tbody></tbody>
        </table>
      </div>
    </div>
    <div>
      <div class="sec-title">지역별 매출 현황 (상위 10)</div>
      <div class="card"><div class="chart-wrap"><canvas id="regionalChart"></canvas></div></div>
    </div>
  </div>

  <!-- RFM 분석 -->
  <div class="section">
    <div class="sec-title">RFM 분석 <span style="font-size:11px;color:var(--muted);font-weight:400">— 최근성(R) · 빈도(F) · 금액(M) 기준 (<span id="rfm-ref"></span> 기준)</span></div>
    <div class="two-col">
      <div class="card">
        <div style="display:grid;grid-template-columns:repeat(3,1fr);gap:10px;margin-bottom:16px" id="rfm-seg-grid"></div>
        <div class="chart-wrap" style="height:220px"><canvas id="rfmChart"></canvas></div>
      </div>
      <div class="card" style="overflow-x:auto">
        <table id="rfm-tbl">
          <thead><tr><th>#</th><th>거래처명</th><th>담당</th><th style="text-align:center">R</th><th style="text-align:center">F</th><th style="text-align:center">M</th><th>매출</th><th>세그먼트</th></tr></thead>
          <tbody></tbody>
        </table>
      </div>
    </div>
  </div>

  <!-- ABC 분석 -->
  <div class="section">
    <div class="sec-title">ABC 분석 <span style="font-size:11px;color:var(--muted);font-weight:400">— 파레토 법칙 기반 거래처 등급</span></div>
    <div class="two-col">
      <div>
        <div style="display:grid;grid-template-columns:repeat(3,1fr);gap:12px;margin-bottom:16px" id="abc-summary"></div>
        <div class="card" style="overflow-x:auto">
          <table id="abc-tbl">
            <thead><tr><th>등급</th><th>거래처명</th><th>담당</th><th>매출</th></tr></thead>
            <tbody></tbody>
          </table>
        </div>
      </div>
      <div class="card">
        <div style="font-size:12px;color:var(--muted);margin-bottom:8px">상위 50개사 파레토 차트</div>
        <div class="chart-wrap" style="height:300px"><canvas id="paretoChart"></canvas></div>
      </div>
    </div>
  </div>

  <!-- 품목 구매패턴 -->
  <div class="section">
    <div class="sec-title">품목 구매패턴 분석</div>
    <div class="two-col">
      <div>
        <div class="sec-title" style="font-size:13px;border-left-color:var(--warning)">카테고리 친화도 TOP 20</div>
        <div class="card" style="overflow-x:auto">
          <table id="affinity-tbl">
            <thead><tr><th>카테고리 A</th><th>카테고리 B</th><th>공동구매 고객</th><th>친화도</th></tr></thead>
            <tbody></tbody>
          </table>
        </div>
      </div>
      <div>
        <div class="sec-title" style="font-size:13px;border-left-color:var(--warning)">연도별 카테고리 매출 비교</div>
        <div class="card">
          <div class="chart-wrap" style="height:300px"><canvas id="catYrChart"></canvas></div>
        </div>
      </div>
    </div>
  </div>

  <!-- 업종별 분석 -->
  <div class="section">
    <div class="sec-title">배송지 기반 업종별 분석 <span style="font-size:11px;color:var(--muted);font-weight:400">— 주소·수령자에서 업종 추정</span></div>
    <div class="two-col">
      <div class="card">
        <div class="chart-wrap" style="height:320px"><canvas id="industryChart"></canvas></div>
      </div>
      <div class="card" style="overflow-x:auto">
        <table id="industry-tbl">
          <thead><tr><th>업종</th><th style="text-align:right">거래처수</th><th style="text-align:right">매출</th><th style="text-align:right">비율</th><th>매출 비중</th></tr></thead>
          <tbody></tbody>
        </table>
      </div>
    </div>
  </div>

</div>

<!-- Modal -->
<div class="overlay" id="overlay" onclick="if(event.target===this)closeModal()">
  <div class="modal">
    <span class="modal-close" onclick="closeModal()">×</span>
    <h3 id="m-title"></h3>
    <div id="m-body"></div>
  </div>
</div>

<script>
const D = __DATA_JSON__;

// Banner
document.getElementById('build-meta').textContent = '생성일: ' + D.build_date + (D.mode==='demo' ? ' · 데모 데이터' : '');
if (D.mode === 'demo') document.getElementById('demo-banner').style.display = 'block';

// ── 채널 탭 ─────────────────────────────────────────────────────────────────
let activeChannel = 'all';
let monthlyChartObj = null;

function getKpi(ch) {
  if (ch === 'all' || !D.channel_kpi || !D.channel_kpi[ch]) return D.kpi;
  return D.channel_kpi[ch];
}
function getMonthlyData(ch) {
  if (ch === 'all' || !D.monthly_by_channel || !D.monthly_by_channel[ch]) return D.monthly_by_year;
  return D.monthly_by_channel[ch];
}

function renderKpi(ch) {
  const kpi = getKpi(ch);
  document.getElementById('k-total').textContent     = (kpi.total||0).toLocaleString();
  document.getElementById('k-churn').textContent     = (kpi.churn||0).toLocaleString();
  document.getElementById('k-crosssell').textContent = (kpi.crosssell||0).toLocaleString();
  document.getElementById('k-upsell').textContent    = (kpi.upsell||0).toLocaleString();
  // Seg bar
  const bar = document.getElementById('seg-bar');
  bar.innerHTML = '';
  const tot = kpi.total || 1;
  [['loyal',kpi.loyal||0],['churn',kpi.churn||0],['crosssell',kpi.crosssell||0],['upsell',kpi.upsell||0]].forEach(([cls,val])=>{
    const pct = (val/tot*100).toFixed(1);
    const el = document.createElement('div');
    el.className = 'seg-item ' + cls;
    el.style.width = pct + '%';
    el.title = cls + ': ' + val + '개사 (' + pct + '%)';
    if (parseFloat(pct) > 4) el.textContent = pct + '%';
    bar.appendChild(el);
  });
}

function renderMonthly(ch) {
  const mData = getMonthlyData(ch);
  if (!monthlyChartObj) return;
  monthlyChartObj.data.datasets.forEach((ds, i) => {
    const yr = yrList[i];
    ds.data = months.map((_,mi) => ((mData[yr]||{})[mi+1+'']||0)/1e6);
  });
  monthlyChartObj.update();
}

function filterTableByChannel(tbId, ch) {
  document.querySelectorAll(`#${tbId} tbody tr`).forEach(tr => {
    tr.style.display = (ch === 'all' || tr.dataset.ch === ch) ? '' : 'none';
  });
}

function setChannel(ch) {
  activeChannel = ch;
  document.querySelectorAll('.ch-tab').forEach(btn => btn.classList.toggle('active', btn.dataset.ch === ch));
  renderKpi(ch);
  renderMonthly(ch);
  filterTableByChannel('churn-tbl', ch);
}

// KPI (초기 렌더)
renderKpi('all');

// Seg bar (초기 렌더 — renderKpi 내에서 처리)

// Monthly chart (동적 연도)
const months = ['1월','2월','3월','4월','5월','6월','7월','8월','9월','10월','11월','12월'];
const YR_STYLES = {
  default0: {borderColor:'#d1d5db',backgroundColor:'transparent',borderDash:[4,3]},
  default1: {borderColor:'#94a3b8',backgroundColor:'transparent',borderDash:[4,3]},
  cur:      {borderColor:'#2563eb',backgroundColor:'rgba(37,99,235,.07)',fill:true},
};
const yrList = D.years || [];
document.getElementById('monthly-title').textContent =
  '월별 매출 추이 (' + yrList.map(y=>y.slice(2)+'년').join(' vs ') + ')';
const monthDatasets = yrList.map((yr, i) => {
  const isCur = yr === D.cur_yr;
  const style = isCur ? YR_STYLES.cur : (i===0 ? YR_STYLES.default0 : YR_STYLES.default1);
  return Object.assign({
    label: yr.slice(2)+'년',
    data: months.map((_,mi)=>((D.monthly_by_year[yr]||{})[mi+1+'']||0)/1e6),
    tension:.35, pointRadius:3,
  }, style);
});
monthlyChartObj = new Chart(document.getElementById('monthlyChart'),{
  type:'line',
  data:{labels:months, datasets:monthDatasets},
  options:{responsive:true,maintainAspectRatio:false,
    plugins:{legend:{position:'top'}},
    scales:{
      y:{ticks:{callback:v=>v.toFixed(0)+'M'},grid:{color:'#f0f0f0'}},
      x:{grid:{display:false}}
    }
  }
});

// Regional chart
new Chart(document.getElementById('regionalChart'),{
  type:'bar',
  data:{
    labels:D.regional.map(r=>r.region),
    datasets:[{label:'매출(백만)',data:D.regional.map(r=>r.total/1e6),backgroundColor:'#2563eb',borderRadius:2}]
  },
  options:{indexAxis:'y',responsive:true,maintainAspectRatio:false,
    plugins:{legend:{display:false}},
    scales:{
      x:{ticks:{callback:v=>v.toFixed(0)+'M'},grid:{color:'#f0f0f0'}},
      y:{grid:{display:false}}
    }
  }
});

// Churn table + 업종 필터
const prevYrLabel = D.prev_yr ? D.prev_yr.slice(2)+'년' : '전년';
document.getElementById('churn-yr-col').textContent = prevYrLabel + ' 매출';
document.getElementById('rep-yr-col').textContent = D.cur_yr ? D.cur_yr.slice(2)+'년 매출' : '최근년 매출';
const churnTb = document.querySelector('#churn-tbl tbody');

const IND_COLORS_JS = {
  '교육':'#f59e0b','의료/복지':'#ef4444','외식/음식':'#f97316','제조/산업':'#3b82f6',
  '종교':'#8b5cf6','공공/기관':'#14b8a6','판매/유통':'#06b6d4',
  '개인/주거':'#9ca3af','기타 사업체':'#6b7280','미분류':'#d1d5db'
};

let activeIndFilter    = 'all';
let activeChFilter     = 'all';
let activeRepFilter    = 'all';
let activeRegionFilter = 'all';
let activeCfChFilter   = 'all';

function churnPassFilter(c) {
  if (activeChFilter     !== 'all' && c.channel  !== activeChFilter)     return false;
  if (activeCfChFilter   !== 'all' && c.channel  !== activeCfChFilter)   return false;
  if (activeIndFilter    !== 'all' && c.industry !== activeIndFilter)    return false;
  if (activeRepFilter    !== 'all' && c.rep      !== activeRepFilter)    return false;
  if (activeRegionFilter !== 'all' && c.region   !== activeRegionFilter) return false;
  return true;
}

const CHURN_DISPLAY_LIMIT = 50;

function renderChurnTable() {
  churnTb.innerHTML = '';
  let seq = 0;
  for (const c of D.churn_top) {
    if (!churnPassFilter(c)) continue;
    if (seq >= CHURN_DISPLAY_LIMIT) break;
    seq++;
    const tr = document.createElement('tr');
    tr.className = 'click';
    if (c.channel)  tr.dataset.ch  = c.channel;
    if (c.industry) tr.dataset.ind = c.industry;
    const hint = c.company_hint ? `<span style="font-size:11px;color:var(--primary);margin-left:4px">(${c.company_hint})</span>` : '';
    const indColor = IND_COLORS_JS[c.industry] || '#888';
    tr.innerHTML =
      `<td style="color:var(--muted);font-size:12px">${seq}</td>` +
      `<td><strong>${c.name}</strong>${hint}</td>` +
      `<td style="font-size:12px;color:var(--primary)">${c.company_hint||'<span style="color:var(--border)">-</span>'}</td>` +
      `<td><span style="font-size:11px;font-weight:600;color:${indColor}">${c.industry||'-'}</span></td>` +
      `<td>${c.rep}</td>` +
      `<td style="font-size:12px;color:var(--muted)">${c.region||'-'}</td>` +
      `<td style="font-size:12px;max-width:200px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis" title="${c.addr||''}">${c.addr||'-'}</td>` +
      `<td style="font-size:12px;white-space:nowrap">${c.mobile||'-'}</td>` +
      `<td class="amount">${Math.round(c.total_prev/1e4).toLocaleString()}만원</td>` +
      `<td style="color:var(--muted)">${c.order_count_prev}건</td>` +
      `<td style="font-size:11px;color:var(--muted)">${(c.cats||[]).join(', ')||'-'}</td>`;
    tr.onclick = () => showModal(c);
    churnTb.appendChild(tr);
  }
  // 전체 필터 통과 건수 (표시 제한 무관)
  const totalFiltered = D.churn_top.filter(c => churnPassFilter(c)).length;
  const badge = document.getElementById('churn-count-badge');
  const parts = [];
  if (activeIndFilter    !== 'all') parts.push(activeIndFilter);
  if (activeRepFilter    !== 'all') parts.push(activeRepFilter);
  if (activeRegionFilter !== 'all') parts.push(activeRegionFilter);
  if (activeCfChFilter   !== 'all') parts.push(activeCfChFilter);
  const label = parts.length ? parts.join(' · ') : '전체';
  const isFiltered = parts.length > 0;
  const limitNote = totalFiltered > CHURN_DISPLAY_LIMIT ? ` (상위 ${CHURN_DISPLAY_LIMIT}개 표시)` : '';
  badge.textContent = `${label} · ${totalFiltered.toLocaleString()}개사${isFiltered ? ' (현재 필터 기준)' : ''}${limitNote} (클릭 시 상세)`;
  document.getElementById('churn-excel-btn').textContent = isFiltered
    ? `📥 [${label}] ${totalFiltered.toLocaleString()}개사 엑셀 저장`
    : `📥 현재 필터 엑셀 저장`;
  document.getElementById('churn-scroll').scrollTop = 0;
}

// 업종 필터 칩 생성 (전체 카운트는 churn_industry_counts 사용)
const filterWrap = document.getElementById('ind-filter-wrap');
const indCounts = D.churn_industry_counts || {};
const totalChurnCount = D.churn_total_count || D.churn_top.length;
const indList = ['all', ...Object.keys(indCounts).sort((a,b)=>(indCounts[b]||0)-(indCounts[a]||0))];
indList.forEach(ind => {
  const btn = document.createElement('button');
  btn.className = 'ind-chip' + (ind==='all'?' active':'');
  const color = IND_COLORS_JS[ind] || '#6b7280';
  const cnt = ind==='all' ? totalChurnCount : (indCounts[ind] || 0);
  if (ind === 'all') {
    btn.style.cssText = 'background:#1e293b;color:#fff;border-color:#1e293b';
    btn.innerHTML = `전체 <span style="opacity:.7">${cnt.toLocaleString()}</span>`;
  } else {
    btn.innerHTML = `<span class="ind-dot" style="background:${color}"></span>${ind} <span style="opacity:.6">${cnt.toLocaleString()}</span>`;
    btn.dataset.color = color;
  }
  btn.onclick = () => {
    activeIndFilter = ind;
    filterWrap.querySelectorAll('.ind-chip').forEach(b => {
      b.classList.remove('active');
      b.style.background = '';
      b.style.color = '';
      b.style.borderColor = '';
    });
    btn.classList.add('active');
    if (ind === 'all') {
      btn.style.cssText = 'background:#1e293b;color:#fff;border-color:#1e293b';
    } else {
      btn.style.background = color;
      btn.style.borderColor = color;
      btn.style.color = '#fff';
    }
    renderChurnTable();
  };
  filterWrap.appendChild(btn);
});

// 드롭다운 옵션 초기화
(function buildDropdowns() {
  const reps    = [...new Set(D.churn_top.map(c => c.rep).filter(Boolean))].sort();
  const regions = [...new Set(D.churn_top.map(c => c.region).filter(Boolean))].sort();
  const chs     = [...new Set(D.churn_top.map(c => c.channel).filter(Boolean))].sort();
  const selRep = document.getElementById('cf-rep');
  const selReg = document.getElementById('cf-region');
  const selCh  = document.getElementById('cf-ch');
  reps.forEach(v    => { const o = document.createElement('option'); o.value = v; o.textContent = v; selRep.appendChild(o); });
  regions.forEach(v => { const o = document.createElement('option'); o.value = v; o.textContent = v; selReg.appendChild(o); });
  chs.forEach(v     => { const o = document.createElement('option'); o.value = v; o.textContent = v; selCh.appendChild(o); });
})();

function setCfFilter() {
  activeRepFilter    = document.getElementById('cf-rep').value;
  activeRegionFilter = document.getElementById('cf-region').value;
  activeCfChFilter   = document.getElementById('cf-ch').value;
  renderChurnTable();
}

function resetAllFilters() {
  activeIndFilter = 'all';
  activeRepFilter = 'all';
  activeRegionFilter = 'all';
  activeCfChFilter = 'all';
  document.getElementById('cf-rep').value    = 'all';
  document.getElementById('cf-region').value = 'all';
  document.getElementById('cf-ch').value     = 'all';
  // 업종 칩 초기화
  document.querySelectorAll('.ind-chip').forEach(b => {
    b.classList.remove('active');
    b.style.background = ''; b.style.color = ''; b.style.borderColor = '';
  });
  const allChip = document.querySelector('.ind-chip');
  if (allChip) { allChip.classList.add('active'); allChip.style.cssText = 'background:#1e293b;color:#fff;border-color:#1e293b'; }
  renderChurnTable();
}

renderChurnTable();

// 드래그 스크롤
(function(){
  const el = document.getElementById('churn-scroll');
  if (!el) return;
  let down = false, startY, startX, scrollY, scrollX;
  el.addEventListener('mousedown', e => {
    down = true; startY = e.clientY; startX = e.clientX;
    scrollY = el.scrollTop; scrollX = el.scrollLeft;
    el.style.cursor = 'grabbing'; e.preventDefault();
  });
  document.addEventListener('mouseup',   () => { down = false; el.style.cursor = 'grab'; });
  document.addEventListener('mousemove', e => {
    if (!down) return;
    el.scrollTop  = scrollY - (e.clientY - startY);
    el.scrollLeft = scrollX - (e.clientX - startX);
  });
})();

// ── 구매빈도 / 매출 상위 테이블 ──────────────────────────────────────────────
const SEG_COLORS = {'충성':'#16a34a','이탈':'#dc2626','크로스셀':'#2563eb','업셀':'#7c3aed'};
function segBadge(seg) {
  const c = SEG_COLORS[seg] || '#888';
  return `<span style="font-size:10px;font-weight:700;color:${c};background:${c}18;padding:2px 6px;border-radius:3px">${seg}</span>`;
}

function renderTopTable(tbId, list, byFreq) {
  const tb = document.querySelector(`#${tbId} tbody`);
  if (!tb || !list) return;
  tb.innerHTML = '';
  list.forEach((c, i) => {
    const tr = document.createElement('tr');
    tr.className = 'click';
    const hint = c.company_hint ? `<span style="font-size:10px;color:var(--primary);margin-left:3px">(${c.company_hint})</span>` : '';
    const rev  = Math.round((c.revenue||0)/1e4).toLocaleString() + '만원';
    const freq = (c.order_count||0) + '건';
    tr.innerHTML = byFreq
      ? `<td style="color:var(--muted);font-size:12px">${i+1}</td>
         <td><strong>${c.name}</strong>${hint}</td>
         <td style="font-size:12px">${c.rep||'-'}</td>
         <td style="font-size:12px">${c.channel||'-'}</td>
         <td style="font-weight:700;color:var(--primary)">${freq}</td>
         <td class="amount" style="font-size:12px">${rev}</td>
         <td>${segBadge(c.segment)}</td>`
      : `<td style="color:var(--muted);font-size:12px">${i+1}</td>
         <td><strong>${c.name}</strong>${hint}</td>
         <td style="font-size:12px">${c.rep||'-'}</td>
         <td style="font-size:12px">${c.channel||'-'}</td>
         <td class="amount" style="font-weight:700;color:var(--primary)">${rev}</td>
         <td style="font-size:12px">${freq}</td>
         <td>${segBadge(c.segment)}</td>`;
    tr.onclick = () => showTopModal(c);
    tb.appendChild(tr);
  });
}

function showTopModal(c) {
  const modal = document.querySelector('#overlay .modal');
  modal.classList.add('wide');
  document.getElementById('m-title').innerHTML =
    c.name + (c.company_hint ? `<span style="font-size:13px;color:var(--primary);font-weight:400;margin-left:8px">${c.company_hint}</span>` : '');

  // 기본정보 섹션
  let html = `<div class="modal-section">기본 정보</div>`;
  html += row('세그먼트', segBadge(c.segment));
  html += row('업종', c.industry || '-');
  html += row('영업담당', c.rep || '-');
  html += row('채널', c.channel || '-');
  html += row('지역', c.region || '-');
  if (c.addr)   html += row('주소', c.addr);
  if (c.mobile) html += row('핸드폰', `<a href="tel:${c.mobile}" style="color:var(--primary)">${c.mobile}</a>`);

  // 구매 요약
  html += `<div class="modal-section">구매 요약</div>`;
  html += row('총 매출', Math.round((c.revenue||0)/1e4).toLocaleString() + '만원');
  html += row('총 주문건수', (c.order_count||0) + '건');
  html += row('첫 구매', c.first_order || '-');
  html += row('최근 구매', c.last_order  || '-');
  html += row('구매 카테고리', (c.cats||[]).join(', ') || '-');

  // 구매이력 테이블
  const orders = c.orders || [];
  if (orders.length) {
    html += `<div class="modal-section">구매 이력 (최근 ${orders.length}건)</div>`;
    html += `<table class="hist-table">
      <thead><tr><th>날짜</th><th>대분류</th><th>중분류</th><th>채널</th><th>판매처</th><th style="text-align:right">금액</th></tr></thead>
      <tbody>`;
    orders.forEach(o => {
      const amt = o.total >= 10000
        ? Math.round(o.total/1e4).toLocaleString() + '만원'
        : o.total.toLocaleString() + '원';
      html += `<tr>
        <td style="white-space:nowrap;color:var(--muted)">${o.date}</td>
        <td>${o.cat1||'-'}</td>
        <td style="color:var(--muted)">${o.cat2||'-'}</td>
        <td style="font-size:11px;white-space:nowrap">${o.channel||'-'}</td>
        <td style="font-size:11px;color:var(--muted);max-width:120px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap" title="${o.site||''}">${o.site||'-'}</td>
        <td style="text-align:right;font-variant-numeric:tabular-nums">${amt}</td>
      </tr>`;
    });
    html += `</tbody></table>`;
  } else {
    html += `<div style="color:var(--muted);font-size:13px;padding:12px 0">이력 없음</div>`;
  }

  document.getElementById('m-body').innerHTML = html;
  document.getElementById('overlay').classList.add('open');
}

// 업종 필터
let activeTopInd = 'all';

function topIndFilter(list) {
  if (activeTopInd === 'all') return list;
  return list.filter(c => c.industry === activeTopInd);
}

function renderBothTopTables() {
  if (D.top_by_freq)    renderTopTable('top-freq-tbl', topIndFilter(D.top_by_freq),    true);
  if (D.top_by_revenue) renderTopTable('top-rev-tbl',  topIndFilter(D.top_by_revenue), false);
}

// 업종 칩 생성
(function(){
  const wrap = document.getElementById('top-ind-filter-wrap');
  if (!wrap) return;
  const allList = [...(D.top_by_freq||[]), ...(D.top_by_revenue||[])];
  const indCnt = {};
  allList.forEach(c => { if (c.industry) indCnt[c.industry] = (indCnt[c.industry]||0) + 1; });
  const inds = ['all', ...Object.keys(indCnt).sort((a,b) => indCnt[b]-indCnt[a])];

  inds.forEach(ind => {
    const btn = document.createElement('button');
    btn.className = 'ind-chip' + (ind==='all' ? ' active' : '');
    const color = ind === 'all' ? '#1e293b' : (IND_COLORS_JS[ind] || '#888');
    const cnt   = ind === 'all' ? allList.length / 2 : indCnt[ind];
    btn.textContent = (ind === 'all' ? '전체' : ind) + ` (${Math.round(cnt)})`;
    if (ind === 'all') btn.style.cssText = 'background:#1e293b;color:#fff;border-color:#1e293b';
    btn.onclick = () => {
      activeTopInd = ind;
      wrap.querySelectorAll('.ind-chip').forEach(b => {
        b.classList.remove('active');
        b.style.background = ''; b.style.color = ''; b.style.borderColor = '';
      });
      btn.classList.add('active');
      if (ind === 'all') {
        btn.style.cssText = 'background:#1e293b;color:#fff;border-color:#1e293b';
      } else {
        btn.style.cssText = `background:${color};color:#fff;border-color:${color}`;
      }
      renderBothTopTables();
    };
    wrap.appendChild(btn);
  });
})();

renderBothTopTables();

// (기존 setChannel에서 churn 필터도 연동)
const _origSetChannel = setChannel;
setChannel = function(ch) {
  activeChFilter = ch;
  _origSetChannel(ch);
  renderChurnTable();
};

// Excel export — SheetJS 기반, 시트 분류 선택 지원
function exportChurnCsv() {
  if (typeof XLSX === 'undefined') { alert('SheetJS 로딩 중입니다. 잠시 후 다시 시도해주세요.'); return; }
  const prevLbl = (D.prev_yr || '25').slice(2) + '년';
  const HEADERS = ['순위','거래처명','추정기업명','업종','영업담당','지역','주소','핸드폰',prevLbl+' 매출','주문건수','구매카테고리','채널'];
  const COL_WIDTHS = [6, 18, 16, 12, 8, 10, 30, 14, 12, 8, 24, 8];

  const GROUP_KEY = { industry:'industry', rep:'rep', region:'region', channel:'channel' };
  const groupBy = document.getElementById('cf-group').value;

  function toRow(seq, c) {
    return [
      seq,
      c.name || '',
      c.company_hint || '',
      c.industry || '',
      c.rep || '',
      c.region || '',
      c.addr || '',
      c.mobile || '',
      Math.round((c.total_prev || 0) / 1e4) + '만원',
      (c.order_count_prev || 0) + '건',
      (c.cats || []).join(', '),
      c.channel || '',
    ];
  }

  function makeSheet(dataRows) {
    const aoa = [HEADERS, ...dataRows];
    const ws = XLSX.utils.aoa_to_sheet(aoa);
    ws['!cols'] = COL_WIDTHS.map(w => ({ wch: w }));
    // 헤더 행 고정
    ws['!freeze'] = { xSplit: 0, ySplit: 1 };
    return ws;
  }

  const wb = XLSX.utils.book_new();
  const filtered = [];
  D.churn_top.forEach(c => { if (churnPassFilter(c)) filtered.push(c); });

  if (groupBy === 'none' || !GROUP_KEY[groupBy]) {
    const dataRows = filtered.map((c, i) => toRow(i + 1, c));
    XLSX.utils.book_append_sheet(wb, makeSheet(dataRows), '재활성화');
  } else {
    const key = GROUP_KEY[groupBy];
    const groups = new Map();
    filtered.forEach(c => {
      const gv = c[key] || '미분류';
      if (!groups.has(gv)) groups.set(gv, []);
      groups.get(gv).push(c);
    });
    // 매출 합계 기준 내림차순 정렬
    const sorted = [...groups.entries()].sort((a, b) =>
      b[1].reduce((s, c) => s + (c.total_prev || 0), 0) -
      a[1].reduce((s, c) => s + (c.total_prev || 0), 0)
    );
    sorted.forEach(([gv, rows]) => {
      const dataRows = rows.map((c, i) => toRow(i + 1, c));
      const sheetName = String(gv).slice(0, 31).replace(/[:\\/?*\[\]]/g, '_');
      XLSX.utils.book_append_sheet(wb, makeSheet(dataRows), sheetName || '기타');
    });
    // 요약 시트 (첫 번째에 삽입)
    const summaryAoa = [['분류','거래처수',prevLbl+' 총매출(만원)','평균 매출(만원)']];
    sorted.forEach(([gv, rows]) => {
      const tot = rows.reduce((s, c) => s + (c.total_prev || 0), 0);
      summaryAoa.push([gv, rows.length, Math.round(tot / 1e4), Math.round(tot / 1e4 / (rows.length || 1))]);
    });
    const wsSum = XLSX.utils.aoa_to_sheet(summaryAoa);
    wsSum['!cols'] = [{wch:16},{wch:10},{wch:18},{wch:16}];
    XLSX.utils.book_append_sheet(wb, wsSum, '요약');
    // 요약 시트를 맨 앞으로
    wb.SheetNames = ['요약', ...wb.SheetNames.filter(n => n !== '요약')];
  }

  const parts = [];
  if (activeIndFilter    !== 'all') parts.push(activeIndFilter);
  if (activeRepFilter    !== 'all') parts.push(activeRepFilter);
  if (activeRegionFilter !== 'all') parts.push(activeRegionFilter);
  if (activeCfChFilter   !== 'all') parts.push(activeCfChFilter);
  const label = parts.length ? parts.join('_') : '전체';
  const groupLabel = groupBy === 'none' ? '' : `_${groupBy}별분류`;
  const date = D.build_date ? D.build_date.slice(0, 10) : '';
  XLSX.writeFile(wb, `재활성화_${label}${groupLabel}_${date}.xlsx`);
}

// Cross-sell matrix
const csTb=document.querySelector('#cs-tbl tbody');
D.crosssell_matrix.forEach(row=>{
  const tr=document.createElement('tr');
  let cells=`<td class="has">${row.current_cat}</td><td style="text-align:center;color:var(--muted)">${row.buyer_count||''}개사</td>`;
  for(let i=0;i<3;i++){
    const o=row.opportunities[i];
    cells+=o?`<td class="opp">${o.cat} <span style="opacity:.6">(${o.co_count}개사 · 친화도 ${o.affinity}%)</span></td>`:`<td class="empty">-</td>`;
  }
  tr.innerHTML=cells;
  csTb.appendChild(tr);
});

// Rep table
const repTb=document.querySelector('#rep-tbl tbody');
D.rep_stats.forEach(r=>{
  const tr=document.createElement('tr');
  tr.innerHTML=`<td><strong>${r.name||'-'}</strong></td>`+
    `<td style="color:var(--success)">${r.loyal}</td>`+
    `<td style="color:var(--danger)">${r.churn}</td>`+
    `<td style="color:var(--primary)">${r.crosssell}</td>`+
    `<td style="color:var(--purple)">${r.upsell}</td>`+
    `<td class="amount">${(r.revenue/1e6).toFixed(1)}M</td>`;
  repTb.appendChild(tr);
});

// ── RFM ────────────────────────────────────────────────────────────────────
if (D.rfm && D.rfm.total > 0) {
  document.getElementById('rfm-ref').textContent = D.rfm.ref_date;
  const RFM_SEG_COLORS = {
    '최우수':'#16a34a','충성고객':'#2563eb','신규/잠재':'#d97706',
    '이탈위험':'#dc2626','관찰필요':'#7c3aed','휴면':'#9ca3af'
  };
  const segGrid = document.getElementById('rfm-seg-grid');
  Object.entries(D.rfm.seg_counts).forEach(([seg, cnt])=>{
    const c2 = RFM_SEG_COLORS[seg]||'#6b7280';
    const d = document.createElement('div');
    d.style.cssText=`border:1px solid ${c2};padding:10px;text-align:center`;
    d.innerHTML=`<div style="font-size:11px;color:${c2};font-weight:600">${seg}</div><div style="font-size:22px;font-weight:700;color:${c2}">${cnt}</div>`;
    segGrid.appendChild(d);
  });
  new Chart(document.getElementById('rfmChart'),{
    type:'radar',
    data:{
      labels:['R (최근성)','F (빈도)','M (금액)'],
      datasets: Object.keys(RFM_SEG_COLORS).filter(s=>D.rfm.seg_counts[s]).slice(0,4).map(seg=>{
        const rows = D.rfm.top.filter(r=>r.segment===seg).slice(0,5);
        const avg = k => rows.length ? rows.reduce((a,r)=>a+r[k],0)/rows.length : 0;
        return {label:seg,data:[avg('R'),avg('F'),avg('M')],
          borderColor:RFM_SEG_COLORS[seg],backgroundColor:RFM_SEG_COLORS[seg]+'22',pointRadius:3};
      })
    },
    options:{responsive:true,maintainAspectRatio:false,
      scales:{r:{min:0,max:5,ticks:{stepSize:1,font:{size:10}}}},
      plugins:{legend:{position:'bottom',labels:{boxWidth:10,font:{size:11}}}}}
  });
  const rfmTb = document.querySelector('#rfm-tbl tbody');
  D.rfm.top.forEach((r,i)=>{
    const c2 = RFM_SEG_COLORS[r.segment]||'#6b7280';
    const tr = document.createElement('tr');
    tr.innerHTML=`<td>${i+1}</td><td><strong>${r.name||'-'}</strong></td><td>${r.rep||'-'}</td>`+
      `<td style="text-align:center;font-weight:700;color:${r.R>=4?'#16a34a':'#dc2626'}">${r.R}</td>`+
      `<td style="text-align:center;font-weight:700;color:${r.F>=4?'#16a34a':'#dc2626'}">${r.F}</td>`+
      `<td style="text-align:center;font-weight:700;color:${r.M>=4?'#16a34a':'#dc2626'}">${r.M}</td>`+
      `<td class="amount">${Math.round(r.monetary/1e4).toLocaleString()}만원</td>`+
      `<td><span style="font-size:11px;font-weight:600;color:${c2}">${r.segment}</span></td>`;
    rfmTb.appendChild(tr);
  });
}

// ── ABC ────────────────────────────────────────────────────────────────────
if (D.abc && D.abc.summary && Object.keys(D.abc.summary).length) {
  const ABC_COLORS = {A:'#16a34a',B:'#2563eb',C:'#9ca3af'};
  const ABC_DESC = {A:'매출 상위 80%',B:'매출 80~95%',C:'매출 하위 5%'};
  const abcSum = document.getElementById('abc-summary');
  ['A','B','C'].forEach(g=>{
    const s = D.abc.summary[g]; if (!s) return;
    const c2 = ABC_COLORS[g];
    const d = document.createElement('div');
    d.style.cssText=`border:2px solid ${c2};padding:16px`;
    d.innerHTML=`<div style="font-size:20px;font-weight:700;color:${c2}">등급 ${g}</div>`+
      `<div style="font-size:12px;color:var(--muted);margin:4px 0">${ABC_DESC[g]}</div>`+
      `<div style="font-size:24px;font-weight:700">${s.count.toLocaleString()}<span style="font-size:12px;color:var(--muted)">개사 (${s.pct_cust}%)</span></div>`+
      `<div style="font-size:13px;color:${c2};font-weight:600">${(s.revenue/1e8).toFixed(1)}억원 (${s.pct_rev}%)</div>`;
    abcSum.appendChild(d);
  });
  const abcTb = document.querySelector('#abc-tbl tbody');
  D.abc.top_a.forEach(r=>{
    const tr = document.createElement('tr');
    tr.innerHTML=`<td><span style="font-weight:700;color:#16a34a">A</span></td>`+
      `<td><strong>${r.name||'-'}</strong></td><td>${r.rep||'-'}</td>`+
      `<td class="amount">${Math.round(r.revenue/1e4).toLocaleString()}만원</td>`;
    abcTb.appendChild(tr);
  });
  if (D.abc.pareto && D.abc.pareto.labels) {
    const gradeColor = g => g==='A'?'#16a34a':g==='B'?'#2563eb':'#9ca3af';
    new Chart(document.getElementById('paretoChart'),{
      type:'bar',
      data:{
        labels: D.abc.pareto.labels,
        datasets:[
          {type:'bar',label:'매출(M)',data:D.abc.pareto.revenue,
           backgroundColor:D.abc.pareto.grades.map(g=>gradeColor(g)+'cc'),
           yAxisID:'y',order:2},
          {type:'line',label:'누적(%)',data:D.abc.pareto.cum_pct,
           borderColor:'#f59e0b',backgroundColor:'transparent',
           yAxisID:'y2',tension:.3,pointRadius:0,borderWidth:2,order:1},
        ]
      },
      options:{responsive:true,maintainAspectRatio:false,
        plugins:{legend:{position:'top',labels:{boxWidth:10,font:{size:11}}}},
        scales:{
          y:{ticks:{callback:v=>v+'M'},grid:{color:'#f0f0f0'}},
          y2:{position:'right',min:0,max:100,ticks:{callback:v=>v+'%'},grid:{display:false}},
          x:{ticks:{display:false},grid:{display:false}}
        }
      }
    });
  }
}

// ── 구매패턴 ────────────────────────────────────────────────────────────────
if (D.purchase_pattern) {
  // 친화도 테이블
  const affTb = document.querySelector('#affinity-tbl tbody');
  (D.purchase_pattern.affinity||[]).forEach(r=>{
    const tr = document.createElement('tr');
    const pct = Math.min(r.affinity, 100);
    tr.innerHTML=`<td><strong>${r.cat_a}</strong></td><td>${r.cat_b}</td>`+
      `<td style="text-align:center">${r.co_count}개사</td>`+
      `<td><span style="font-weight:600;color:var(--primary)">${r.affinity}%</span>`+
      `<div class="gap-bar"><div class="gap-fill" style="width:${pct}%;background:var(--primary)"></div></div></td>`;
    affTb.appendChild(tr);
  });

  // 연도별 카테고리 매출 차트
  const yrCatData = D.purchase_pattern.yr_cat_rev || {};
  const yrKeys = Object.keys(yrCatData);
  if (yrKeys.length > 0) {
    const allCats2 = [...new Set(yrKeys.flatMap(y=>yrCatData[y].map(r=>r.cat)))].slice(0,10);
    const YR_COLORS2 = ['#d1d5db','#94a3b8','#2563eb','#7c3aed'];
    new Chart(document.getElementById('catYrChart'),{
      type:'bar',
      data:{
        labels: allCats2,
        datasets: yrKeys.map((yr,i)=>({
          label: yr.slice(2)+'년',
          data: allCats2.map(cat=>{ const r=yrCatData[yr].find(x=>x.cat===cat); return r?r.total/1e6:0; }),
          backgroundColor: YR_COLORS2[i]||'#888',
          borderRadius:2,
        }))
      },
      options:{responsive:true,maintainAspectRatio:false,
        plugins:{legend:{position:'top',labels:{boxWidth:10,font:{size:11}}}},
        scales:{
          y:{ticks:{callback:v=>v.toFixed(0)+'M'},grid:{color:'#f0f0f0'}},
          x:{grid:{display:false},ticks:{font:{size:11}}}
        }
      }
    });
  }
}

// ── 업종별 분석 ──────────────────────────────────────────────────────────────
if (D.industry && D.industry.length > 0) {
  const totalRev = D.industry.reduce((a, r) => a + r.revenue, 0) || 1;
  const totalCnt = D.industry.reduce((a, r) => a + r.count, 0) || 1;
  new Chart(document.getElementById('industryChart'), {
    type: 'doughnut',
    data: {
      labels: D.industry.map(r => r.name),
      datasets: [{
        data: D.industry.map(r => r.revenue),
        backgroundColor: D.industry.map(r => r.color || '#888'),
        borderWidth: 2,
        borderColor: '#fff',
      }]
    },
    options: {
      responsive: true, maintainAspectRatio: false,
      plugins: {
        legend: { position: 'right', labels: { boxWidth: 12, font: { size: 12 }, padding: 10 } },
        tooltip: {
          callbacks: {
            label: ctx => ` ${ctx.label}: ${Math.round(ctx.raw/1e6).toLocaleString()}M (${(ctx.raw/totalRev*100).toFixed(1)}%)`
          }
        }
      }
    }
  });
  const indTb = document.querySelector('#industry-tbl tbody');
  D.industry.forEach(r => {
    const revPct = (r.revenue / totalRev * 100).toFixed(1);
    const barW = Math.min(100, parseFloat(revPct));
    const c2 = r.color || '#888';
    const tr = document.createElement('tr');
    tr.innerHTML =
      `<td><span style="display:inline-block;width:9px;height:9px;background:${c2};border-radius:2px;margin-right:6px;vertical-align:middle"></span><strong>${r.name}</strong></td>` +
      `<td style="text-align:right;color:var(--muted)">${r.count.toLocaleString()}개사</td>` +
      `<td class="amount" style="text-align:right">${Math.round(r.revenue/1e6).toFixed(1)}M</td>` +
      `<td style="text-align:right;color:var(--muted)">${revPct}%</td>` +
      `<td style="min-width:80px"><div class="gap-bar"><div class="gap-fill" style="width:${barW}%;background:${c2}"></div></div></td>`;
    indTb.appendChild(tr);
  });
}

// Modal
function showModal(c){
  document.querySelector('#overlay .modal').classList.remove('wide');
  document.getElementById('m-title').textContent=c.name;
  document.getElementById('m-body').innerHTML=
    (c.company_hint ? row('추정 기업명',`<span style="color:var(--primary);font-weight:600">${c.company_hint}</span>`) : '')+
    row('영업담당',c.rep)+row('지역',c.region)+
    row('주소',c.addr||'-')+
    (c.mobile ? row('핸드폰', `<a href="tel:${c.mobile}" style="color:var(--primary)">${c.mobile}</a>`) : '')+
    row((c.prev_yr_label||prevYrLabel)+' 총 매출',Math.round(c.total_prev/1e4).toLocaleString()+'만원')+
    row('주문 건수',c.order_count_prev+'건')+
    row('구매 카테고리',(c.cats||[]).join(', ')||'-')+
    row('마지막 주문',c.last_order)+
    `<div class="mrow"><span class="lbl">액션 제안</span><span class="action-badge">재활성화 연락 필요</span></div>`;
  document.getElementById('overlay').classList.add('open');
}
function row(label,val){return`<div class="mrow"><span class="lbl">${label}</span><span>${val}</span></div>`;}
function closeModal(){
  const ov = document.getElementById('overlay');
  ov.classList.remove('open');
  ov.querySelector('.modal').classList.remove('wide');
}
</script>
</body>
</html>"""

# ─────────────────────────── 메인 ───────────────────────────────────────────

def main():
    args = sys.argv[1:]
    path_25 = path_26 = None

    # ── 파일 수집: ERP(연도별) / 플레이오토(기간별) 분리 ──────────────────────
    year_paths   = {}   # {year: Path}  — ERP
    playauto_paths = [] # [Path]        — 플레이오토

    def _sniff_format(path):
        """첫 줄 헤더만 읽어서 포맷 감지"""
        try:
            for enc in ('utf-8-sig', 'cp949', 'utf-8'):
                try:
                    header = pd.read_csv(path, nrows=0, dtype=str, encoding=enc)
                    return 'playauto' if _is_playauto(header) else 'erp'
                except Exception:
                    continue
        except Exception:
            pass
        return 'erp'

    if HAS_PANDAS:
        here = Path(__file__).parent
        for d in [here / 'data', here]:
            if not d.exists(): continue
            for p in sorted(d.glob('*')):
                if p.suffix.lower() not in ('.xlsx', '.xls', '.csv'): continue
                fmt = _sniff_format(p)
                if fmt == 'playauto':
                    playauto_paths.append(p)
                else:
                    stem = p.stem
                    for yr in [2024, 2025, 2026, 2027, 2023]:
                        yr_str_4 = str(yr)
                        yr_str_2 = yr_str_4[2:]
                        if (yr_str_4 in stem or
                                (stem.startswith(yr_str_2) and re.match(r'^\d{4}', stem))):
                            if yr not in year_paths:
                                year_paths[yr] = p
                            break

    if HAS_PANDAS and (len(year_paths) >= 2 or (year_paths and playauto_paths)):
        dfs = {}

        # 1) ERP 파일 로드
        for yr in sorted(year_paths):
            p = year_paths[yr]
            print(f"[ERP {yr}] {p.name}")
            df = load_file(p)
            print(f"   {len(df):,}건 로드")
            dfs[yr] = df

        # 2) 플레이오토 파일 로드 → ERP에 없는 연도만 추가
        if playauto_paths:
            print(f"\n[플레이오토] {len(playauto_paths)}개 파일 처리 중...")
            pa_by_year = defaultdict(list)
            for p in sorted(playauto_paths):
                print(f"  {p.name}")
                df = load_file(p)
                df['_yr'] = df['order_date'].dt.year.where(df['order_date'].notna(), 0).astype(int)
                for yr, grp in df.groupby('_yr'):
                    if yr < 2000: continue
                    pa_by_year[yr].append(grp.drop(columns=['_yr'], errors='ignore'))

            for yr in sorted(pa_by_year):
                if yr in dfs:
                    print(f"  [{yr}] ERP 자료 있음 → 플레이오토 건너뜀")
                    continue
                merged = pd.concat(pa_by_year[yr], ignore_index=True)
                print(f"  [{yr}] 플레이오토 {len(merged):,}건 추가")
                dfs[yr] = merged

        data = analyze_real(dfs)
    else:
        if HAS_PANDAS:
            print("[주의] 데이터 파일을 찾지 못했습니다. 데모 모드로 실행합니다.")
            print("   data/ 폴더에 연도명 포함된 CSV/엑셀을 넣으세요.")
        customers, o25, o26 = build_demo_data()
        data = analyze_demo(customers, o25, o26)

    data['build_date'] = datetime.now().strftime('%Y-%m-%d %H:%M')

    # HTML JSON에서 대용량 export 전용 키 제거
    html_data = {k: v for k, v in data.items() if not k.startswith('_')}
    html = HTML.replace('__DATA_JSON__', json.dumps(html_data, ensure_ascii=False))
    OUTPUT_FILE.write_text(html, encoding='utf-8')

    # 거래처_2.csv 생성 (기업 추정 포함 전체 거래처 목록)
    if data.get('mode') == 'real' and HAS_PANDAS:
        _export_customer2(data, Path(__file__).parent / '거래처_2.csv')
        action_path = Path(__file__).parent / f"sales_action_list_{datetime.now().strftime('%Y%m%d')}.xlsx"
        _export_action_list(data, action_path)

    k = data['kpi']
    print(f"\n[완료] 대시보드 생성: {OUTPUT_FILE.resolve()}")
    print(f"   전체 {k['total']}개사  충성 {k['loyal']}  이탈 {k['churn']}  크로스셀 {k['crosssell']}  업셀 {k['upsell']}")

if __name__ == '__main__':
    main()
