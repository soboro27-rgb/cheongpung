"""
두 파일 비교 분석:
  - 카카오뱅크견적서송부_20260511.xlsx : 이번 7월 투찰 예정 단가
  - 2512_대상 장비 세부 목록...xlsx    : 2025년 12월 유찰 이력
"""
import openpyxl
from collections import defaultdict

# ── 카카오뱅크 파일 파싱 ─────────────────────────────────────
# col[14]=순번, col[17]=제조사, col[18]=모델명, col[20]=구매일자
# col[21]=장부가액, col[23]=매입단가(투찰단가), col[24]=CPU
wb1 = openpyxl.load_workbook(
    'data/카카오뱅크견적서송부_20260511.xlsx', data_only=True)
ws1 = wb1.active

new_items = []
for row in ws1.iter_rows(min_row=15, values_only=True):
    seq = row[14] if len(row) > 14 else None
    if not isinstance(seq, int):
        continue
    mfr   = str(row[17]).strip() if len(row) > 17 and row[17] else ''
    model = str(row[18]).strip() if len(row) > 18 and row[18] else ''
    mdate = row[20]              if len(row) > 20 else None
    book  = row[21]              if len(row) > 21 else None
    bid   = row[23]              if len(row) > 23 else None
    cpu   = str(row[24]).strip() if len(row) > 24 and row[24] else ''
    if not model:
        continue
    year = str(mdate)[:4] if mdate else ''
    new_items.append({
        'seq': seq, 'mfr': mfr, 'model': model,
        'year': year, 'book': book, 'bid': bid, 'cpu': cpu
    })

# 모델별 집계
new_sum = defaultdict(lambda: {
    'cnt': 0, 'mfr': '', 'bids': set(),
    'books': set(), 'years': set(), 'cpu': ''
})
for it in new_items:
    d = new_sum[it['model']]
    d['cnt']  += 1
    d['mfr']   = it['mfr']
    if it['bid']:  d['bids'].add(it['bid'])
    if it['book']: d['books'].add(it['book'])
    if it['year']: d['years'].add(it['year'])
    if it['cpu'] and not d['cpu']: d['cpu'] = it['cpu']

# ── 과거(2512) 파일 파싱 ─────────────────────────────────────
# col[0]=No, col[1]=제조사, col[2]=품명(모델명), col[3]=특이사항, col[4]=매입금액
wb2 = openpyxl.load_workbook(
    'data/2512_대상 장비 세부 목록_2025 전산단말기 정리 및 매각(수정)(1) 발송1.xlsx',
    data_only=True)
ws2 = wb2.worksheets[0]

past_items = []
for row in ws2.iter_rows(min_row=3, values_only=True):
    if not row or row[2] is None:
        continue
    model = str(row[2]).strip()
    price = row[4] if isinstance(row[4], (int, float)) else None
    if model and price:
        past_items.append({'model': model, 'price': price})

past_sum = defaultdict(lambda: {'cnt': 0, 'prices': []})
for it in past_items:
    past_sum[it['model']]['cnt']    += 1
    past_sum[it['model']]['prices'].append(it['price'])

# ────────────────────────────────────────────────────────────
# 출력 1: 이번 카카오뱅크 투찰 목록
# ────────────────────────────────────────────────────────────
total_bid  = sum(it['bid']  for it in new_items if it['bid'])
total_book = sum(it['book'] for it in new_items if it['book'])

print("=" * 125)
print(" [카카오뱅크] 이번 7월 투찰 예정 목록")
print("=" * 125)
print(f" 총 {len(new_items)}대  |  장부가 합계: {total_book:>15,.0f}원  |  투찰 합계: {total_bid:>13,.0f}원")
print("=" * 125)
print(f"  {'#':<4} {'모델명':<38} {'제조사':<8} {'대수':>4} {'장부가(원)':>13} {'투찰단가(원)':>12}  {'구매연도':<10} CPU 사양")
print("-" * 125)
for i, (model, d) in enumerate(
        sorted(new_sum.items(), key=lambda x: -x[1]['cnt']), 1):
    bids  = ' / '.join(f"{int(v):,}" for v in sorted(d['bids']))
    books = ' / '.join(f"{int(v):,}" for v in sorted(d['books']))
    years = ', '.join(sorted(d['years']))
    cpu   = d['cpu'][:42] if d['cpu'] else '-'
    print(f"  {i:<4} {model:<38} {d['mfr']:<8} {d['cnt']:>4} "
          f"{books:>13} {bids:>12}  {years:<10} {cpu}")
print("=" * 125)

# ────────────────────────────────────────────────────────────
# 출력 2: 과거(12월) 유찰 이력
# ────────────────────────────────────────────────────────────
print()
print("=" * 80)
print(" [2512 과거] 2025년 12월 유찰 이력")
print("=" * 80)
print(f"  {'모델명':<40} {'대수':>4} {'유찰단가(원)'}")
print("-" * 80)
for model, d in sorted(past_sum.items(), key=lambda x: -x[1]['cnt']):
    prices = sorted(set(d['prices']))
    pstr = ' / '.join(f"{int(p):,}" for p in prices)
    print(f"  {model:<40} {d['cnt']:>4}  {pstr}")
print("=" * 80)

# ────────────────────────────────────────────────────────────
# 출력 3: 모델별 교차 비교
# ────────────────────────────────────────────────────────────
print()
print("=" * 120)
print(" [비교] 과거 유찰단가  vs  이번 투찰예정단가  (★ = 주의 필요)")
print("=" * 120)
print(f"  {'모델명':<38} {'대수':>4} {'과거유찰(원)':>14} {'이번투찰(원)':>13}  {'차이(원)':>11}  판단")
print("-" * 120)

matched = unmatched = 0
for model, d in sorted(new_sum.items(), key=lambda x: -x[1]['cnt']):
    new_bid = min(d['bids']) if d['bids'] else None
    past    = past_sum.get(model)

    new_str = f"{int(new_bid):,}" if new_bid else '-'

    if past and past['prices']:
        matched += 1
        past_min = min(past['prices'])
        past_str = ' / '.join(f"{int(p):,}" for p in sorted(set(past['prices'])))
        if new_bid:
            diff = new_bid - past_min
            diff_str = f"{int(diff):+,}"
            if new_bid < past_min:
                judge = "★ 이번이 낮음 (하향 투찰)"
            elif new_bid == past_min:
                judge = "= 동일"
            else:
                judge = "▲ 이번이 높음"
        else:
            diff_str, judge = '-', '투찰단가 미입력'
        print(f"  {model:<38} {d['cnt']:>4} {past_str:>14} {new_str:>13}  {diff_str:>11}  {judge}")
    else:
        unmatched += 1
        print(f"  {model:<38} {d['cnt']:>4} {'(과거없음)':>14} {new_str:>13}  {'':>11}  신규 모델")

print("=" * 120)
print(f"  총 {len(new_sum)}개 모델  |  과거 매칭: {matched}개  |  신규: {unmatched}개")
print()
print("  ※ 과거 유찰은 '해당 단가로 입찰했지만 탈락'을 의미합니다.")
print("  ※ 낙찰받으려면 과거 유찰단가보다 높게 써야 가능성이 높아집니다.")
