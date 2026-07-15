#!/usr/bin/env python3
"""
자산매각 입찰 단가 산출 도구 (Estimate Tool)

사용법:
  python estimate.py             # data/ 폴더 파일로 견적서 자동 생성 (또는 템플릿 생성)
  python estimate.py --template  # 입력 템플릿만 새로 생성
"""

import sys
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl 설치 필요: pip install openpyxl")
    sys.exit(1)

BASE_DIR  = Path(__file__).resolve().parent
DATA_DIR  = BASE_DIR / "data"
OUT_DIR   = BASE_DIR / "output"
DATA_DIR.mkdir(exist_ok=True)
OUT_DIR.mkdir(exist_ok=True)

# ── 색상·스타일 상수 ──────────────────────────────────────────
GREEN       = "005B30"
BLUE_DARK   = "1565C0"
YELLOW_DARK = "F57F17"
C_WHITE     = "FFFFFF"

def hfill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def hfont(hex_color="000000", bold=False, size=11):
    return Font(color=hex_color, bold=bold, size=size)

H_FILL  = hfill(GREEN)        # 기본 헤더 (녹색)
B_FILL  = hfill(BLUE_DARK)    # 참조 컬럼 헤더 (파랑)
Y_FILL  = hfill(YELLOW_DARK)  # 입력 컬럼 헤더 (노랑)
REF_BG  = hfill("E3F2FD")     # 참조값 행 배경 (연파랑)
INP_BG  = hfill("FFF9C4")     # 입력 셀 배경 (연노랑)
WIN_BG  = hfill("E8F5E9")     # 낙찰 행 배경
LOSE_BG = hfill("FFF3E0")     # 유찰 행 배경
ALT_BG  = hfill("F5F5F5")     # 짝수 행 배경

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT   = Alignment(horizontal="left",   vertical="center")

CATEGORIES = ["PC", "노트북", "서버", "태블릿", "모바일", "프린터", "복합기", "기타전산기기", "기타"]

NUM_FMT   = '#,##0'
DATE_FMT  = 'YYYY-MM-DD'


# ── 헬퍼 ──────────────────────────────────────────────────────
def col_val(row, idx):
    if idx is None or idx >= len(row) or row[idx] is None:
        return ""
    return str(row[idx]).strip()

def col_num(row, idx, default=0):
    if idx is None or idx >= len(row) or row[idx] is None:
        return default
    try:
        return float(row[idx])
    except (ValueError, TypeError):
        return default

def find_col(headers, *keywords):
    for kw in keywords:
        for i, h in enumerate(headers):
            if kw in h:
                return i
    return None

def set_header(ws, row, col, value, fill, font_color=C_WHITE, size=11):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = fill
    c.font = Font(color=font_color, bold=True, size=size)
    c.alignment = CENTER
    return c

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


# ── 템플릿 생성 ────────────────────────────────────────────────
def create_templates():
    # 1. 과거 입찰 데이터 템플릿
    p = DATA_DIR / "과거입찰데이터.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "입찰이력"

    cols = [
        ("입찰일",       14),
        ("고객사명",     18),
        ("카테고리",     14),
        ("모델명",       28),
        ("제조사",       14),
        ("제조연도",     10),
        ("수량",          8),
        ("상태",          8),
        ("투찰단가(원)", 16),
        ("낙찰여부",     10),
        ("비고",         28),
    ]
    for ci, (h, w) in enumerate(cols, 1):
        set_header(ws, 1, ci, h, H_FILL)
        set_col_width(ws, ci, w)
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    ex = hfill("F0F7F2")
    samples = [
        ["2025-12-01", "새마을금고중앙회", "PC",    "OptiPlex 3090",       "Dell",   2020, 10, "중", 80000,  "유찰", ""],
        ["2025-12-01", "새마을금고중앙회", "PC",    "ProDesk 400 G7",      "HP",     2021,  8, "중", 90000,  "유찰", ""],
        ["2025-12-01", "새마을금고중앙회", "노트북","EliteBook 840 G7",    "HP",     2020,  5, "중", 150000, "유찰", ""],
        ["2025-12-01", "새마을금고중앙회", "노트북","ThinkPad E14 Gen2",   "Lenovo", 2021,  3, "중", 130000, "유찰", ""],
        ["2025-12-01", "새마을금고중앙회", "프린터","LaserJet Pro M404n",  "HP",     2019,  4, "하", 20000,  "유찰", ""],
        ["2025-12-01", "새마을금고중앙회", "복합기","WorkCentre 6515",     "Xerox",  2018,  2, "하", 30000,  "유찰", ""],
    ]
    for ri, row in enumerate(samples, 2):
        for ci, v in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.fill = ex
            cell.alignment = CENTER if ci not in [2, 4, 5] else LEFT
            if ci == 9:
                cell.number_format = NUM_FMT

    # 작성요령 시트
    ws2 = wb.create_sheet("작성요령")
    ws2["A1"] = "과거입찰데이터 작성 요령"
    ws2["A1"].font = Font(bold=True, size=13)
    notes = [
        ("입찰일",       "입찰 참여 날짜 (예: 2025-12-01)"),
        ("고객사명",     "입찰 발주 고객사명"),
        ("카테고리",     f"다음 중 하나: {', '.join(CATEGORIES)}"),
        ("모델명",       "장비 모델명 (예: OptiPlex 3090)"),
        ("제조사",       "제조사명 (예: Dell, HP, Lenovo)"),
        ("제조연도",     "4자리 연도 (예: 2020)"),
        ("수량",         "대수 (정수)"),
        ("상태",         "상 / 중 / 하"),
        ("투찰단가(원)", "대당 투찰 금액(원). 예: 80000"),
        ("낙찰여부",     "낙찰 또는 유찰"),
        ("비고",         "특이사항 (자유 입력)"),
    ]
    for i, (f, d) in enumerate(notes, 3):
        ws2[f"A{i}"] = f
        ws2[f"A{i}"].font = Font(bold=True)
        ws2[f"B{i}"] = d
    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 55

    wb.save(p)
    print(f"  [생성] {p.name}")

    # 2. 신규 자산 리스트 템플릿
    p2 = DATA_DIR / "신규자산리스트.xlsx"
    wb2 = openpyxl.Workbook()
    ws3 = wb2.active
    ws3.title = "자산목록"
    cols2 = [
        ("순번",     6),
        ("카테고리", 14),
        ("모델명",   28),
        ("제조사",   14),
        ("제조연도", 10),
        ("수량",      8),
        ("상태",      8),
        ("비고",     28),
    ]
    for ci, (h, w) in enumerate(cols2, 1):
        set_header(ws3, 1, ci, h, H_FILL)
        set_col_width(ws3, ci, w)
    ws3.row_dimensions[1].height = 22
    ws3.freeze_panes = "A2"

    for ri in range(2, 7):
        ws3.cell(ri, 1, ri - 1).alignment = CENTER

    wb2.save(p2)
    print(f"  [생성] {p2.name}")

    print()
    print("템플릿 생성 완료!")
    print(f"  → {DATA_DIR}")
    print()
    print("다음 단계:")
    print("  1. data/과거입찰데이터.xlsx 에 과거 입찰 이력 입력")
    print("  2. data/신규자산리스트.xlsx 에 고객사 자산 목록 입력")
    print("  3. python estimate.py 실행 → output/ 에 견적서 생성")


# ── 데이터 로드 ─────────────────────────────────────────────────
def load_history(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]

    ci_date   = find_col(headers, "입찰일")
    ci_cust   = find_col(headers, "고객사")
    ci_cat    = find_col(headers, "카테고리")
    ci_model  = find_col(headers, "모델명")
    ci_mfr    = find_col(headers, "제조사")
    ci_year   = find_col(headers, "제조연도", "연도")
    ci_qty    = find_col(headers, "수량")
    ci_cond   = find_col(headers, "상태")
    ci_price  = find_col(headers, "투찰단가", "단가")
    ci_result = find_col(headers, "낙찰")
    ci_note   = find_col(headers, "비고")

    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        price = col_num(row, ci_price)
        if price <= 0:
            continue
        year_val = col_num(row, ci_year)
        items.append({
            "date":       col_val(row, ci_date),
            "customer":   col_val(row, ci_cust),
            "category":   col_val(row, ci_cat),
            "model":      col_val(row, ci_model),
            "manufacturer": col_val(row, ci_mfr),
            "year":       int(year_val) if year_val else None,
            "quantity":   int(col_num(row, ci_qty, 1)),
            "condition":  col_val(row, ci_cond) or "중",
            "unit_price": price,
            "result":     col_val(row, ci_result),
            "note":       col_val(row, ci_note),
        })
    return items


def load_assets(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]

    ci_seq   = find_col(headers, "순번")
    ci_cat   = find_col(headers, "카테고리")
    ci_model = find_col(headers, "모델명")
    ci_mfr   = find_col(headers, "제조사")
    ci_year  = find_col(headers, "제조연도", "연도")
    ci_qty   = find_col(headers, "수량")
    ci_cond  = find_col(headers, "상태")
    ci_note  = find_col(headers, "비고")

    items = []
    for ri, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        cat = col_val(row, ci_cat)
        if not cat:
            continue
        year_val = col_num(row, ci_year)
        qty = int(col_num(row, ci_qty, 1)) or 1
        seq_val = col_num(row, ci_seq)
        items.append({
            "seq":          int(seq_val) if seq_val else ri - 1,
            "category":     cat,
            "model":        col_val(row, ci_model),
            "manufacturer": col_val(row, ci_mfr),
            "year":         int(year_val) if year_val else None,
            "quantity":     qty,
            "condition":    col_val(row, ci_cond) or "중",
            "note":         col_val(row, ci_note),
        })
    return items


# ── 통계 계산 ──────────────────────────────────────────────────
def price_stats(history, category, condition=None):
    """카테고리(+상태) 기준 과거 단가 통계 반환"""
    hits = [h for h in history if h["category"] == category]
    if condition:
        c_hits = [h for h in hits if h["condition"] == condition]
        if c_hits:
            hits = c_hits

    if not hits:
        return None

    prices = [h["unit_price"] for h in hits]
    win    = [h["unit_price"] for h in hits if "낙찰" in h["result"]]
    lose   = [h["unit_price"] for h in hits if "유찰" in h["result"]]

    def avg(lst): return round(sum(lst) / len(lst)) if lst else None

    return {
        "min":        int(min(prices)),
        "max":        int(max(prices)),
        "avg":        avg(prices),
        "count":      len(hits),
        "win_avg":    avg(win),
        "lose_avg":   avg(lose),
        "win_count":  len(win),
        "lose_count": len(lose),
    }


# ── 견적서 생성 ────────────────────────────────────────────────
def generate_estimate(assets, history, out_path):
    wb = openpyxl.Workbook()

    # ── 시트1: 견적입력 ────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "견적입력"
    ws1.freeze_panes = "H2"
    ws1.sheet_view.showGridLines = True

    # 헤더 정의 (fill, label, width)
    hdr1 = [
        (H_FILL, "순번",         6),
        (H_FILL, "카테고리",    14),
        (H_FILL, "모델명",      28),
        (H_FILL, "제조사",      14),
        (H_FILL, "제조연도",    10),
        (H_FILL, "수량",         8),
        (H_FILL, "상태",         8),
        (B_FILL, "과거최저(원)", 14),   # H
        (B_FILL, "과거최고(원)", 14),   # I
        (B_FILL, "과거평균(원)", 14),   # J
        (B_FILL, "낙찰평균(원)", 14),   # K
        (B_FILL, "유찰평균(원)", 14),   # L
        (B_FILL, "참조건수",     9),    # M
        (Y_FILL, "★ 투찰단가",  16),   # N  ← 사용자 입력
        (Y_FILL, "소계(원)",     16),   # O
        (H_FILL, "비고",         28),   # P
    ]
    for ci, (fill, label, width) in enumerate(hdr1, 1):
        set_header(ws1, 1, ci, label, fill)
        set_col_width(ws1, ci, width)
    ws1.row_dimensions[1].height = 24

    # 안내 메모
    note_cell = ws1.cell(1, 18, "※ N열(★투찰단가)에 단가를 입력하면 O열 소계와 최종견적서가 자동 완성됩니다.")
    note_cell.font = Font(color="C62828", bold=True, size=10)
    note_cell.alignment = LEFT
    ws1.column_dimensions["R"].width = 60

    n_col = 14   # N: 투찰단가 입력
    o_col = 15   # O: 소계

    for ri, asset in enumerate(assets, 2):
        st = price_stats(history, asset["category"], asset["condition"])
        bg = ALT_BG if ri % 2 == 0 else None

        vals = [
            asset["seq"],
            asset["category"],
            asset["model"],
            asset["manufacturer"],
            asset["year"],
            asset["quantity"],
            asset["condition"],
            st["min"]        if st else None,   # H
            st["max"]        if st else None,   # I
            st["avg"]        if st else None,   # J
            st["win_avg"]    if st else None,   # K
            st["lose_avg"]   if st else None,   # L
            st["count"]      if st else 0,      # M
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws1.cell(ri, ci, v)
            cell.alignment = LEFT if ci in [3, 4, 16] else CENTER
            if bg and ci < n_col:
                cell.fill = bg
            if ci in range(8, 14):
                cell.fill = REF_BG
                if v is not None and ci != 13:
                    cell.number_format = NUM_FMT

        # N열: 투찰단가 입력 셀
        inp = ws1.cell(ri, n_col)
        inp.fill = INP_BG
        inp.alignment = CENTER
        inp.number_format = NUM_FMT

        # O열: 소계 수식
        qty_ref   = f"F{ri}"
        price_ref = f"N{ri}"
        sub = ws1.cell(ri, o_col, f'=IF({price_ref}="","",{qty_ref}*{price_ref})')
        sub.fill = INP_BG
        sub.alignment = CENTER
        sub.number_format = NUM_FMT

        # P열: 비고
        ws1.cell(ri, 16, asset["note"]).alignment = LEFT

    # 합계 행
    last = len(assets) + 1
    total_r = last + 1
    ws1.cell(total_r, n_col - 1, "합  계").font = Font(bold=True, size=11)
    ws1.cell(total_r, n_col - 1).alignment = CENTER
    tot = ws1.cell(total_r, o_col, f"=SUM(O2:O{last})")
    tot.font = Font(bold=True, size=12, color=GREEN)
    tot.number_format = NUM_FMT
    tot.fill = WIN_BG
    tot.alignment = CENTER
    ws1.row_dimensions[total_r].height = 26

    # ── 시트2: 과거데이터참조 ──────────────────────────────────
    ws2 = wb.create_sheet("과거데이터참조")
    ws2.freeze_panes = "A2"

    hdr2 = [
        ("입찰일",       14),
        ("고객사",       18),
        ("카테고리",     14),
        ("모델명",       28),
        ("제조사",       14),
        ("제조연도",     10),
        ("수량",          8),
        ("상태",          8),
        ("투찰단가(원)", 16),
        ("낙찰여부",     10),
        ("비고",         28),
    ]
    for ci, (h, w) in enumerate(hdr2, 1):
        set_header(ws2, 1, ci, h, H_FILL)
        set_col_width(ws2, ci, w)
    ws2.row_dimensions[1].height = 22

    sorted_hist = sorted(history, key=lambda h: (h["category"], h["condition"], -h["unit_price"]))
    for ri, h in enumerate(sorted_hist, 2):
        bg = WIN_BG if "낙찰" in h["result"] else LOSE_BG
        row_vals = [h["date"], h["customer"], h["category"], h["model"],
                    h["manufacturer"], h["year"], h["quantity"],
                    h["condition"], h["unit_price"], h["result"], h["note"]]
        for ci, v in enumerate(row_vals, 1):
            cell = ws2.cell(ri, ci, v)
            cell.fill = bg
            cell.alignment = LEFT if ci in [2, 4, 5] else CENTER
            if ci == 9:
                cell.number_format = NUM_FMT

    # 카테고리별 통계 요약표
    gap = len(sorted_hist) + 3
    ws2.cell(gap, 1, "▼ 카테고리·상태별 단가 요약").font = Font(bold=True, size=11, color=GREEN)
    gap += 1

    sum_hdrs = ["카테고리", "상태", "건수", "최저단가", "최고단가", "평균단가", "낙찰평균", "유찰평균", "낙찰건", "유찰건"]
    sum_widths = [14, 8, 7, 13, 13, 13, 13, 13, 8, 8]
    for ci, (h, w) in enumerate(zip(sum_hdrs, sum_widths), 1):
        set_header(ws2, gap, ci, h, hfill("37474F"))
        ws2.column_dimensions[get_column_letter(ci)].width = w
    gap += 1

    for cat in sorted(set(h["category"] for h in history)):
        for cond in ["상", "중", "하"]:
            st = price_stats(history, cat, cond)
            if not st or st["count"] == 0:
                continue
            row_d = [cat, cond, st["count"], st["min"], st["max"], st["avg"],
                     st["win_avg"] or "-", st["lose_avg"] or "-",
                     st["win_count"], st["lose_count"]]
            for ci, v in enumerate(row_d, 1):
                cell = ws2.cell(gap, ci, v)
                cell.alignment = CENTER
                if ci in [4, 5, 6]:
                    cell.number_format = NUM_FMT
                    cell.fill = REF_BG
                if ci == 7 and isinstance(v, int):
                    cell.number_format = NUM_FMT
                    cell.fill = WIN_BG
                if ci == 8 and isinstance(v, int):
                    cell.number_format = NUM_FMT
                    cell.fill = LOSE_BG
            gap += 1

    # ── 시트3: 최종견적서 ──────────────────────────────────────
    ws3 = wb.create_sheet("최종견적서")
    ws3.sheet_view.showGridLines = False
    ws3.sheet_view.showRowColHeaders = False

    # 제목
    ws3.merge_cells("B2:J2")
    ws3["B2"] = "IT 자산 매각 투찰 견적서"
    ws3["B2"].font = Font(bold=True, size=18, color=C_WHITE)
    ws3["B2"].fill = H_FILL
    ws3["B2"].alignment = CENTER
    ws3.row_dimensions[2].height = 42

    # 견적 정보
    ws3.row_dimensions[3].height = 8
    ws3.row_dimensions[4].height = 20
    ws3["B4"] = f"견적일자: {datetime.now().strftime('%Y년 %m월 %d일')}"
    ws3["B4"].font = Font(size=11)
    ws3["H4"] = "담당자:"
    ws3["H4"].font = Font(size=11)
    ws3.row_dimensions[5].height = 8

    # 테이블 헤더 (row 6)
    q_hdrs = [
        ("순번",         6),
        ("카테고리",    14),
        ("모델명",      28),
        ("제조사",      14),
        ("제조연도",    10),
        ("수량",         8),
        ("상태",         8),
        ("투찰단가(원)", 16),
        ("소계(원)",    16),
        ("비고",        24),
    ]
    COL_OFFSET = 2   # B열부터 시작
    for ci, (h, w) in enumerate(q_hdrs, COL_OFFSET):
        set_header(ws3, 6, ci, h, H_FILL, size=11)
        ws3.column_dimensions[get_column_letter(ci)].width = w
    ws3.row_dimensions[6].height = 24
    ws3.column_dimensions["A"].width = 2   # 좌측 여백

    for ri, asset in enumerate(assets, 7):
        sheet1_row = ri - 7 + 2   # 견적입력 시트의 해당 행
        bg = ALT_BG if ri % 2 == 1 else None

        static_vals = [
            asset["seq"],
            asset["category"],
            asset["model"],
            asset["manufacturer"],
            asset["year"],
            asset["quantity"],
            asset["condition"],
        ]
        for ci, v in enumerate(static_vals, COL_OFFSET):
            cell = ws3.cell(ri, ci, v)
            cell.alignment = LEFT if ci in [COL_OFFSET + 2, COL_OFFSET + 3] else CENTER
            if bg:
                cell.fill = bg

        # 투찰단가: 견적입력 시트 N열 참조
        price_cell = ws3.cell(ri, COL_OFFSET + 7)
        price_cell.value = f"=견적입력!N{sheet1_row}"
        price_cell.number_format = NUM_FMT
        price_cell.alignment = CENTER
        if bg: price_cell.fill = bg

        # 소계
        sub_ci = COL_OFFSET + 8
        f_qty   = get_column_letter(COL_OFFSET + 5)
        f_price = get_column_letter(COL_OFFSET + 7)
        sub_cell = ws3.cell(ri, sub_ci)
        sub_cell.value = f'=IF({f_price}{ri}="","",{f_qty}{ri}*{f_price}{ri})'
        sub_cell.number_format = NUM_FMT
        sub_cell.alignment = CENTER
        if bg: sub_cell.fill = bg

        # 비고
        ws3.cell(ri, COL_OFFSET + 9, asset["note"]).alignment = LEFT

        ws3.row_dimensions[ri].height = 19

    # 합계 행
    last3 = len(assets) + 6
    total3_r = last3 + 1
    ws3.row_dimensions[total3_r].height = 32
    tot_start = get_column_letter(COL_OFFSET)
    tot_end   = get_column_letter(COL_OFFSET + 7)
    ws3.merge_cells(f"{tot_start}{total3_r}:{tot_end}{total3_r}")
    ws3.cell(total3_r, COL_OFFSET, "합  계").font = Font(bold=True, size=13)
    ws3.cell(total3_r, COL_OFFSET).alignment = CENTER
    ws3.cell(total3_r, COL_OFFSET).fill = WIN_BG

    sub_col_letter = get_column_letter(COL_OFFSET + 8)
    grand = ws3.cell(total3_r, COL_OFFSET + 8)
    grand.value = f"=SUM({sub_col_letter}7:{sub_col_letter}{last3})"
    grand.font = Font(bold=True, size=15, color=GREEN)
    grand.number_format = NUM_FMT
    grand.alignment = CENTER
    grand.fill = WIN_BG

    wb.save(out_path)
    print(f"  [생성] {out_path.name}")
    print()
    print("사용 방법:")
    print("  1. '견적입력' 시트 열기")
    print("  2. H~M열(파란색) 과거 단가 참고")
    print("  3. N열(★투찰단가, 노란색)에 단가 직접 입력")
    print("  4. '최종견적서' 시트에서 완성된 견적서 확인 및 출력")


# ── 메인 ──────────────────────────────────────────────────────
def main():
    args = sys.argv[1:]

    if "--template" in args:
        print("입력 템플릿 생성 중...")
        create_templates()
        return

    history_path = DATA_DIR / "과거입찰데이터.xlsx"
    asset_path   = DATA_DIR / "신규자산리스트.xlsx"

    # 직접 파일 지정
    if len(args) >= 2:
        asset_path   = Path(args[0])
        history_path = Path(args[1])
    elif len(args) == 1:
        asset_path = Path(args[0])

    # 파일 없으면 템플릿 생성
    if not asset_path.exists() or not history_path.exists():
        print("입력 파일이 없어 템플릿을 생성합니다...")
        create_templates()
        return

    print(f"신규 자산 리스트 : {asset_path.name}")
    print(f"과거 입찰 데이터 : {history_path.name}")

    history = load_history(history_path)
    assets  = load_assets(asset_path)

    if not assets:
        print("오류: 신규 자산 리스트가 비어있습니다.")
        return

    print(f"  → 신규 자산 {len(assets)}건 / 과거 이력 {len(history)}건 로드")
    print()

    out_name = f"견적서_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    out_path = OUT_DIR / out_name
    generate_estimate(assets, history, out_path)
    print(f"완료: output/{out_name}")


if __name__ == "__main__":
    main()
