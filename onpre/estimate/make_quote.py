"""
카카오뱅크 자산매각 입찰 견적서 생성
- 과거 유찰 단가 비교
- 네이버쇼핑 / 다나와 시세 자동 조회 (크롤링)
- ★최종단가 입력 칸 포함
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
from datetime import datetime
import requests
from bs4 import BeautifulSoup
import time
import json
import re

# ── 설정 ─────────────────────────────────────────────────────
FETCH_MARKET_PRICE = True   # False 로 바꾸면 시세 조회 생략 (빠른 재생성용)
DELAY_SEC = 0.8             # 요청 간격 (초)

# ── 스타일 ───────────────────────────────────────────────────
def fill(hex_c):
    return PatternFill(start_color=hex_c, end_color=hex_c, fill_type="solid")

GREEN    = fill("005B30")
BLUE_H   = fill("1565C0")
PURPLE_H = fill("6A1B9A")
YELLOW_H = fill("F57F17")
RED_H    = fill("B71C1C")
ALT_BG   = fill("F5F5F5")
WARN_BG  = fill("FFF3E0")
OK_BG    = fill("E8F5E9")
INP_BG   = fill("FFF9C4")
REF_BG   = fill("E3F2FD")
MKT_BG   = fill("F3E5F5")

def thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")

def hdr(ws, r, c, val, bg, sz=10, width=None):
    cell = ws.cell(r, c, val)
    cell.fill = bg
    cell.font = Font(color="FFFFFF", bold=True, size=sz)
    cell.alignment = CENTER
    if width:
        ws.column_dimensions[get_column_letter(c)].width = width

# ── 시세 조회 ─────────────────────────────────────────────────
_BROWSER = {
    'User-Agent': (
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
        'AppleWebKit/537.36 (KHTML, like Gecko) '
        'Chrome/124.0.0.0 Safari/537.36'
    ),
    'Accept-Language': 'ko-KR,ko;q=0.9,en-US;q=0.8',
}
_cache = {}


def _extract_prices(text_list, lo=1000, hi=100_000_000):
    prices = []
    for t in text_list:
        t = re.sub(r'[^\d]', '', t)
        if t:
            try:
                p = int(t)
                if lo < p < hi:
                    prices.append(p)
            except Exception:
                pass
    return prices


def danawa_price(model: str, mfr: str = ''):
    """다나와 검색 HTML 파싱 → (최저가, 상위5개 가격 목록) 반환
    노트북/PC 기준: 100,000원 미만은 부품으로 간주하여 제외"""
    key = f'dw_{mfr}_{model}'
    if key in _cache:
        return _cache[key]
    result = (None, [])
    try:
        query = f'{mfr} {model}'.strip() if mfr else model
        url = f'https://search.danawa.com/dsearch.php?query={requests.utils.quote(query)}&limit=5'
        r = requests.get(url, headers=_BROWSER, timeout=9)
        soup = BeautifulSoup(r.text, 'html.parser')
        prices = []
        for tag in soup.select('.price_sect')[:8]:
            nums = re.findall(r'[\d,]+', tag.get_text())
            for n in nums:
                try:
                    v = int(n.replace(',', ''))
                    if 100_000 <= v < 100_000_000:   # 10만원 이상만 유효
                        prices.append(v)
                        break
                except Exception:
                    pass
        if prices:
            result = (min(prices), sorted(set(prices))[:5])
    except Exception:
        pass
    _cache[key] = result
    time.sleep(DELAY_SEC)
    return result


def naver_link(model: str) -> str:
    """네이버 쇼핑 검색 URL 반환 (스크래핑 불가 — 링크만 제공)"""
    return f'https://search.shopping.naver.com/search/all?query={requests.utils.quote(model)}'


# ── 데이터 로드 ───────────────────────────────────────────────
import glob, os

def _find_files():
    """data/ 폴더에서 과거이력 파일과 신규자산 파일 자동 감지"""
    all_xlsx = glob.glob('data/*.xlsx')
    past_file = new_file = None
    for f in all_xlsx:
        bn = os.path.basename(f)
        # 과거이력 파일: 이름에 '과거' / '이력' / '2512' 포함 or 날짜가 과거(2025)
        if any(k in bn for k in ('2512', '과거', '이력', '유찰')):
            past_file = f
        else:
            # 신규자산 파일: 나머지 중 가장 최근 수정본
            if new_file is None or os.path.getmtime(f) > os.path.getmtime(new_file):
                new_file = f
    return new_file, past_file

_new_file, _past_file = _find_files()
if not _new_file:
    raise FileNotFoundError("data/ 폴더에 신규 자산 엑셀 파일이 없습니다.")
if not _past_file:
    raise FileNotFoundError("data/ 폴더에 과거 이력 엑셀 파일이 없습니다. (파일명에 '2512' 또는 '과거'가 포함되어야 합니다)")

print("[ 1/3 ] 엑셀 데이터 로드 중...")
print(f"        신규자산: {os.path.basename(_new_file)}")
print(f"        과거이력: {os.path.basename(_past_file)}")

wb1 = openpyxl.load_workbook(_new_file, data_only=True)
ws_src = wb1.active

new_items = []
for row in ws_src.iter_rows(min_row=15, values_only=True):
    seq = row[14] if len(row) > 14 else None
    if not isinstance(seq, int):
        continue
    mfr   = str(row[17]).strip() if len(row) > 17 and row[17] else ''
    model = str(row[18]).strip() if len(row) > 18 and row[18] else ''
    mdate = row[20] if len(row) > 20 else None
    book  = row[21] if len(row) > 21 else None
    bid   = row[23] if len(row) > 23 else None
    cpu   = str(row[24]).strip() if len(row) > 24 and row[24] else ''
    if not model:
        continue
    year = str(mdate)[:4] if mdate else ''
    new_items.append({'seq': seq, 'mfr': mfr, 'model': model,
                      'year': year, 'book': book, 'bid': bid, 'cpu': cpu})

new_sum = defaultdict(lambda: {
    'cnt': 0, 'mfr': '', 'bids': set(),
    'books': set(), 'years': set(), 'cpu': ''
})
for it in new_items:
    d = new_sum[it['model']]
    d['cnt']  += 1
    d['mfr']   = it['mfr']
    if it['bid']:  d['bids'].add(it['bid'])
    if it['book']: d['books'].add(it['book'])
    if it['year']: d['years'].add(it['year'])
    if it['cpu'] and not d['cpu']: d['cpu'] = it['cpu']

wb2 = openpyxl.load_workbook(_past_file, data_only=True)
past_sum = defaultdict(lambda: {'cnt': 0, 'prices': []})
for row in wb2.worksheets[0].iter_rows(min_row=3, values_only=True):
    if not row or row[2] is None:
        continue
    m = str(row[2]).strip()
    p = row[4] if isinstance(row[4], (int, float)) else None
    if m and p:
        past_sum[m]['cnt']    += 1
        past_sum[m]['prices'].append(p)

sorted_models = sorted(new_sum.items(), key=lambda x: -x[1]['cnt'])
print(f"        카카오뱅크 {len(new_items)}대 / {len(sorted_models)}종 로드 완료")

# ── 시세 조회 ─────────────────────────────────────────────────
mkt_prices = {}   # model → (naver_price, danawa_price)

if FETCH_MARKET_PRICE:
    total = len(sorted_models)
    print(f"\n[ 2/3 ] 다나와 시세 조회 중 ({total}개 모델)...")
    for i, (model, _) in enumerate(sorted_models, 1):
        mfr_name = new_sum[model]['mfr']
        dw_min, dw_list = danawa_price(model, mfr_name)
        nv_url = naver_link(model)
        mkt_prices[model] = (dw_min, dw_list, nv_url)
        dw_str = f"{dw_min:>12,.0f}원" if dw_min else "         없음"
        print(f"  [{i:>2}/{total}] {model:<38}  다나와:{dw_str}")
else:
    print("\n[ 2/3 ] 시세 조회 생략 (FETCH_MARKET_PRICE=False)")
    for model, _ in sorted_models:
        mkt_prices[model] = (None, [], naver_link(model))

# ── 엑셀 생성 ────────────────────────────────────────────────
print("\n[ 3/3 ] 견적서 엑셀 생성 중...")
wb = openpyxl.Workbook()

# ════════════════════════════════════════════════════════════
# 시트 1: 투찰단가 수정
#
# 열 배치:
#  A  순번      G  초록(자산정보)
#  B  제조사
#  C  모델명
#  D  대수
#  E  구매연도
#  F  장부가
#  G  CPU사양
#  H  과거유찰단가   } 파란색 (과거이력)
#  I  과거대수       }
#  J  낙찰여부       }
#  K  네이버최저가   } 보라색 (온라인 시세)
#  L  다나와시세      }
#  M  현재예정가    } 노란색 (투찰단가)
#  N  ★최종단가    }  ← 사용자 입력
#  O  판단          빨간색
# ════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "투찰단가 수정"
ws1.freeze_panes = "A3"
ws1.sheet_view.showGridLines = False

# 행1 그룹 헤더
groups = [
    ("A1:G1", "카카오뱅크 자산 정보",       GREEN),
    ("H1:J1", "과거 유찰 이력 (2025.12)",   BLUE_H),
    ("K1:L1", "온라인 시세 참고",           PURPLE_H),
    ("M1:N1", "★ 이번 투찰 단가",          YELLOW_H),
    ("O1:O1", "판단",                       RED_H),
]
for rng, label, bg in groups:
    ws1.merge_cells(rng)
    c = ws1[rng.split(':')[0]]
    c.value = label
    c.fill = bg
    c.font = Font(color="FFFFFF", bold=True, size=11)
    c.alignment = CENTER
ws1.row_dimensions[1].height = 24

# 행2 세부 헤더
COL_DEF = [
    # (label, width, bg)
    ("순번",            5,  GREEN),
    ("제조사",          9,  GREEN),
    ("모델명",         36,  GREEN),
    ("대수",            6,  GREEN),
    ("구매연도",       10,  GREEN),
    ("장부가(원)",     14,  GREEN),
    ("CPU 사양",       30,  GREEN),
    ("과거유찰(원)",   14,  BLUE_H),
    ("과거대수",        8,  BLUE_H),
    ("낙찰여부",        8,  BLUE_H),
    ("다나와최저(원)", 14,  PURPLE_H),
    ("네이버쇼핑링크", 18,  PURPLE_H),
    ("현재예정(원)",   14,  YELLOW_H),
    ("★최종단가(원)", 16,  YELLOW_H),
    ("판단",           22,  RED_H),
]
for ci, (label, width, bg) in enumerate(COL_DEF, 1):
    hdr(ws1, 2, ci, label, bg, sz=10, width=width)
ws1.row_dimensions[2].height = 22

# 데이터 행
for ri, (model, d) in enumerate(sorted_models, 3):
    past       = past_sum.get(model, {})
    past_min   = min(past.get('prices', [None]) or [None])
    past_cnt   = past.get('cnt', 0)
    new_bid    = min(d['bids']) if d['bids'] else None
    dw_p, _dw_list, nv_url = mkt_prices.get(model, (None, [], ''))

    if past_min and new_bid:
        diff = new_bid - past_min
        if diff < -50000:
            judge, row_bg = "★★ 크게 하향 — 검토 필요", WARN_BG
        elif diff < 0:
            judge, row_bg = "★ 하향 — 주의", WARN_BG
        elif diff == 0:
            judge, row_bg = "동일", None
        else:
            judge, row_bg = "상향", OK_BG
    elif not past_min:
        judge, row_bg = "신규 모델", None
    else:
        judge, row_bg = "단가 미입력", WARN_BG

    vals = [
        ri - 2,                              # A 순번
        d['mfr'],                            # B 제조사
        model,                               # C 모델명
        d['cnt'],                            # D 대수
        ', '.join(sorted(d['years'])),       # E 구매연도
        min(d['books']) if d['books'] else None,  # F 장부가
        d['cpu'],                            # G CPU
        past_min,                            # H 과거유찰
        past_cnt or None,                    # I 과거대수
        "유찰" if past_cnt else None,        # J 낙찰여부
        dw_p,                                # K 다나와 최저가
        "→ 네이버쇼핑 검색",                # L 네이버 링크 (하이퍼링크)
        new_bid,                             # M 현재예정
        None,                                # N ★최종단가 (사용자 입력)
        judge,                               # O 판단
    ]
    for ci, v in enumerate(vals, 1):
        cell = ws1.cell(ri, ci, v)
        cell.alignment = LEFT if ci in [3, 7, 15] else CENTER
        # 배경색
        if ci == 14:
            cell.fill = INP_BG
        elif ci in [8, 9, 10]:
            cell.fill = REF_BG
        elif ci in [11, 12]:
            cell.fill = MKT_BG
        elif row_bg:
            cell.fill = row_bg
        elif ri % 2 == 0:
            cell.fill = ALT_BG
        # 숫자 포맷
        if ci in [6, 8, 11, 13, 14]:
            cell.number_format = '#,##0'

    # L열(12): 네이버 쇼핑 하이퍼링크
    if nv_url:
        link_cell = ws1.cell(ri, 12)
        link_cell.hyperlink = nv_url
        link_cell.font = Font(color="1565C0", underline="single", size=9)

    ws1.row_dimensions[ri].height = 18

# 합계 행
last_r = len(sorted_models) + 2
tot_r  = last_r + 1
ws1.cell(tot_r, 3, "합  계").font = Font(bold=True, size=11)
ws1.cell(tot_r, 4, sum(d['cnt'] for _, d in sorted_models)).font = Font(bold=True)
ws1.cell(tot_r, 4).alignment = CENTER
ws1.cell(tot_r, 6,  f"=SUMPRODUCT(D3:D{last_r},F3:F{last_r})").number_format = '#,##0'
ws1.cell(tot_r, 6).font  = Font(bold=True, color="005B30")
ws1.cell(tot_r, 14, f"=SUMPRODUCT(D3:D{last_r},N3:N{last_r})").number_format = '#,##0'
ws1.cell(tot_r, 14).font = Font(bold=True, color="005B30", size=12)
ws1.row_dimensions[tot_r].height = 28

ws1.cell(1, 17,
    "사용법: N열(★최종단가)에 단가 입력 → 최종견적서 자동 반영  "
    "| H열=12월 유찰단가  K열=다나와 신품 최저가  L열=네이버쇼핑 바로가기 링크"
)
ws1.cell(1, 17).font = Font(color="C62828", bold=True, size=10)
ws1.column_dimensions["Q"].width = 90

# ════════════════════════════════════════════════════════════
# 시트 2: 최종견적서
# ════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("최종견적서")
ws2.sheet_view.showGridLines = False

ws2.merge_cells("B2:K2")
ws2["B2"] = "IT 자산 매각 투찰 견적서 (카카오뱅크)"
ws2["B2"].fill = GREEN
ws2["B2"].font = Font(color="FFFFFF", bold=True, size=16)
ws2["B2"].alignment = CENTER
ws2.row_dimensions[2].height = 40

ws2["B3"] = f"견적일자: {datetime.now().strftime('%Y년 %m월 %d일')}"
ws2["B3"].font = Font(size=11)
ws2["H3"] = "작성: Re.New.All PC"
ws2["H3"].font = Font(size=11)
ws2.row_dimensions[3].height = 20
ws2.row_dimensions[4].height = 8

Q_HDRS = [
    ("순번",           5),
    ("제조사",         9),
    ("모델명",        36),
    ("대수",           6),
    ("구매연도",      10),
    ("장부가(원)",    14),
    ("CPU 사양",      28),
    ("투찰단가(원)",  16),
    ("소계(원)",      16),
    ("비고",          14),
]
OFF = 2
for ci, (h, w) in enumerate(Q_HDRS, OFF):
    hdr(ws2, 5, ci, h, GREEN, sz=11, width=w)
    ws2.row_dimensions[5].height = 22
ws2.column_dimensions["A"].width = 2

for ri, (model, d) in enumerate(sorted_models, 6):
    ws1_row = ri - 6 + 3
    bg = ALT_BG if ri % 2 == 0 else None

    for ci, v in enumerate([
        ri - 5,
        d['mfr'],
        model,
        d['cnt'],
        ', '.join(sorted(d['years'])),
        min(d['books']) if d['books'] else None,
        d['cpu'],
    ], OFF):
        cell = ws2.cell(ri, ci, v)
        cell.alignment = LEFT if ci in [OFF+2, OFF+6] else CENTER
        if bg: cell.fill = bg
        if ci == OFF + 5: cell.number_format = '#,##0'

    # 투찰단가: 시트1 N열(14) 참조
    pc = ws2.cell(ri, OFF + 7)
    pc.value = f"=투찰단가수정!N{ws1_row}"
    pc.number_format = '#,##0'
    pc.alignment = CENTER
    if bg: pc.fill = bg

    # 소계
    qty_c = get_column_letter(OFF + 3)
    pri_c = get_column_letter(OFF + 7)
    sc = ws2.cell(ri, OFF + 8)
    sc.value = f'=IF({pri_c}{ri}="","",{qty_c}{ri}*{pri_c}{ri})'
    sc.number_format = '#,##0'
    sc.alignment = CENTER
    if bg: sc.fill = bg

    ws2.row_dimensions[ri].height = 19

last_q = len(sorted_models) + 5
tot_q  = last_q + 1
ws2.row_dimensions[tot_q].height = 32
ws2.merge_cells(f"{get_column_letter(OFF)}{tot_q}:{get_column_letter(OFF+7)}{tot_q}")
ws2.cell(tot_q, OFF, "합  계").font = Font(bold=True, size=13)
ws2.cell(tot_q, OFF).alignment = CENTER
ws2.cell(tot_q, OFF).fill = fill("E8F5E9")
sub_col = get_column_letter(OFF + 8)
grand = ws2.cell(tot_q, OFF + 8)
grand.value = f"=SUM({sub_col}6:{sub_col}{last_q})"
grand.font   = Font(bold=True, size=15, color="005B30")
grand.number_format = '#,##0'
grand.alignment = CENTER
grand.fill = fill("E8F5E9")

# ════════════════════════════════════════════════════════════
# 시트 3: 개별 자산 전체 목록
# ════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("전체자산목록")
ws3.freeze_panes = "A2"

for ci, (h, w) in enumerate([
    ("순번", 6), ("제조사", 9), ("모델명", 36),
    ("구매연도", 10), ("장부가(원)", 14), ("현예정단가", 12), ("CPU 사양", 32),
], 1):
    hdr(ws3, 1, ci, h, GREEN, sz=11, width=w)
ws3.row_dimensions[1].height = 22

for ri, it in enumerate(new_items, 2):
    bg = ALT_BG if ri % 2 == 0 else None
    for ci, v in enumerate(
        [it['seq'], it['mfr'], it['model'], it['year'],
         it['book'], it['bid'], it['cpu']], 1
    ):
        cell = ws3.cell(ri, ci, v)
        cell.alignment = LEFT if ci in [3, 7] else CENTER
        if bg: cell.fill = bg
        if ci in [5, 6]: cell.number_format = '#,##0'
    ws3.row_dimensions[ri].height = 17

# ── 저장 ─────────────────────────────────────────────────────
out = f"output/카카오뱅크_견적서_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
wb.save(out)

total_book = sum(it['book'] for it in new_items if it['book'])
total_bid  = sum(it['bid']  for it in new_items if it['bid'])
print(f"\n완료: {out}")
print(f"  장부가 합계   : {total_book:>15,.0f} 원")
print(f"  현재 예정 합계: {total_bid:>15,.0f} 원")
print(f"  (장부가 대비  : {total_bid/total_book*100:.1f}%)")
print()
print("  N열(★최종단가)에 단가 입력 → 최종견적서 자동 반영")
