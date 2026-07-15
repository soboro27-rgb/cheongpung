import re, os, io, base64
from collections import defaultdict
from datetime import datetime, date
from functools import wraps

from flask import (Flask, render_template, request, redirect, url_for,
                   flash, send_file, jsonify)
from flask_login import (LoginManager, login_user, logout_user,
                         login_required, current_user)
from models import db, User, PurchaseOrder, QRCode, Inspection, InspectionEdit
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from dotenv import load_dotenv
import qrcode

load_dotenv()

app = Flask(__name__)
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'wwm-dev-secret')
app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv('DATABASE_URL', 'sqlite:///wwm.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)

login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = '로그인이 필요합니다.'
login_manager.login_message_category = 'warning'


@login_manager.user_loader
def load_user(uid):
    return User.query.get(int(uid))


# ── decorators ─────────────────────────────────────────────────────────────

def admin_required(f):
    @wraps(f)
    def dec(*a, **kw):
        if not current_user.is_admin:
            flash('관리자 권한이 필요합니다.', 'danger')
            return redirect(url_for('dashboard'))
        return f(*a, **kw)
    return dec


# ── QR helpers ─────────────────────────────────────────────────────────────

def make_qr_string(purchase_date, vendor_name_en, manager_last_en, seq):
    date_part = purchase_date.strftime('%Y_%m%d')
    vendor_part = re.sub(r'[^a-zA-Z0-9]', '', vendor_name_en.lower())[:8]
    manager_part = re.sub(r'[^a-zA-Z]', '', manager_last_en.lower())[:2]
    return f'{date_part}_{vendor_part}_{manager_part}_{seq:03d}'


def parse_qr_string(qr_string):
    parts = qr_string.strip().split('_')
    if len(parts) != 5:
        raise ValueError('QR 형식 오류')
    year, mmdd, vendor, manager, seq_s = parts
    purchase_date = datetime.strptime(year + mmdd, '%Y%m%d').date()
    return {'purchase_date': purchase_date, 'vendor': vendor,
            'manager': manager, 'seq': int(seq_s)}


def qr_to_png_b64(text):
    qr = qrcode.QRCode(box_size=5, border=2,
                       error_correction=qrcode.constants.ERROR_CORRECT_M)
    qr.add_data(text)
    qr.make(fit=True)
    img = qr.make_image(fill_color='black', back_color='white')
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    return base64.b64encode(buf.getvalue()).decode()


# ── auth ───────────────────────────────────────────────────────────────────

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        pw = request.form.get('password', '')
        user = User.query.filter_by(username=username, is_active=True).first()
        if user and user.check_password(pw):
            login_user(user, remember=True)
            return redirect(url_for('dashboard'))
        flash('아이디 또는 비밀번호가 올바르지 않습니다.', 'danger')
    return render_template('login.html')


@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))


# ── dashboard ──────────────────────────────────────────────────────────────

@app.route('/')
@login_required
def dashboard():
    if current_user.is_admin:
        orders = PurchaseOrder.query.order_by(
            PurchaseOrder.purchase_date.desc()).limit(10).all()
        total_qr = QRCode.query.count()
        locked = QRCode.query.filter_by(is_locked=True).count()
        unlocked = total_qr - locked
        today_cnt = Inspection.query.filter(
            db.func.date(Inspection.saved_at) == date.today()).count()
        return render_template('dashboard.html', active='dashboard',
                               orders=orders, total_qr=total_qr,
                               locked=locked, unlocked=unlocked,
                               today_cnt=today_cnt)
    else:
        my = (Inspection.query
              .filter_by(inspector_id=current_user.id)
              .order_by(Inspection.saved_at.desc())
              .limit(10).all())
        my_total = Inspection.query.filter_by(inspector_id=current_user.id).count()
        return render_template('dashboard.html', active='dashboard',
                               my=my, my_total=my_total)


# ── purchase orders ────────────────────────────────────────────────────────

@app.route('/orders')
@login_required
@admin_required
def order_list():
    orders = PurchaseOrder.query.order_by(
        PurchaseOrder.purchase_date.desc()).all()
    return render_template('order/list.html', active='orders', orders=orders)


@app.route('/orders/new', methods=['GET', 'POST'])
@login_required
@admin_required
def order_new():
    if request.method == 'POST':
        vendor_en = request.form.get('vendor_name_en', '').strip()
        manager_en = request.form.get('manager_last_en', '').strip()
        pdate = datetime.strptime(request.form['purchase_date'], '%Y-%m-%d').date()
        qty = int(request.form.get('quantity', 1))

        # auto order_no
        prefix = pdate.strftime('%Y%m%d')
        existing = PurchaseOrder.query.filter(
            PurchaseOrder.order_no.like(f'{prefix}%')).count()
        order_no = f'{prefix}-{existing+1:03d}'

        po = PurchaseOrder(
            order_no=order_no,
            vendor_name=request.form.get('vendor_name', ''),
            vendor_name_en=vendor_en,
            manager_last_en=manager_en,
            purchase_date=pdate,
            quantity=qty,
            memo=request.form.get('memo', ''),
            created_by_id=current_user.id,
        )
        db.session.add(po)
        db.session.flush()

        for i in range(1, qty + 1):
            qs = make_qr_string(pdate, vendor_en, manager_en, i)
            db.session.add(QRCode(
                qr_string=qs,
                purchase_order_id=po.id,
                seq=i,
            ))

        db.session.commit()
        flash(f'계약건 등록 완료 — {order_no} / QR {qty}매 생성', 'success')
        return redirect(url_for('order_detail', id=po.id))

    return render_template('order/form.html', active='orders',
                           today=date.today().isoformat())


@app.route('/orders/<int:id>')
@login_required
@admin_required
def order_detail(id):
    po = PurchaseOrder.query.get_or_404(id)
    qr_data = [(q, qr_to_png_b64(q.qr_string)) for q in po.qr_codes]
    return render_template('order/detail.html', active='orders',
                           po=po, qr_data=qr_data)


@app.route('/orders/<int:id>/labels')
@login_required
@admin_required
def order_labels(id):
    return redirect(url_for('label_print', order_id=id))


@app.route('/label/print')
@login_required
@admin_required
def label_print():
    sel_order = request.args.get('order_id', '', type=str)
    orders = PurchaseOrder.query.order_by(PurchaseOrder.purchase_date.desc()).all()
    labels = []
    if sel_order:
        po = PurchaseOrder.query.get_or_404(int(sel_order))
        for qr in po.qr_codes:
            if not qr.is_locked or not qr.inspection:
                continue
            insp = qr.inspection
            try:
                p = parse_qr_string(qr.qr_string)
            except Exception:
                p = {}
            labels.append({
                'qr_string': qr.qr_string,
                'qr_img': 'data:image/png;base64,' + qr_to_png_b64(qr.qr_string),
                'grade': insp.grade,
                'quality': insp.quality,
                'maker': insp.maker,
                'model_name': insp.model_name,
                'cpu': insp.cpu,
                'ram': insp.ram,
                'storage': insp.storage,
                'vendor_name': po.vendor_name,
                'purchase_date': str(p.get('purchase_date', '')),
            })
    return render_template('label/print.html', active='label',
                           orders=orders, labels=labels, sel_order=sel_order)


# ── inspect (inspector flow) ───────────────────────────────────────────────

@app.route('/scan')
@login_required
def scan():
    return render_template('inspect/scan.html', active='scan')


@app.route('/inspect/<qr_string>', methods=['GET', 'POST'])
@login_required
def inspect(qr_string):
    qr = QRCode.query.filter_by(qr_string=qr_string).first()
    if not qr:
        flash(f'QR을 찾을 수 없습니다: {qr_string}', 'danger')
        return redirect(url_for('scan'))

    # 락 상태 체크
    if qr.is_locked and not current_user.is_admin:
        return render_template('inspect/locked.html', qr=qr, active='scan')

    if request.method == 'POST':
        insp = qr.inspection or Inspection(qr_id=qr.id,
                                           inspector_id=current_user.id)

        def fbool(field, default=None):
            v = request.form.get(field, '')
            if v == 'O': return True
            if v == 'X': return False
            return default

        insp.maker = request.form.get('maker', '')
        insp.model_name = request.form.get('model_name', '')
        insp.cpu = request.form.get('cpu', '')
        insp.ram = request.form.get('ram', '')
        insp.storage = request.form.get('storage', '')
        insp.vga = request.form.get('vga', '')
        insp.screen_size = request.form.get('screen_size', '')
        insp.resolution = request.form.get('resolution', '')
        insp.lcd_condition = request.form.get('lcd_condition', '')
        insp.keyboard_touch = request.form.get('keyboard_touch', '')
        insp.battery_loss_rate = float(request.form.get('battery_loss_rate') or 0)
        insp.battery_45_checked = 'battery_45_checked' in request.form
        insp.notes = request.form.get('notes', '')
        insp.has_adapter = fbool('has_adapter')
        insp.is_barasi = fbool('is_barasi')
        insp.quality = request.form.get('quality', '')
        insp.grade = request.form.get('grade', '')
        price = request.form.get('purchase_price', '')
        insp.purchase_price = int(price) if price.strip() else None
        insp.saved_at = datetime.now()

        if not qr.inspection:
            db.session.add(insp)
        db.session.flush()

        # 즉시 락 설정
        qr.is_locked = True
        qr.locked_at = datetime.now()
        qr.locked_by_id = current_user.id
        db.session.commit()

        flash('검수 저장 완료. QR이 잠금 처리되었습니다.', 'success')
        return redirect(url_for('scan'))

    try:
        parsed = parse_qr_string(qr_string)
    except Exception:
        parsed = {}

    return render_template('inspect/form.html', active='scan',
                           qr=qr, parsed=parsed,
                           insp=qr.inspection)


# ── history (inspector's own) ──────────────────────────────────────────────

@app.route('/history')
@login_required
def history():
    if current_user.is_admin:
        return redirect(url_for('admin_inspections'))
    records = (Inspection.query
               .filter_by(inspector_id=current_user.id)
               .order_by(Inspection.saved_at.desc()).all())
    return render_template('history/list.html', active='history',
                           records=records)


# ── admin: inspections ────────────────────────────────────────────────────

@app.route('/admin/inspections')
@login_required
@admin_required
def admin_inspections():
    order_id = request.args.get('order_id', '', type=str)
    inspector_id = request.args.get('inspector_id', '', type=str)
    df = request.args.get('date_from', '')
    dt = request.args.get('date_to', '')
    quality = request.args.get('quality', '')

    query = Inspection.query.join(QRCode).join(PurchaseOrder, isouter=True)
    if order_id:
        query = query.filter(QRCode.purchase_order_id == int(order_id))
    if inspector_id:
        query = query.filter(Inspection.inspector_id == int(inspector_id))
    if df:
        query = query.filter(Inspection.saved_at >= datetime.strptime(df, '%Y-%m-%d'))
    if dt:
        query = query.filter(Inspection.saved_at <= datetime.strptime(dt + ' 23:59:59', '%Y-%m-%d %H:%M:%S'))
    if quality:
        query = query.filter(Inspection.quality == quality)

    records = query.order_by(Inspection.saved_at.desc()).all()
    orders = PurchaseOrder.query.order_by(PurchaseOrder.purchase_date.desc()).all()
    inspectors = User.query.filter_by(role='INSPECTOR').all()
    return render_template('admin/inspections.html', active='inspections',
                           records=records, orders=orders, inspectors=inspectors,
                           f_order=order_id, f_inspector=inspector_id,
                           f_df=df, f_dt=dt, f_quality=quality)


@app.route('/admin/inspections/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_edit(id):
    insp = Inspection.query.get_or_404(id)

    if request.method == 'POST':
        fields = {
            'maker': insp.maker, 'model_name': insp.model_name,
            'cpu': insp.cpu, 'ram': insp.ram, 'storage': insp.storage,
            'vga': insp.vga, 'screen_size': insp.screen_size,
            'resolution': insp.resolution, 'lcd_condition': insp.lcd_condition,
            'keyboard_touch': insp.keyboard_touch,
            'battery_loss_rate': str(insp.battery_loss_rate or ''),
            'notes': insp.notes or '',
            'quality': insp.quality or '', 'grade': insp.grade or '',
            'purchase_price': str(insp.purchase_price or ''),
        }

        def fbool(f, default=None):
            v = request.form.get(f, '')
            if v == 'O': return True
            if v == 'X': return False
            return default

        new_vals = {
            'maker': request.form.get('maker', ''),
            'model_name': request.form.get('model_name', ''),
            'cpu': request.form.get('cpu', ''),
            'ram': request.form.get('ram', ''),
            'storage': request.form.get('storage', ''),
            'vga': request.form.get('vga', ''),
            'screen_size': request.form.get('screen_size', ''),
            'resolution': request.form.get('resolution', ''),
            'lcd_condition': request.form.get('lcd_condition', ''),
            'keyboard_touch': request.form.get('keyboard_touch', ''),
            'battery_loss_rate': request.form.get('battery_loss_rate', ''),
            'notes': request.form.get('notes', ''),
            'quality': request.form.get('quality', ''),
            'grade': request.form.get('grade', ''),
            'purchase_price': request.form.get('purchase_price', ''),
        }

        # 수정이력 기록
        for f, old in fields.items():
            new = new_vals.get(f, '')
            if str(old or '') != str(new or ''):
                db.session.add(InspectionEdit(
                    inspection_id=insp.id,
                    editor_id=current_user.id,
                    field_name=f,
                    old_value=str(old or ''),
                    new_value=str(new or ''),
                ))

        # 실제 반영
        insp.maker = new_vals['maker']
        insp.model_name = new_vals['model_name']
        insp.cpu = new_vals['cpu']
        insp.ram = new_vals['ram']
        insp.storage = new_vals['storage']
        insp.vga = new_vals['vga']
        insp.screen_size = new_vals['screen_size']
        insp.resolution = new_vals['resolution']
        insp.lcd_condition = new_vals['lcd_condition']
        insp.keyboard_touch = new_vals['keyboard_touch']
        r = new_vals['battery_loss_rate']
        insp.battery_loss_rate = float(r) if r.strip() else None
        insp.battery_45_checked = 'battery_45_checked' in request.form
        insp.notes = new_vals['notes']
        insp.has_adapter = fbool('has_adapter')
        insp.is_barasi = fbool('is_barasi')
        insp.quality = new_vals['quality']
        insp.grade = new_vals['grade']
        p = new_vals['purchase_price']
        insp.purchase_price = int(p) if p.strip() else None

        db.session.commit()
        flash('수정 완료. 수정이력이 기록되었습니다.', 'success')
        return redirect(url_for('admin_inspections'))

    try:
        parsed = parse_qr_string(insp.qr_code.qr_string)
    except Exception:
        parsed = {}
    return render_template('admin/edit.html', active='inspections',
                           insp=insp, parsed=parsed)


# ── admin: missed (미완료 현황) ────────────────────────────────────────────

@app.route('/admin/pending')
@login_required
@admin_required
def admin_pending():
    order_id = request.args.get('order_id', '', type=str)
    query = QRCode.query.filter_by(is_locked=False)
    if order_id:
        query = query.filter(QRCode.purchase_order_id == int(order_id))
    qr_codes = (query.join(PurchaseOrder, isouter=True)
                .order_by(PurchaseOrder.purchase_date.desc()).all())
    orders = PurchaseOrder.query.order_by(PurchaseOrder.purchase_date.desc()).all()
    return render_template('admin/pending.html', active='pending',
                           qr_codes=qr_codes, orders=orders, f_order=order_id)


# ── admin: users ───────────────────────────────────────────────────────────

@app.route('/admin/users')
@login_required
@admin_required
def admin_users():
    users = User.query.order_by(User.created_at).all()
    return render_template('admin/users.html', active='users', users=users)


@app.route('/admin/users/new', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_user_new():
    if request.method == 'POST':
        username = request.form['username'].strip()
        if User.query.filter_by(username=username).first():
            flash('이미 사용 중인 아이디입니다.', 'danger')
        else:
            u = User(username=username,
                     name=request.form['name'],
                     role=request.form.get('role', 'INSPECTOR'))
            u.set_password(request.form['password'])
            db.session.add(u); db.session.commit()
            flash(f'{u.name} 계정 생성 완료', 'success')
            return redirect(url_for('admin_users'))
    return render_template('admin/user_form.html', active='users')


@app.route('/admin/users/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_user_edit(id):
    u = User.query.get_or_404(id)
    if request.method == 'POST':
        u.name = request.form['name']
        u.role = request.form.get('role', 'INSPECTOR')
        u.is_active = request.form.get('is_active', '0') == '1'
        pw = request.form.get('password', '').strip()
        if pw:
            u.set_password(pw)
        db.session.commit()
        flash(f'{u.name} 정보가 수정되었습니다.', 'success')
        return redirect(url_for('admin_users'))
    return render_template('admin/user_form.html', active='users', user=u)


@app.route('/admin/users/<int:id>/toggle', methods=['POST'])
@login_required
@admin_required
def admin_user_toggle(id):
    u = User.query.get_or_404(id)
    if u.id != current_user.id:
        u.is_active = not u.is_active
        db.session.commit()
        flash(f'{u.name} {"활성화" if u.is_active else "비활성화"}', 'success')
    return redirect(url_for('admin_users'))


@app.route('/admin/users/<int:id>/reset', methods=['POST'])
@login_required
@admin_required
def admin_user_reset(id):
    u = User.query.get_or_404(id)
    pw = request.form.get('new_password', '')
    if pw:
        u.set_password(pw); db.session.commit()
        flash(f'{u.name} 비밀번호 변경 완료', 'success')
    return redirect(url_for('admin_users'))


# ── export ─────────────────────────────────────────────────────────────────

@app.route('/export')
@login_required
@admin_required
def export_index():
    orders = PurchaseOrder.query.order_by(PurchaseOrder.purchase_date.desc()).all()
    return render_template('export/index.html', active='export', orders=orders)


@app.route('/export/excel')
@login_required
@admin_required
def export_excel():
    order_id = request.args.get('order_id', '')
    df = request.args.get('date_from', '')
    dt = request.args.get('date_to', '')
    records = Inspection.query.join(QRCode)
    if order_id:
        records = records.filter(QRCode.purchase_order_id == int(order_id))
    if df:
        records = records.filter(Inspection.saved_at >= datetime.strptime(df, '%Y-%m-%d'))
    if dt:
        records = records.filter(Inspection.saved_at <= datetime.strptime(dt + ' 23:59:59', '%Y-%m-%d %H:%M:%S'))
    records = records.order_by(Inspection.saved_at).all()

    wb = Workbook()
    ws = wb.active; ws.title = '검수데이터'
    headers = ['QR코드', '매입일자', '매입처', '담당자', '입고번호',
               '제조사', '모델명', 'CPU', 'RAM', '저장장치', 'VGA',
               '사이즈', '해상도', '액정상태', '키보드터치', '배터리손실률',
               '45%방전', '특이사항', '아답터', '바라시', '불량·양품',
               '등급', '매입가', '검수자', '저장일시']

    hf = PatternFill(fill_type='solid', fgColor='1E293B')
    hfont = Font(bold=True, color='FFFFFF', size=10)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci, h)
        c.fill = hf; c.font = hfont
        c.alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 22

    for ri, r in enumerate(records, 2):
        qr = r.qr_code
        try:
            p = parse_qr_string(qr.qr_string)
        except Exception:
            p = {}
        row = [
            qr.qr_string,
            p.get('purchase_date', ''), qr.order.vendor_name if qr.order else '',
            p.get('manager', ''), p.get('seq', ''),
            r.maker, r.model_name, r.cpu, r.ram, r.storage, r.vga,
            r.screen_size, r.resolution, r.lcd_condition, r.keyboard_touch,
            r.battery_loss_rate, 'O' if r.battery_45_checked else 'X',
            r.notes,
            'O' if r.has_adapter else 'X',
            'O' if r.is_barasi else 'X',
            r.quality, r.grade, r.purchase_price,
            r.inspector.name if r.inspector else '',
            r.saved_at.strftime('%Y-%m-%d %H:%M') if r.saved_at else '',
        ]
        for ci, v in enumerate(row, 1):
            ws.cell(ri, ci, v)

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f"검수데이터_{datetime.now().strftime('%Y%m%d%H%M')}.xlsx"
    return send_file(buf, download_name=fname, as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── init ───────────────────────────────────────────────────────────────────

def init_db():
    with app.app_context():
        db.create_all()
        if not User.query.filter_by(username='admin').first():
            a = User(username='admin', name='관리자', role='ADMIN')
            a.set_password('admin1234')
            db.session.add(a); db.session.commit()
            print('Admin created: admin / admin1234')


if __name__ == '__main__':
    init_db()
    app.run(debug=True, port=5050)
