"""
다음카페 중고IT 업체 단가 수집기 v4
- 브라우저 프로파일 저장 (browser_profile/) → 로그인 한 번만
- 카페 검색창(글쓴이) → 최근 6개월 게시글 → 구매/판매가 추출 → Excel
"""

import json
import re
import urllib.parse
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
from bs4 import BeautifulSoup
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from playwright.sync_api import Page, sync_playwright

# ── 설정 ─────────────────────────────────────────────────────────────
EMAIL    = "favorte@hanmail.net"
PASSWORD = "!cbj90436122"
CAFE_URL = "https://cafe.daum.net/usedexport"
PROFILE_DIR = Path(__file__).parent / "browser_profile"
PROFILE_DIR.mkdir(exist_ok=True)

VENDORS = [
    "에이스피씨", "블루트레이더스", "성일",
    "인디컴", "판다컴", "컴기공", "컴퓨터구조대",
]

VENDOR_COLORS = {
    "에이스피씨":    "DBEAFE",
    "블루트레이더스": "DCFCE7",
    "성일":          "FEF9C3",
    "인디컴":        "FFE4E6",
    "판다컴":        "EDE9FE",
    "컴기공":        "FFEDD5",
    "컴퓨터구조대":  "CCFBF1",
}

CUTOFF  = datetime.now() - timedelta(days=90)
OUT_DIR = Path(__file__).parent / "output"
OUT_DIR.mkdir(exist_ok=True)

# ── 날짜 유틸 ─────────────────────────────────────────────────────────

def parse_date(s: str) -> datetime | None:
    s = s.strip()
    now = datetime.now()
    for fmt in ("%Y.%m.%d", "%y.%m.%d", "%Y-%m-%d", "%m.%d", "%m/%d"):
        try:
            dt = datetime.strptime(s, fmt)
            if fmt in ("%m.%d", "%m/%d"):
                dt = dt.replace(year=now.year)
            return dt
        except ValueError:
            continue
    if re.match(r"^\d{1,2}:\d{2}", s):
        return now
    return None


def within_6m(date_str: str) -> bool:
    if not date_str:
        return True
    dt = parse_date(date_str)
    return dt is None or dt >= CUTOFF


# ── 프레임 유틸 ──────────────────────────────────────────────────────

def eval_all_frames(page: Page, js: str) -> list:
    results = []
    seen = set()
    for frame in page.frames:
        try:
            r = frame.evaluate(js)
            if not r:
                continue
            for item in (r if isinstance(r, list) else [r]):
                key = item.get("href", "") if isinstance(item, dict) else str(item)
                if key and key not in seen:
                    seen.add(key)
                    results.append(item)
        except Exception:
            pass
    return results


# ── 로그인 상태 확인 ─────────────────────────────────────────────────

def check_login_status(page: Page) -> bool:
    """내부 프레임의 로그인 버튼 텍스트로 로그인 상태 판단"""
    for frame in page.frames:
        if "_c21_" not in frame.url:
            continue
        try:
            result = frame.evaluate("""
                () => {
                    const el = document.querySelector('#loginout');
                    if (!el) return 'unknown';
                    return el.textContent.trim() === '로그인' ? 'out' : 'in';
                }
            """)
            if result == "in":
                return True
            if result == "out":
                return False
        except Exception:
            pass
    return False


# ── 로그인 ────────────────────────────────────────────────────────────

def login(page: Page):
    print("[로그인] 카페 접속 중...")
    page.goto(CAFE_URL, wait_until="domcontentloaded")
    page.wait_for_timeout(4000)

    if check_login_status(page):
        print("[로그인] 로그인됨 (저장된 세션)")
        return

    print("=" * 50)
    print("[로그인 필요] 브라우저에서 카카오 계정으로 로그인하세요.")
    print("  1. 화면의 '로그인' 버튼 클릭")
    print("  2. 카카오 이메일/비밀번호 입력 후 로그인")
    print("  3. 카페 메인 화면이 보이면 아래 Enter 누르기")
    print("=" * 50)
    input(">>> 로그인 완료 후 Enter: ")

    # 카카오 OAuth 리다이렉트가 끝날 때까지 대기
    try:
        page.wait_for_url(lambda url: "cafe.daum.net" in url, timeout=15000)
    except Exception:
        pass
    try:
        page.wait_for_load_state("domcontentloaded", timeout=10000)
    except Exception:
        pass
    page.wait_for_timeout(3000)

    print("[로그인] 완료 — 다음 실행부터는 자동 로그인됩니다.")


# ── 검색 ─────────────────────────────────────────────────────────────

_JS_EXTRACT_POSTS = r"""
() => {
    const results = [];
    const seen = new Set();

    // cafesearch 결과 테이블: td.subject / td.searchpreview_subject 안의 링크
    let anchors = Array.from(document.querySelectorAll(
        'td.subject a[href], td.searchpreview_subject a[href]'
    ));
    // fallback: datanum 포함 링크 또는 /usedexport/.../숫자 패턴
    if (anchors.length === 0) {
        document.querySelectorAll('a[href]').forEach(a => {
            const h = a.href || '';
            if (h.includes('datanum=') || h.match(/\/usedexport\/[A-Za-z0-9_]+\/\d+/))
                anchors.push(a);
        });
    }

    anchors.forEach(a => {
        const href = a.href || '';
        if (!href || seen.has(href)) return;
        if (!href.includes('datanum=') && !href.match(/\/usedexport\/[A-Za-z0-9_]+\/\d+/)) return;
        seen.add(href);

        const row = a.closest('tr.list_row_info') || a.closest('tr') || a.closest('li') || a.parentElement;

        let author = '';
        if (row) {
            const el = row.querySelector('.search_nick a, .td_name a, .nick a, .author');
            if (el) author = el.textContent.trim();
        }

        let date = '';
        if (row) {
            const el = row.querySelector('td.date, .td_date, .date');
            if (el) date = el.textContent.trim();
        }
        if (!date) {
            const txt = row ? row.textContent : a.textContent;
            const dm = txt.match(/(\d{4}[.]\d{2}[.]\d{2}|\d{2}[.]\d{2}[.]\d{2})/);
            if (dm) date = dm[0];
        }

        let boardName = '';
        if (row) {
            const next = row.nextElementSibling;
            if (next) {
                const el = next.querySelector('a[href*="bbs_list"]');
                if (el) boardName = el.textContent.trim();
            }
        }

        const fldM = href.match(/[?&]fldid=([^&]+)/);
        const board = fldM ? fldM[1] : '';

        results.push({
            href, board, boardName,
            title: a.textContent.trim().substring(0, 80),
            author, date,
        });
    });
    return results;
}
"""


def _goto_safe(page: Page, url: str):
    """이동 중 OAuth 리다이렉트 충돌 시 재시도"""
    for _att in range(3):
        try:
            page.goto(url, wait_until="domcontentloaded")
            return
        except Exception as e:
            if "interrupted by another navigation" in str(e) and _att < 2:
                try:
                    page.wait_for_load_state("domcontentloaded", timeout=10000)
                except Exception:
                    pass
                page.wait_for_timeout(1500)
            else:
                raise


def _debug_frame(frame, vendor: str, page=None):
    """cafesearch 프레임 HTML + 링크 덤프"""
    try:
        ddir = Path(__file__).parent / "debug"
        ddir.mkdir(exist_ok=True)
        safe = vendor[:10].replace("/", "_")
        if page:
            page.screenshot(path=str(ddir / f"search_{safe}.png"))
        html = frame.evaluate("() => document.documentElement.outerHTML")
        (ddir / f"search_{safe}.html").write_text(html, encoding="utf-8", errors="replace")
        all_links = frame.evaluate("""
            () => Array.from(document.querySelectorAll('a[href]'))
                .map(a => a.href || '').filter(h => h && !h.startsWith('javascript'))
                .slice(0, 40)
        """)
        post_links = [l for l in all_links if "usedexport" in l or "bbs_read" in l or "datanum" in l]
        if post_links:
            print(f"  [후보 링크] {post_links[:5]}")
        else:
            print(f"  [프레임 링크 샘플] {all_links[:8]}")
        print(f"  → debug/search_{safe}.html 저장됨")
    except Exception as e:
        print(f"  [디버그 실패] {e}")


def search_vendor(page: Page, vendor: str, grpid: str) -> list[dict]:
    """카페 iframe 내에서 제목 검색 → cafesearch 프레임 유지하며 수집"""
    posts = []
    seen_urls: set[str] = set()
    encoded = urllib.parse.quote(vendor)
    print(f"\n[검색] {vendor}")

    # 1. 카페 외부 래퍼 로드 (iframe 컨텍스트 확보)
    _goto_safe(page, CAFE_URL)
    page.wait_for_timeout(3000)

    # 2. _c21_ 내부 프레임에서 검색 폼 제출
    submitted = False
    for fr in page.frames:
        if "_c21_" not in fr.url:
            continue
        try:
            el = fr.query_selector("input[name='search_left_query']")
            if not el:
                continue
            el.fill(vendor)
            page.wait_for_timeout(300)
            el.press("Enter")
            page.wait_for_timeout(5000)
            submitted = True
            break
        except Exception as e:
            print(f"  [폼 제출 오류] {e}")

    if not submitted:
        print(f"  [오류] 검색 폼을 찾지 못함 — 프레임: {[f.url[:50] for f in page.frames]}")
        return posts

    # 3. cafesearch 프레임 확보
    search_frame = None
    for fr in page.frames:
        if "cafesearch" in fr.url:
            search_frame = fr
            break

    if not search_frame:
        print(f"  [오류] cafesearch 프레임 없음 — 프레임: {[f.url[:60] for f in page.frames]}")
        return posts

    search_base_url = re.sub(r"[&?]pagenum=\d+", "", search_frame.url)
    print(f"  검색 URL: {search_base_url[:100]}")

    # 4. 페이지 순회 (프레임 컨텍스트 유지)
    for page_no in range(1, 50):
        if page_no > 1:
            next_url = search_base_url + f"&pagenum={page_no}"
            try:
                search_frame.evaluate(f"location.href = {json.dumps(next_url)}")
                page.wait_for_timeout(2500)
                # 프레임 재참조 (navigate 후 frame 객체 갱신)
                for fr in page.frames:
                    if "cafesearch" in fr.url:
                        search_frame = fr
                        break
            except Exception:
                break

        items = search_frame.evaluate(_JS_EXTRACT_POSTS)
        print(f"  페이지 {page_no}: {len(items)}건 탐지")

        if page_no == 1 and len(items) == 0:
            _debug_frame(search_frame, vendor, page)
            break

        if not items:
            break

        hit_old = False
        added = 0
        for it in items:
            href = it.get("href", "")
            if not href or href in seen_urls:
                continue
            date_s = it.get("date", "")
            if not within_6m(date_s):
                hit_old = True
                continue
            seen_urls.add(href)
            # bbs_nsread / bbs_read → 직접 접근 가능한 표준 카페 URL로 변환
            fld_m = re.search(r'fldid=([^&]+)', href)
            num_m = re.search(r'datanum=(\d+)', href)
            if ("bbs_nsread" in href or "bbs_read" in href) and fld_m and num_m:
                full_url = f"https://cafe.daum.net/usedexport/{fld_m.group(1)}/{num_m.group(1)}"
            else:
                full_url = href if href.startswith("http") else f"https://cafe.daum.net{href}"
            posts.append({
                "vendor":    vendor,
                "board":     it.get("board", ""),
                "boardName": it.get("boardName", ""),
                "title":     it.get("title", ""),
                "url":       full_url,
                "date":      date_s,
            })
            added += 1

        print(f"  → {added}건 추가 (누적 {len(posts)}건)")
        if hit_old:
            print("  6개월 이전 게시글 발견, 종료")
            break

    return posts


# ── 게시글 파싱 ───────────────────────────────────────────────────────

def detect_trade_type(title: str, board_name: str) -> str:
    txt = title + " " + board_name
    if any(k in txt for k in ["구매", "매입", "삽니다", "구합"]):
        return "구매가"
    if any(k in txt for k in ["판매", "팝니다", "셀러", "팔아요"]):
        return "판매가"
    return "단가"


def extract_prices_from_text(text: str) -> list[dict]:
    rows = []
    current_section = ""
    SKIP_PREFIXES = ("상호", "지역", "전화", "소량", "부가세", "담당", "연락처", "주소")
    SKIP_EXACT = {"품목", "제품명", "모델", "규격", "단가", "금액", "항목", "분류", "품명", "수량", "비고"}

    for line in text.splitlines():
        line = line.strip()
        if not line or len(line) < 3:
            continue

        # 섹션 헤더: ---CPU---, ===메모리=== 등
        sect_m = re.match(r'^[-=*\s]{2,}(.{1,20}?)[-=*\s]{2,}$', line)
        if sect_m:
            sec = sect_m.group(1).strip().lstrip('-').rstrip('-').strip()
            if sec and not re.search(r'\d', sec):
                current_section = sec
            continue

        # 정보성 헤더 줄 건너뜀
        if any(line.startswith(p) for p in SKIP_PREFIXES):
            continue

        # 가격 패턴: 한국식 점 구분(395.000원, 80.000) 또는 쉼표(50,000원) 또는 순수 숫자(50000원)
        price_m = re.search(
            r'(\d{1,3}(?:[.,]\d{3})+|\d{5,})\s*원?(?=\s|$|[^0-9.,])',
            line
        )
        if not price_m:
            continue

        raw = price_m.group(1).replace(".", "").replace(",", "")
        if not raw.isdigit():
            continue
        price_val = int(raw)
        if not (500 < price_val < 100_000_000):
            continue

        # 수량 추출 (숫자+EA/개/pcs)
        qty = ""
        qty_m = re.search(r'(\d+)\s*(?:EA|ea|개|pcs)', line, re.IGNORECASE)
        if qty_m and qty_m.start() < price_m.start():
            qty = qty_m.group(1) + "EA"
            item_end = qty_m.start()
        else:
            item_end = price_m.start()

        item = line[:item_end].strip()
        item = re.sub(r'\s{2,}', ' ', item).rstrip('(').strip()
        if not item or len(item) < 2 or item in SKIP_EXACT:
            continue

        rows.append({
            "카테고리": current_section,
            "품목": item,
            "수량": qty,
            "단가": price_val,
            "비고": "",
        })
    return rows


def parse_post(page: Page, post: dict) -> list[dict]:
    trade_type = detect_trade_type(post["title"], post.get("boardName", ""))
    try:
        page.goto(post["url"], wait_until="domcontentloaded")
        page.wait_for_timeout(3000)
    except Exception:
        return []

    actual_author = post["vendor"]
    for frame in page.frames:
        try:
            au = frame.evaluate("""
                () => {
                    const el = document.querySelector(
                        '.profile-info .nick, .writer_info .nick, .author-name, '
                        + '.td_name, [class*="nick"]:first-of-type, .profile_nick'
                    );
                    return el ? el.textContent.trim() : '';
                }
            """)
            if au:
                actual_author = au
                break
        except Exception:
            pass

    content_html = ""
    for frame in page.frames:
        for sel in [
            ".article-body", "#article-body", ".se-main-container",
            ".faceTextViewer", "#contentDiv", ".article_viewer",
            ".postViewArea", "[class*='article-body']", "[class*='article-content']",
        ]:
            try:
                el = frame.query_selector(sel)
                if el:
                    h = el.inner_html()
                    if len(h) > 100:
                        content_html = h
                        break
            except Exception:
                pass
        if content_html:
            break

    if not content_html:
        try:
            content_html = page.content()
        except Exception:
            return []

    soup = BeautifulSoup(content_html, "html.parser")
    price_rows: list[dict] = []

    for table in soup.find_all("table"):
        try:
            dfs = pd.read_html(str(table), header=0)
            for df in dfs:
                df = df.dropna(how="all").reset_index(drop=True)
                if df.shape[1] < 2 or len(df) < 1:
                    continue
                if all(str(c).startswith("Unnamed") for c in df.columns):
                    df.columns = [str(v) for v in df.iloc[0].values]
                    df = df[1:].reset_index(drop=True)
                for _, r in df.iterrows():
                    vals = [str(v).strip() for v in r.values]
                    item = vals[0] if vals[0] not in ("nan", "-", "", "None") else ""
                    if not item or len(item) < 2:
                        continue
                    if any(h in item for h in ["품목", "제품", "모델", "규격", "항목"]):
                        continue
                    price = vals[1] if len(vals) > 1 and vals[1] != "nan" else ""
                    note  = " / ".join(v for v in vals[2:] if v not in ("nan", "", "None"))
                    price_rows.append({"품목": item, "수량": "", "단가": price, "비고": note})
        except Exception:
            pass

    if not price_rows:
        price_rows = extract_prices_from_text(soup.get_text(separator="\n"))

    if not price_rows:
        price_rows = [{"품목": "(본문 확인 필요)", "수량": "", "단가": "", "비고": ""}]

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    return [{
        "게시일":    post["date"],
        "업체":      post["vendor"],
        "거래유형":  trade_type,
        "게시판":    post.get("boardName", post.get("board", "")),
        "카테고리":  pr.get("카테고리", ""),
        "품목":      pr["품목"],
        "수량":      pr.get("수량", ""),
        "단가":      pr["단가"],
        "비고":      pr.get("비고", ""),
        "게시글제목": post["title"],
        "URL":       post["url"],
        "수집일시":  now_str,
    } for pr in price_rows]


# ── Excel 저장 ────────────────────────────────────────────────────────

def _style_ws(ws, hdr_hex: str, row_hex: str = "FFFFFF"):
    hfill = PatternFill("solid", fgColor=hdr_hex)
    rfill = PatternFill("solid", fgColor=row_hex)
    for cell in ws[1]:
        cell.fill = hfill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    if row_hex != "FFFFFF":
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.fill = rfill
    for col_cells in ws.columns:
        w = min(max(len(str(c.value or "")) for c in col_cells) + 2, 50)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = w


def save_excel(all_data: list[dict]) -> Path | None:
    if not all_data:
        print("[결과] 수집된 데이터가 없습니다.")
        return None
    df = pd.DataFrame(all_data)
    df["단가"] = pd.to_numeric(df["단가"], errors="coerce")
    ts   = datetime.now().strftime("%Y%m%d_%H%M")
    path = OUT_DIR / f"중고IT단가_{ts}.xlsx"
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="전체")
        _style_ws(writer.sheets["전체"], "1E3A5F")

        # 거래유형별 시트
        for trade_type, sheet_name, color in [("판매가", "판매목록", "1E5F3A"), ("구매가", "매입목록", "3A3A8F")]:
            tdf = df[df["거래유형"] == trade_type]
            if not tdf.empty:
                tdf.to_excel(writer, index=False, sheet_name=sheet_name)
                _style_ws(writer.sheets[sheet_name], color)

        for vendor in VENDORS:
            vdf = df[df["업체"] == vendor]
            if vdf.empty:
                continue
            vdf.to_excel(writer, index=False, sheet_name=vendor[:31])
            _style_ws(
                writer.sheets[vendor[:31]],
                hdr_hex="334155",
                row_hex=VENDOR_COLORS.get(vendor, "FFFFFF"),
            )
    print(f"\n[저장] {len(all_data)}건 → {path}")
    return path


# ── 메인 ─────────────────────────────────────────────────────────────

def main():
    all_data: list[dict] = []

    with sync_playwright() as pw:
        # 브라우저 프로파일 저장 → 로그인 세션 유지
        ctx = pw.chromium.launch_persistent_context(
            user_data_dir=str(PROFILE_DIR),
            headless=False,
            slow_mo=80,
            locale="ko-KR",
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            args=["--disable-blink-features=AutomationControlled"],
        )
        page = ctx.new_page()
        page.set_default_timeout(20000)

        # 1. 로그인
        login(page)

        # 2. grpid 추출
        grpid = "MMhX"
        for frame in page.frames:
            m = re.search(r"grpid=([A-Za-z0-9]+)", frame.url)
            if m:
                grpid = m.group(1)
                break
        print(f"[설정] grpid = {grpid}")

        # 3. 업체별 검색
        all_posts: list[dict] = []
        seen_urls: set[str] = set()

        for vendor in VENDORS:
            posts = search_vendor(page, vendor, grpid)
            for p in posts:
                if p["url"] not in seen_urls:
                    seen_urls.add(p["url"])
                    all_posts.append(p)

        print(f"\n[수집] 총 {len(all_posts)}개 게시글 발견")

        # 4. 게시글 본문 파싱
        print("\n[파싱] 게시글 본문 처리 중...")
        for i, post in enumerate(all_posts, 1):
            print(
                f"  [{i}/{len(all_posts)}] [{post['vendor']}] "
                f"{post['title'][:35]}  ({post['date']})"
            )
            records = parse_post(page, post)
            all_data.extend(records)
            print(f"    → {len(records)}건 추출")

        print(f"\n[완료] 총 {len(all_data)}건 수집")
        input("\n브라우저를 닫으려면 Enter: ")
        ctx.close()

    save_excel(all_data)


if __name__ == "__main__":
    main()
