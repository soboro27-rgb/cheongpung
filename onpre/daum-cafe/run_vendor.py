"""
특정 업체 최근 N일 게시글 수집 → 기준표 단가 자동 기재

실행: py run_vendor.py
"""
import re
import sys
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from playwright.sync_api import sync_playwright

sys.path.insert(0, str(Path(__file__).parent))
import scraper as _s

# ── 설정 ─────────────────────────────────────────────────────────────
VENDOR    = "에이스피씨"
MAX_POSTS = 3             # 최신 게시글 몇 개까지 조회할지
REF_PATH  = Path(__file__).parent / "vendors" / "에이스피씨_기준표.xlsx"


# ── 매칭 유틸 ─────────────────────────────────────────────────────────

def normalize(text: str) -> str:
    t = str(text).lower().replace("\xa0", " ")
    t = re.sub(r"이상$", "", t.strip())          # "이상" 제거
    t = re.sub(r"[^a-z0-9가-힣]", "", t)        # 특수문자 제거
    return t.strip()


def tokenize(text: str) -> list[str]:
    n = normalize(text)
    return re.findall(r"[a-z]+\d*|\d+[a-z]*|[가-힣]+", n)


def match_score(ref_item: str, ref_brand: str, scr_item: str) -> float:
    ref_tokens = tokenize(ref_item)
    if not ref_tokens:
        return 0.0
    scr_n = normalize(scr_item)

    # 토큰 매칭 점수
    token_score = sum(1 for t in ref_tokens if t in scr_n) / len(ref_tokens)

    # 브랜드 보정
    brand_bonus = 0.0
    brand_n = normalize(ref_brand)
    if brand_n in ("외산", "기타"):
        # 외산 = 삼성·하이닉스 이외
        if "삼성" not in scr_n:
            brand_bonus = 0.1
    elif brand_n and brand_n in scr_n:
        brand_bonus = 0.2

    return token_score + brand_bonus


def find_best_match(ref_item: str, ref_brand: str,
                    scraped_items: list[dict]) -> dict | None:
    ref_tokens = tokenize(ref_item)
    min_score  = max(0.5, (len(ref_tokens) - 1) / len(ref_tokens)) if ref_tokens else 0.5

    best_score, best = 0.0, None
    for item in scraped_items:
        sc = match_score(ref_item, ref_brand, item.get("품목", ""))
        if sc >= min_score and sc > best_score:
            best_score, best = sc, item
    return best


# ── 기준표 기재 ───────────────────────────────────────────────────────

def fill_reference(scraped_data: list[dict]):
    if not REF_PATH.exists():
        print(f"[오류] 기준표 없음: {REF_PATH}")
        return

    wb = openpyxl.load_workbook(REF_PATH)
    ws = wb.active

    # 추가 헤더 컬럼 (없으면 생성)
    extra_cols = {5: "수집단가", 6: "게시일", 7: "거래유형"}
    hdr_fill = PatternFill("solid", fgColor="1E3A5F")
    hdr_font = Font(color="FFFFFF", bold=True)
    for col, name in extra_cols.items():
        if ws.cell(row=1, column=col).value != name:
            c = ws.cell(row=1, column=col, value=name)
            c.fill = hdr_fill
            c.font = hdr_font
            c.alignment = Alignment(horizontal="center")

    # 수집 품목 → 최신 기준으로 dict화 {norm_key: row}
    latest: dict[str, dict] = {}
    for row in scraped_data:
        item  = row.get("품목", "")
        price = row.get("단가")
        date  = row.get("게시일", "")
        if not item or not price:
            continue
        key = normalize(item)
        if key not in latest or date > latest[key].get("게시일", ""):
            latest[key] = row

    scr_list = list(latest.values())
    print(f"[매칭] 수집 품목 {len(scr_list)}종")

    filled = 0
    for r in range(2, ws.max_row + 1):
        ref_item  = ws.cell(row=r, column=3).value
        ref_brand = ws.cell(row=r, column=2).value or ""
        if not ref_item or str(ref_item).startswith("▶"):
            continue

        matched = find_best_match(str(ref_item), str(ref_brand), scr_list)
        if matched:
            ws.cell(row=r, column=5).value = matched.get("단가")
            ws.cell(row=r, column=6).value = matched.get("게시일", "")
            ws.cell(row=r, column=7).value = matched.get("거래유형", "")
            # 수집단가 컬럼 숫자 표시 서식
            ws.cell(row=r, column=5).number_format = "#,##0"
            filled += 1

    # 컬럼 너비 자동 조정
    for col, width in zip([5, 6, 7], [12, 10, 8]):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    wb.save(REF_PATH)
    print(f"[완료] {filled}개 항목 기재 → {REF_PATH.name}")


# ── 메인 ─────────────────────────────────────────────────────────────

def main():
    # 날짜 필터 비활성화 — 최신 게시글 MAX_POSTS개만 가져옴
    _s.CUTOFF = datetime.now() - timedelta(days=3650)

    all_data: list[dict] = []

    with sync_playwright() as pw:
        ctx = pw.chromium.launch_persistent_context(
            user_data_dir=str(_s.PROFILE_DIR),
            headless=False,
            slow_mo=80,
            locale="ko-KR",
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            args=["--disable-blink-features=AutomationControlled"],
        )
        page = ctx.new_page()
        page.set_default_timeout(20000)

        _s.login(page)

        # grpid 추출
        grpid = "MMhX"
        for frame in page.frames:
            m = re.search(r"grpid=([A-Za-z0-9]+)", frame.url)
            if m:
                grpid = m.group(1)
                break

        # 검색 후 최신 MAX_POSTS개만 사용
        all_posts = _s.search_vendor(page, VENDOR, grpid)
        posts = all_posts[:MAX_POSTS]
        print(f"\n[수집] {VENDOR} 최신 {len(posts)}개 게시글 처리 (전체 {len(all_posts)}개 중)")

        for i, post in enumerate(posts, 1):
            print(f"  [{i}/{len(posts)}] {post['title'][:40]}  ({post['date']})")
            records = _s.parse_post(page, post)
            all_data.extend(records)
            print(f"    → {len(records)}건 추출")

        print(f"\n[합계] {len(all_data)}건 추출")
        input("\n브라우저 닫으려면 Enter: ")
        ctx.close()

    if all_data:
        fill_reference(all_data)
    else:
        print("[결과] 수집된 데이터 없음")


if __name__ == "__main__":
    main()
