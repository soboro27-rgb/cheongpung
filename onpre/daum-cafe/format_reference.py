"""
기준표 포맷 변환: 좌우 병렬 → 카테고리/브랜드/품목/매입단가 세로 목록
입력: vendors/기준표작성중_매입완료.xlsx
출력: vendors/에이스피씨_기준표.xlsx
"""
import re
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path

SRC = Path(__file__).parent / "vendors" / "기준표작성중_매입완료.xlsx"
DST = Path(__file__).parent / "vendors" / "에이스피씨_기준표.xlsx"

CAT_COLORS = {
    "CPU":       ("BDD7EE", "DEEAF6"),
    "메모리":    ("C6EFCE", "E2F0D9"),
    "노트북메모리": ("FFEB9C", "FFF2CC"),
    "그래픽카드": ("FCE4D6", "FCF4F0"),
    "SSD/HDD":  ("DDD9F3", "EDEBF7"),
    "메인보드":  ("D9E1F2", "EDF1F9"),
}
DEFAULT_COLORS = ("E2E2E2", "F5F5F5")


def clean(v):
    if v is None:
        return ""
    return str(v).replace("\xa0", " ").strip()


def parse_source():
    wb = openpyxl.load_workbook(SRC)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    items = []
    current_cat = ""
    brand_map = {}  # {col_index(0-based): brand_name}

    for row in rows:
        a = clean(row[0])

        if a:
            # 카테고리 헤더 행: A열에 카테고리명, B/D/F/H열에 브랜드명
            current_cat = a
            brand_map = {}
            for i in range(1, min(len(row), 9), 2):
                v = clean(row[i])
                if v and v != "단가":
                    brand_map[i] = v
        else:
            # 데이터 행
            for col_i, brand in brand_map.items():
                v = clean(row[col_i]) if col_i < len(row) else ""
                if not v:
                    continue
                # 품목명에 내포된 가격 정보 제거 (예: "라이젠 1300   2개 10,000원")
                item = re.sub(r"\s{2,}\d+개\s+[\d,.]+원?.*$", "", v).strip()
                item = re.sub(r"\s{2,}", " ", item).strip()
                if not item:
                    continue
                items.append({
                    "카테고리": current_cat,
                    "브랜드":   brand,
                    "품목":     item,
                    "매입단가": "",
                })

    return items


def make_excel(items):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "기준표"

    # ── 헤더 행 ──
    headers = ["카테고리", "브랜드", "품목", "매입단가"]
    hdr_fill = PatternFill("solid", fgColor="1E3A5F")
    hdr_font = Font(color="FFFFFF", bold=True, size=11)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill  = hdr_fill
        cell.font  = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    # ── 데이터 행 ──
    prev_cat = ""
    r = 2

    for item in items:
        cat   = item["카테고리"]
        hcol, dcol = CAT_COLORS.get(cat, DEFAULT_COLORS)

        # 카테고리가 바뀔 때 구분 헤더 삽입
        if cat != prev_cat:
            sect_fill = PatternFill("solid", fgColor=hcol)
            sect_font = Font(bold=True, size=11)
            for c in range(1, 5):
                cell = ws.cell(row=r, column=c)
                cell.fill = sect_fill
                cell.font = sect_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=r, column=1).value = f"▶ {cat}"
            ws.merge_cells(f"A{r}:D{r}")
            ws.row_dimensions[r].height = 18
            r += 1
            prev_cat = cat

        # 데이터 셀
        alt_fill = PatternFill("solid", fgColor=dcol if r % 2 == 0 else "FFFFFF")
        for c, key in enumerate(["카테고리", "브랜드", "품목", "매입단가"], 1):
            cell = ws.cell(row=r, column=c, value=item[key])
            cell.fill = alt_fill
            cell.alignment = Alignment(vertical="center",
                                       horizontal="right" if key == "매입단가" else "left")
        r += 1

    # ── 컬럼 너비 ──
    for col, width in zip([1, 2, 3, 4], [14, 16, 38, 12]):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.auto_filter.ref = f"A1:D{r - 1}"
    ws.freeze_panes = "A2"

    wb.save(DST)
    print(f"[완료] {len(items)}개 품목 → {DST}")


if __name__ == "__main__":
    items = parse_source()
    print(f"[파싱] {len(items)}건")
    make_excel(items)
