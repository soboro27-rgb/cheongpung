from fastapi import APIRouter, Request, Depends
from fastapi.responses import RedirectResponse, HTMLResponse, JSONResponse, StreamingResponse
from sqlalchemy.orm import Session
from database import get_db
import models
from models import UserRole, AppStatus
from auth import require_wm, require_super_admin
from config import templates
import bcrypt
from datetime import datetime
import openpyxl
from io import BytesIO

router = APIRouter()


def _hash(password: str) -> str:
    return bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()


def _check(request: Request):
    u = require_wm(request)
    if not u:
        return None, RedirectResponse("/login", status_code=302)
    return u, None


def _check_super(request: Request):
    u = require_super_admin(request)
    if not u:
        return None, RedirectResponse("/login", status_code=302)
    return u, None


# ─── 대시보드 ──────────────────────────────────────────

@router.get("/dashboard", response_class=HTMLResponse)
def dashboard(request: Request, db: Session = Depends(get_db)):
    u, redir = _check(request)
    if redir: return redir

    total_apps      = db.query(models.Application).count()
    pending_apps    = db.query(models.Application).filter(models.Application.status == AppStatus.REQUESTED).count()
    in_progress     = db.query(models.Application).filter(
        models.Application.status.in_([AppStatus.RECEIVED, AppStatus.SCHEDULED, AppStatus.COLLECTED, AppStatus.QUOTED])
    ).count()
    completed_apps  = db.query(models.Application).filter(
        models.Application.status.in_([AppStatus.APPROVED, AppStatus.WIPED, AppStatus.SETTLED, AppStatus.CLOSED])
    ).count()
    total_dealers   = db.query(models.Dealer).filter(models.Dealer.is_active == True).count()
    total_customers = db.query(models.Customer).filter(models.Customer.is_active == True).count()

    recent = (
        db.query(models.Application)
        .order_by(models.Application.updated_at.desc())
        .limit(10).all()
    )

    return templates.TemplateResponse(request, "wm/dashboard.html", {"session": request.session,
        "total_apps": total_apps, "pending_apps": pending_apps,
        "in_progress": in_progress, "completed_apps": completed_apps,
        "total_dealers": total_dealers, "total_customers": total_customers,
        "recent": recent})


# ─── IDC 센터 관리 ─────────────────────────────────────

@router.get("/idc-centers", response_class=HTMLResponse)
def idc_list(request: Request, db: Session = Depends(get_db)):
    u, redir = _check(request)
    if redir: return redir
    centers = db.query(models.IdcCenter).order_by(models.IdcCenter.name).all()
    return templates.TemplateResponse(request, "wm/idc_centers.html", {"session": request.session, "centers": centers})


@router.post("/idc-centers")
async def idc_create(request: Request, db: Session = Depends(get_db)):
    u, redir = _check_super(request)
    if redir: return redir
    form = await request.form()
    db.add(models.IdcCenter(
        name=form.get("name", "").strip(),
        address=form.get("address", "").strip(),
        security_notes=form.get("security_notes", "").strip(),
        work_hours=form.get("work_hours", "").strip(),
        contact_name=form.get("contact_name", "").strip(),
        contact_phone=form.get("contact_phone", "").strip(),
    ))
    db.commit()
    return RedirectResponse("/wm/idc-centers", status_code=302)


@router.post("/idc-centers/{center_id}/delete")
def idc_delete(request: Request, center_id: int, db: Session = Depends(get_db)):
    u, redir = _check_super(request)
    if redir: return redir
    c = db.query(models.IdcCenter).filter(models.IdcCenter.id == center_id).first()
    if c:
        c.is_active = False
        db.commit()
    return RedirectResponse("/wm/idc-centers", status_code=302)


# ─── 운영사 관리 ────────────────────────────────────────

@router.get("/operators", response_class=HTMLResponse)
def operator_list(request: Request, db: Session = Depends(get_db)):
    u, redir = _check(request)
    if redir: return redir
    operators = db.query(models.Operator).order_by(models.Operator.name).all()
    return templates.TemplateResponse(request, "wm/operators.html", {"session": request.session,
        "operators": operators})


@router.post("/operators")
async def operator_create(request: Request, db: Session = Depends(get_db)):
    u, redir = _check_super(request)
    if redir: return redir
    form = await request.form()
    try:
        fee_value = float(form.get("fee_value", 0) or 0)
    except ValueError:
        fee_value = 0.0

    db.add(models.Operator(
        name=form.get("name", "").strip(),
        business_no=form.get("business_no", "").strip(),
        manager_name=form.get("manager_name", "").strip(),
        manager_phone=form.get("manager_phone", "").strip(),
        manager_email=form.get("manager_email", "").strip(),
        operator_code=form.get("operator_code", "").strip().upper(),
        fee_type=form.get("fee_type", "percent"),
        fee_value=fee_value,
    ))
    db.commit()
    return RedirectResponse("/wm/operators", status_code=302)


@router.get("/operators/{operator_id}", response_class=HTMLResponse)
def operator_detail(request: Request, operator_id: int, db: Session = Depends(get_db)):
    u, redir = _check(request)
    if redir: return redir
    operator = db.query(models.Operator).filter(models.Operator.id == operator_id).first()
    if not operator:
        return RedirectResponse("/wm/operators", status_code=302)
    dealers = db.query(models.Dealer).filter(models.Dealer.operator_id == operator_id).all()
    users   = db.query(models.User).filter(models.User.operator_id == operator_id).all()
    return templates.TemplateResponse(request, "wm/operator_detail.html", {"session": request.session,
        "operator": operator, "dealers": dealers, "users": users})


# ─── 딜러 관리 ────────────────────────────────────────

@router.get("/dealers", response_class=HTMLResponse)
def dealer_list(request: Request, db: Session = Depends(get_db)):
    u, redir = _check(request)
    if redir: return redir
    dealers = db.query(models.Dealer).order_by(models.Dealer.name).all()
    centers = db.query(models.IdcCenter).filter(models.IdcCenter.is_active == True).all()
    operators = db.query(models.Operator).filter(models.Operator.is_active == True).order_by(models.Operator.name).all()
    return templates.TemplateResponse(request, "wm/dealers.html", {"session": request.session,
        "dealers": dealers, "centers": centers, "operators": operators})


@router.post("/dealers")
async def dealer_create(request: Request, db: Session = Depends(get_db)):
    u, redir = _check_super(request)
    if redir: return redir
    form = await request.form()
    try:
        fee_value = float(form.get("fee_value", 0) or 0)
    except ValueError:
        fee_value = 0.0
    try:
        operator_id = int(form.get("operator_id") or 0) or None
    except (ValueError, TypeError):
        operator_id = None

    dealer = models.Dealer(
        operator_id=operator_id,
        name=form.get("name", "").strip(),
        business_no=form.get("business_no", "").strip(),
        manager_name=form.get("manager_name", "").strip(),
        manager_phone=form.get("manager_phone", "").strip(),
        manager_email=form.get("manager_email", "").strip(),
        dealer_code=form.get("dealer_code", "").strip().upper(),
        settlement_type=form.get("settlement_type", "direct"),
        fee_type=form.get("fee_type", "percent"),
        fee_value=fee_value,
    )
    db.add(dealer)
    db.flush()

    # 담당 IDC 연결
    for cid in form.getlist("idc_center_ids[]"):
        try:
            db.add(models.DealerIdcCenter(dealer_id=dealer.id, idc_center_id=int(cid)))
        except (ValueError, TypeError):
            pass

    db.commit()
    return RedirectResponse("/wm/dealers", status_code=302)


@router.get("/dealers/{dealer_id}", response_class=HTMLResponse)
def dealer_detail(request: Request, dealer_id: int, db: Session = Depends(get_db)):
    u, redir = _check(request)
    if redir: return redir
    dealer = db.query(models.Dealer).filter(models.Dealer.id == dealer_id).first()
    if not dealer:
        return RedirectResponse("/wm/dealers", status_code=302)
    centers   = db.query(models.IdcCenter).filter(models.IdcCenter.is_active == True).all()
    customers = db.query(models.Customer).filter(models.Customer.dealer_id == dealer_id).all()
    users     = db.query(models.User).filter(models.User.dealer_id == dealer_id).all()
    operators = db.query(models.Operator).filter(models.Operator.is_active == True).order_by(models.Operator.name).all()
    return templates.TemplateResponse(request, "wm/dealer_detail.html", {"session": request.session,
        "dealer": dealer, "centers": centers,
        "customers": customers, "users": users, "operators": operators})


@router.post("/dealers/{dealer_id}/operator")
async def dealer_set_operator(request: Request, dealer_id: int, db: Session = Depends(get_db)):
    """딜러의 소속 운영사 재지정 (없음 선택 시 3단 구조로 되돌림)"""
    u, redir = _check_super(request)
    if redir: return redir
    dealer = db.query(models.Dealer).filter(models.Dealer.id == dealer_id).first()
    if not dealer:
        return RedirectResponse("/wm/dealers", status_code=302)
    form = await request.form()
    try:
        operator_id = int(form.get("operator_id") or 0) or None
    except (ValueError, TypeError):
        operator_id = None
    dealer.operator_id = operator_id
    db.commit()
    return RedirectResponse(f"/wm/dealers/{dealer_id}?success=operator_updated", status_code=302)


# ─── 고객사 관리 (WM 전체 조회) ───────────────────────

@router.get("/customers", response_class=HTMLResponse)
def customer_list(request: Request, dealer_id: int = 0, db: Session = Depends(get_db)):
    u, redir = _check(request)
    if redir: return redir
    q = db.query(models.Customer)
    if dealer_id:
        q = q.filter(models.Customer.dealer_id == dealer_id)
    customers = q.order_by(models.Customer.name).all()
    dealers   = db.query(models.Dealer).filter(models.Dealer.is_active == True).all()
    centers   = db.query(models.IdcCenter).filter(models.IdcCenter.is_active == True).all()
    return templates.TemplateResponse(request, "wm/customers.html", {"session": request.session,
        "customers": customers, "dealers": dealers, "centers": centers,
        "filter_dealer_id": dealer_id})


@router.post("/customers")
async def customer_create(request: Request, db: Session = Depends(get_db)):
    u, redir = _check_super(request)
    if redir: return redir
    form = await request.form()
    try:
        idc_id = int(form.get("default_idc_center_id") or 0) or None
    except (ValueError, TypeError):
        idc_id = None

    db.add(models.Customer(
        dealer_id=int(form.get("dealer_id")),
        name=form.get("name", "").strip(),
        customer_code=form.get("customer_code", "").strip().upper(),
        business_no=form.get("business_no", "").strip(),
        manager_name=form.get("manager_name", "").strip(),
        manager_phone=form.get("manager_phone", "").strip(),
        manager_email=form.get("manager_email", "").strip(),
        default_idc_center_id=idc_id,
    ))
    db.commit()
    return RedirectResponse("/wm/customers", status_code=302)


# ─── 사용자(계정) 관리 ────────────────────────────────

@router.get("/users", response_class=HTMLResponse)
def user_list(request: Request, db: Session = Depends(get_db)):
    u, redir = _check_super(request)
    if redir: return redir
    users     = db.query(models.User).order_by(models.User.created_at.desc()).all()
    operators = db.query(models.Operator).filter(models.Operator.is_active == True).all()
    dealers   = db.query(models.Dealer).filter(models.Dealer.is_active == True).all()
    customers = db.query(models.Customer).filter(models.Customer.is_active == True).all()
    return templates.TemplateResponse(request, "wm/users.html", {"session": request.session,
        "users": users, "operators": operators, "dealers": dealers, "customers": customers})


@router.post("/users")
async def user_create(request: Request, db: Session = Depends(get_db)):
    u, redir = _check_super(request)
    if redir: return redir
    form = await request.form()
    role_str = form.get("role", "")

    try:
        role = UserRole(role_str)
    except ValueError:
        return RedirectResponse("/wm/users?error=invalid_role", status_code=302)

    operator_id = None
    dealer_id = None
    customer_id = None
    if role in (UserRole.OPERATOR_ADMIN, UserRole.OPERATOR_STAFF):
        try: operator_id = int(form.get("operator_id") or 0) or None
        except: pass
    elif role in (UserRole.DEALER_ADMIN, UserRole.DEALER_STAFF):
        try: dealer_id = int(form.get("dealer_id") or 0) or None
        except: pass
    elif role in (UserRole.CUSTOMER_ADMIN, UserRole.CUSTOMER_STAFF):
        try: customer_id = int(form.get("customer_id") or 0) or None
        except: pass
        # 고객사 소속 → 딜러도 연결
        if customer_id:
            cust = db.query(models.Customer).filter(models.Customer.id == customer_id).first()
            if cust:
                dealer_id = cust.dealer_id

    password = form.get("password", "")
    if not password:
        return RedirectResponse("/wm/users?error=no_password", status_code=302)

    db.add(models.User(
        login_id=form.get("login_id", "").strip(),
        password_hash=_hash(password),
        name=form.get("name", "").strip(),
        email=form.get("email", "").strip(),
        phone=form.get("phone", "").strip(),
        role=role,
        operator_id=operator_id,
        dealer_id=dealer_id,
        customer_id=customer_id,
    ))
    db.commit()
    return RedirectResponse("/wm/users", status_code=302)


@router.post("/users/{user_id}/toggle")
def user_toggle(request: Request, user_id: int, db: Session = Depends(get_db)):
    u, redir = _check_super(request)
    if redir: return redir
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if user and user.id != u["user_id"]:
        user.is_active = not user.is_active
        db.commit()
    return RedirectResponse("/wm/users", status_code=302)


@router.post("/users/{user_id}/reset-password")
async def user_reset_password(request: Request, user_id: int, db: Session = Depends(get_db)):
    u, redir = _check_super(request)
    if redir: return redir
    form = await request.form()
    new_pw = str(form.get("new_password", "")).strip()
    if len(new_pw) < 6:
        return RedirectResponse("/wm/users?err=short_pw", status_code=302)
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if user:
        user.password_hash = _hash(new_pw)
        db.commit()
    return RedirectResponse("/wm/users?reset=ok", status_code=302)


@router.post("/users/{user_id}/edit")
async def user_edit(request: Request, user_id: int, db: Session = Depends(get_db)):
    u, redir = _check_super(request)
    if redir: return redir
    form = await request.form()
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        return RedirectResponse("/wm/users", status_code=302)

    role_str = form.get("role", "")
    try:
        role = UserRole(role_str)
    except ValueError:
        return RedirectResponse("/wm/users?error=invalid_role", status_code=302)

    operator_id = None
    dealer_id = None
    customer_id = None
    if role in (UserRole.OPERATOR_ADMIN, UserRole.OPERATOR_STAFF):
        try: operator_id = int(form.get("operator_id") or 0) or None
        except: pass
    elif role in (UserRole.DEALER_ADMIN, UserRole.DEALER_STAFF):
        try: dealer_id = int(form.get("dealer_id") or 0) or None
        except: pass
    elif role in (UserRole.CUSTOMER_ADMIN, UserRole.CUSTOMER_STAFF):
        try: customer_id = int(form.get("customer_id") or 0) or None
        except: pass
        if customer_id:
            cust = db.query(models.Customer).filter(models.Customer.id == customer_id).first()
            if cust:
                dealer_id = cust.dealer_id

    name = form.get("name", "").strip()
    if name:
        user.name = name
    user.email = form.get("email", "").strip()
    user.phone = form.get("phone", "").strip()
    user.role = role
    user.operator_id = operator_id
    user.dealer_id = dealer_id
    user.customer_id = customer_id
    db.commit()
    return RedirectResponse("/wm/users?edited=ok", status_code=302)


@router.post("/users/{user_id}/delete")
def user_delete(request: Request, user_id: int, db: Session = Depends(get_db)):
    u, redir = _check_super(request)
    if redir: return redir
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if user and user.id != u["user_id"]:
        # 이 계정을 참조하는 이력 레코드는 삭제하지 않고 참조만 해제
        # (로그인 이력·수정 이력 자체는 감사 목적으로 보존)
        db.query(models.LoginLog).filter(models.LoginLog.user_id == user_id).update({"user_id": None})
        db.query(models.AssetEditLog).filter(models.AssetEditLog.user_id == user_id).update({"user_id": None})
        db.query(models.ServerAsset).filter(models.ServerAsset.locked_by == user_id).update({"locked_by": None})
        db.delete(user)
        db.commit()
    return RedirectResponse("/wm/users?deleted=ok", status_code=302)


# ─── 전체 신청 목록 ───────────────────────────────────

@router.get("/applications", response_class=HTMLResponse)
def app_list(request: Request, status: str = "", dealer_id: int = 0, db: Session = Depends(get_db)):
    u, redir = _check(request)
    if redir: return redir
    q = db.query(models.Application)
    if status:
        try: q = q.filter(models.Application.status == AppStatus(status))
        except ValueError: pass
    if dealer_id:
        q = q.filter(models.Application.dealer_id == dealer_id)
    apps    = q.order_by(models.Application.updated_at.desc()).all()
    dealers = db.query(models.Dealer).filter(models.Dealer.is_active == True).all()
    return templates.TemplateResponse(request, "wm/applications.html", {"session": request.session,
        "applications": apps, "dealers": dealers,
        "current_status": status, "filter_dealer_id": dealer_id})


@router.get("/settlements", response_class=HTMLResponse)
def settlement_list(request: Request, dealer_id: int = 0, db: Session = Depends(get_db)):
    u, redir = _check(request)
    if redir: return redir
    q = (db.query(models.Settlement)
         .join(models.Application, models.Settlement.application_id == models.Application.id))
    if dealer_id:
        q = q.filter(models.Application.dealer_id == dealer_id)
    settlements = q.order_by(models.Settlement.created_at.desc()).all()
    dealers = db.query(models.Dealer).filter(models.Dealer.is_active == True).all()
    total_amount = sum(s.total_amount for s in settlements)
    total_customer = sum(s.customer_amount for s in settlements)
    return templates.TemplateResponse(request, "wm/settlements.html", {
        "session": request.session,
        "settlements": settlements, "dealers": dealers,
        "filter_dealer_id": dealer_id,
        "total_amount": total_amount, "total_customer": total_customer,
    })


@router.post("/applications/{app_id}/receive")
def receive_app(request: Request, app_id: int, db: Session = Depends(get_db)):
    u, redir = _check(request)
    if redir: return redir
    app = db.query(models.Application).filter(
        models.Application.id == app_id,
        models.Application.status == AppStatus.REQUESTED,
    ).first()
    if app:
        app.status = AppStatus.RECEIVED
        app.received_at = datetime.now()
        app.updated_at = datetime.now()
        db.commit()
    return RedirectResponse(f"/wm/applications/{app_id}", status_code=302)


@router.get("/applications/{app_id}", response_class=HTMLResponse)
def app_detail(request: Request, app_id: int, db: Session = Depends(get_db)):
    u, redir = _check(request)
    if redir: return redir
    app = db.query(models.Application).filter(models.Application.id == app_id).first()
    if not app:
        return RedirectResponse("/wm/applications", status_code=302)
    centers = db.query(models.IdcCenter).filter(models.IdcCenter.is_active == True).all()
    return templates.TemplateResponse(request, "wm/application_detail.html", {"session": request.session,
        "app": app, "centers": centers})


@router.post("/applications/{app_id}/schedule")
async def set_schedule(request: Request, app_id: int, db: Session = Depends(get_db)):
    u, redir = _check(request)
    if redir: return redir
    form = await request.form()
    app = db.query(models.Application).filter(
        models.Application.id == app_id,
        models.Application.status == AppStatus.RECEIVED,
    ).first()
    if app:
        if not app.schedule:
            db.add(models.VisitSchedule(application_id=app_id))
            db.flush(); db.refresh(app)
        app.schedule.visit_date     = form.get("visit_date", "")
        app.schedule.visit_time     = form.get("visit_time", "")
        app.schedule.collector_name = form.get("collector_name", "")
        app.schedule.collector_phone = form.get("collector_phone", "")
        app.schedule.notes          = form.get("notes", "")
        app.status = AppStatus.SCHEDULED
        app.updated_at = datetime.now()
        try:
            idc_id = int(form.get("idc_center_id") or 0) or None
            app.idc_center_id = idc_id
        except (ValueError, TypeError):
            pass
        db.commit()
    return RedirectResponse(f"/wm/applications/{app_id}", status_code=302)


@router.post("/applications/{app_id}/schedule-edit")
async def edit_schedule(request: Request, app_id: int, db: Session = Depends(get_db)):
    u, redir = _check(request)
    if redir: return redir
    form = await request.form()
    app = db.query(models.Application).filter(models.Application.id == app_id).first()
    if not app or not app.schedule:
        return RedirectResponse(f"/wm/applications/{app_id}", status_code=302)

    app.schedule.visit_date      = form.get("visit_date", "")
    app.schedule.visit_time      = form.get("visit_time", "")
    app.schedule.collector_name  = form.get("collector_name", "")
    app.schedule.collector_phone = form.get("collector_phone", "")
    app.schedule.notes           = form.get("notes", "")
    try:
        idc_id = int(form.get("idc_center_id") or 0) or None
        app.idc_center_id = idc_id
    except (ValueError, TypeError):
        pass
    app.updated_at = datetime.now()
    db.commit()
    return RedirectResponse(f"/wm/applications/{app_id}?schedule_ok=1", status_code=302)


@router.post("/applications/{app_id}/estimated-price")
async def save_estimated_price(request: Request, app_id: int, db: Session = Depends(get_db)):
    """수거 전 예정단가 저장 (received/scheduled 상태)"""
    u, redir = _check(request)
    if redir: return redir
    form = await request.form()
    app = db.query(models.Application).filter(
        models.Application.id == app_id,
        models.Application.status.in_([AppStatus.RECEIVED, AppStatus.SCHEDULED]),
    ).first()
    if app:
        for asset in app.assets:
            raw = form.get(f"eprice_{asset.id}", "")
            try:
                asset.unit_price_estimated = float(str(raw).replace(",", "") or 0)
            except ValueError:
                pass
        app.updated_at = datetime.now()
        db.commit()
    return RedirectResponse(f"/wm/applications/{app_id}", status_code=302)


@router.post("/applications/{app_id}/collect")
def mark_collected(request: Request, app_id: int, db: Session = Depends(get_db)):
    u, redir = _check(request)
    if redir: return redir
    app = db.query(models.Application).filter(
        models.Application.id == app_id,
        models.Application.status == AppStatus.SCHEDULED,
    ).first()
    if app:
        app.status = AppStatus.COLLECTED
        app.collected_at = datetime.now()
        app.updated_at = datetime.now()
        db.commit()
    return RedirectResponse(f"/wm/applications/{app_id}", status_code=302)


# ─── Phase 2: 검수 / 단가확정 / 정산 ───────────────────

def _gen_qr(dealer_code: str, customer_code: str, seq: int) -> str:
    from datetime import date
    d = date.today().strftime("%Y_%m%d")
    return f"{d}_{dealer_code}_{customer_code}_{seq:04d}"


@router.post("/applications/{app_id}/bulk-price")
async def bulk_price(request: Request, app_id: int, db: Session = Depends(get_db)):
    """검수자: 자산별 확정 단가 일괄 입력"""
    u = require_wm(request)
    if not u:
        return RedirectResponse("/login", status_code=302)
    if u["role"] not in ("super_admin", "wm_inspector"):
        return RedirectResponse(f"/wm/applications/{app_id}?err=no_perm", status_code=302)

    form = await request.form()
    app = db.query(models.Application).filter(models.Application.id == app_id).first()
    if not app:
        return RedirectResponse("/wm/applications", status_code=302)

    for asset in app.assets:
        key = f"price_{asset.id}"
        raw = str(form.get(key, "")).replace(",", "").strip()
        if not raw:
            continue
        try:
            price = float(raw)
        except ValueError:
            continue

        # QUOTED 이후 잠긴 자산 → super_admin만 수정 가능 + 감사 로그
        if asset.is_locked:
            if u["role"] != "super_admin":
                continue
            if abs(price - asset.unit_price_confirmed) > 0.01:
                db.add(models.AssetEditLog(
                    asset_id=asset.id,
                    user_id=u["user_id"],
                    field_name="unit_price_confirmed",
                    old_value=str(asset.unit_price_confirmed),
                    new_value=str(price),
                ))
        asset.unit_price_confirmed = price
        asset.updated_at = datetime.now()

        # QR 코드가 아직 없으면 생성
        if not asset.qr_code and app.dealer and app.customer:
            asset.qr_code = _gen_qr(
                app.dealer.dealer_code,
                app.customer.customer_code,
                asset.id,
            )

    db.commit()
    return RedirectResponse(f"/wm/applications/{app_id}", status_code=302)


@router.post("/applications/{app_id}/quote")
def quote_app(request: Request, app_id: int, db: Session = Depends(get_db)):
    """단가 확정 → QUOTED: 모든 자산 잠금 + 상태 전이"""
    u = require_wm(request)
    if not u:
        return RedirectResponse("/login", status_code=302)
    if u["role"] not in ("super_admin", "wm_inspector"):
        return RedirectResponse(f"/wm/applications/{app_id}?err=no_perm", status_code=302)

    app = db.query(models.Application).filter(
        models.Application.id == app_id,
        models.Application.status == AppStatus.COLLECTED,
    ).first()
    if not app:
        return RedirectResponse(f"/wm/applications/{app_id}", status_code=302)

    # 미입력 자산 체크
    unprice = [a for a in app.assets if a.unit_price_confirmed <= 0]
    if unprice:
        return RedirectResponse(f"/wm/applications/{app_id}?err=unprice", status_code=302)

    now = datetime.now()
    for asset in app.assets:
        asset.is_locked = True
        asset.locked_at = now
        asset.locked_by = u["user_id"]
        asset.updated_at = now

    app.status = AppStatus.QUOTED
    app.quoted_at = now
    app.updated_at = now
    db.commit()
    return RedirectResponse(f"/wm/applications/{app_id}", status_code=302)


@router.post("/applications/{app_id}/wipe")
def mark_wiped(request: Request, app_id: int, db: Session = Depends(get_db)):
    """데이터 파기 완료"""
    u, redir = _check(request)
    if redir: return redir
    app = db.query(models.Application).filter(
        models.Application.id == app_id,
        models.Application.status == AppStatus.APPROVED,
    ).first()
    if app:
        app.status = AppStatus.WIPED
        app.updated_at = datetime.now()
        db.commit()
    return RedirectResponse(f"/wm/applications/{app_id}", status_code=302)


@router.post("/applications/{app_id}/settle")
async def settle_app(request: Request, app_id: int, db: Session = Depends(get_db)):
    """정산 완료 처리 + Settlement 생성/갱신"""
    u, redir = _check_super(request)
    if redir: return redir

    form = await request.form()
    app = db.query(models.Application).filter(
        models.Application.id == app_id,
        models.Application.status.in_([AppStatus.WIPED, AppStatus.APPROVED]),
    ).first()
    if not app:
        return RedirectResponse(f"/wm/applications/{app_id}", status_code=302)

    # 금액 계산
    total = sum(a.unit_price_confirmed * a.quantity for a in app.assets)
    dealer = app.dealer
    operator = dealer.operator if (dealer and dealer.operator_id) else None

    if operator:
        if operator.fee_type.value == "percent":
            operator_fee = total * operator.fee_value / 100
        else:
            operator_fee = operator.fee_value
    else:
        operator_fee = 0.0
    after_operator = total - operator_fee

    if dealer and dealer.fee_type.value == "percent":
        fee = after_operator * dealer.fee_value / 100
    elif dealer:
        fee = dealer.fee_value
    else:
        fee = 0.0
    customer_amount = after_operator - fee

    stype = dealer.settlement_type if dealer else models.SettlementType.DIRECT

    notes = str(form.get("pricing_notes", "")).strip()

    s = app.settlement
    if not s:
        s = models.Settlement(application_id=app.id)
        db.add(s)
        db.flush()

    s.settlement_type  = stype
    s.total_amount     = total
    s.operator_fee_amount = operator_fee
    s.dealer_fee_amount = fee
    s.customer_amount  = customer_amount
    s.pricing_notes    = notes
    s.wm_paid          = True
    s.wm_paid_at       = datetime.now()

    app.status     = AppStatus.SETTLED
    app.updated_at = datetime.now()
    db.commit()
    return RedirectResponse(f"/wm/applications/{app_id}", status_code=302)


@router.post("/applications/{app_id}/close")
def close_app(request: Request, app_id: int, db: Session = Depends(get_db)):
    """종결"""
    u, redir = _check_super(request)
    if redir: return redir
    app = db.query(models.Application).filter(
        models.Application.id == app_id,
        models.Application.status == AppStatus.SETTLED,
    ).first()
    if app:
        app.status = AppStatus.CLOSED
        app.updated_at = datetime.now()
        db.commit()
    return RedirectResponse(f"/wm/applications/{app_id}", status_code=302)


# ─── Phase 3: 데이터 파기 기록 + 인증서 ───────────────

@router.post("/applications/{app_id}/wipe-detail")
async def wipe_detail(request: Request, app_id: int, db: Session = Depends(get_db)):
    """파기 세부 정보 저장 (자산별)"""
    u, redir = _check(request)
    if redir: return redir
    form = await request.form()
    app = db.query(models.Application).filter(models.Application.id == app_id).first()
    if not app:
        return RedirectResponse("/wm/applications", status_code=302)

    wipe_date       = form.get("wipe_date", "")
    wipe_method_val = form.get("wipe_method", "software")
    is_on_site      = bool(form.get("is_on_site"))
    tech_name       = form.get("technician_name", "").strip()
    tech_company    = form.get("technician_company", "").strip()

    for asset in app.assets:
        # 이미 기록이 있으면 갱신, 없으면 생성
        rec = asset.wipe_record
        if not rec:
            rec = models.DataWipeRecord(asset_id=asset.id)
            db.add(rec)
        rec.wipe_date           = wipe_date
        rec.wipe_method         = wipe_method_val
        rec.is_on_site          = is_on_site
        rec.technician_name     = tech_name
        rec.technician_company  = tech_company
        rec.cert_number         = f"WR-{app.id:05d}-{asset.id:04d}"
        rec.updated_at          = datetime.now()

    # APPROVED 상태라면 WIPED로 전이
    if app.status == AppStatus.APPROVED:
        app.status = AppStatus.WIPED
        app.updated_at = datetime.now()

    db.commit()
    return RedirectResponse(f"/wm/applications/{app_id}", status_code=302)


@router.get("/applications/{app_id}/wipe-cert", response_class=HTMLResponse)
def wipe_cert(request: Request, app_id: int, db: Session = Depends(get_db)):
    """데이터 파기 인증서 (인쇄용 HTML)"""
    u, redir = _check(request)
    if redir: return redir
    app = db.query(models.Application).filter(models.Application.id == app_id).first()
    if not app:
        return RedirectResponse("/wm/applications", status_code=302)
    return templates.TemplateResponse(request, "wm/wipe_cert.html", {
        "app": app,
        "issued_at": datetime.now().strftime("%Y년 %m월 %d일"),
    })


@router.get("/applications/{app_id}/edit-logs", response_class=HTMLResponse)
def edit_logs(request: Request, app_id: int, db: Session = Depends(get_db)):
    """단가확정 후 수정이력 조회"""
    u, redir = _check(request)
    if redir: return redir
    app = db.query(models.Application).filter(models.Application.id == app_id).first()
    if not app:
        return RedirectResponse("/wm/applications", status_code=302)
    logs = (
        db.query(models.AssetEditLog)
        .join(models.ServerAsset)
        .filter(models.ServerAsset.application_id == app_id)
        .order_by(models.AssetEditLog.changed_at.desc())
        .all()
    )
    return templates.TemplateResponse(request, "wm/edit_logs.html", {
        "app": app, "logs": logs,
    })


# ─── 서버 시세 기준가 ─────────────────────────────────

@router.get("/price-refs", response_class=HTMLResponse)
def price_refs(request: Request, db: Session = Depends(get_db)):
    u, redir = _check(request)
    if redir: return redir
    refs = db.query(models.ServerPriceRef).order_by(
        models.ServerPriceRef.category, models.ServerPriceRef.base_price
    ).all()
    return templates.TemplateResponse(request, "wm/price_refs.html", {"session": request.session, "refs": refs})


@router.post("/price-refs")
async def price_ref_create(request: Request, db: Session = Depends(get_db)):
    u, redir = _check_super(request)
    if redir: return redir
    form = await request.form()
    try:
        price = int(str(form.get("base_price", 0)).replace(",", "") or 0)
    except ValueError:
        price = 0
    db.add(models.ServerPriceRef(
        category=form.get("category", "서버"),
        manufacturer=form.get("manufacturer", "").strip(),
        model_code=form.get("model_code", "").strip(),
        model_display=form.get("model_display", "").strip(),
        spec_summary=form.get("spec_summary", "").strip(),
        base_price=price,
        updated_at=datetime.now(),
    ))
    db.commit()
    return RedirectResponse("/wm/price-refs", status_code=302)


@router.post("/price-refs/{ref_id}/delete")
def price_ref_delete(request: Request, ref_id: int, db: Session = Depends(get_db)):
    u, redir = _check_super(request)
    if redir: return redir
    ref = db.query(models.ServerPriceRef).filter(models.ServerPriceRef.id == ref_id).first()
    if ref:
        db.delete(ref); db.commit()
    return RedirectResponse("/wm/price-refs", status_code=302)


# ─── 시세 자동완성 API (고객사 신청서 + 검수용) ────────

@router.get("/api/price-refs", response_class=JSONResponse)
def price_ref_api(request: Request, q: str = "", category: str = "", db: Session = Depends(get_db)):
    query = db.query(models.ServerPriceRef).filter(models.ServerPriceRef.is_active == True)
    if category:
        query = query.filter(models.ServerPriceRef.category == category)
    if q:
        like = f"%{q}%"
        query = query.filter(
            models.ServerPriceRef.model_code.ilike(like) |
            models.ServerPriceRef.model_display.ilike(like) |
            models.ServerPriceRef.manufacturer.ilike(like)
        )
    refs = query.order_by(models.ServerPriceRef.base_price).limit(20).all()
    return JSONResponse([{
        "id": r.id, "category": r.category,
        "manufacturer": r.manufacturer, "model_code": r.model_code,
        "model_display": r.model_display, "spec_summary": r.spec_summary,
        "base_price": r.base_price,
    } for r in refs])


@router.get("/price-refs/template")
def price_ref_template(request: Request):
    u, redir = _check(request)
    if redir: return redir

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "기준가"
    ws.append(["분류", "제조사", "모델코드", "표시명", "스펙요약(CPU/RAM)", "기준가(원)"])
    ws.append(["서버", "Dell", "R740", "PowerEdge R740", "Xeon Gold 6230 / 128GB DDR4", 500000])
    ws.append(["스토리지", "NetApp", "FAS2750", "FAS2750 24-Bay", "24×1.8TB SAS", 1200000])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename*=UTF-8''price_ref_template.xlsx"},
    )


@router.post("/price-refs/bulk")
async def price_ref_bulk(request: Request, db: Session = Depends(get_db)):
    u, redir = _check_super(request)
    if redir: return redir

    form = await request.form()
    file = form.get("file")
    if not file or not file.filename:
        return RedirectResponse("/wm/price-refs?error=no_file", status_code=302)

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(BytesIO(content))
    except Exception:
        return RedirectResponse("/wm/price-refs?error=invalid_file", status_code=302)

    ws = wb.active
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        category     = str(row[0] or "서버").strip()
        manufacturer = str(row[1] or "").strip()
        model_code   = str(row[2] or "").strip()
        model_display = str(row[3] or "").strip()
        spec_summary = str(row[4] or "").strip()
        try:
            base_price = int(str(row[5] or 0).replace(",", "").split(".")[0] or 0)
        except (ValueError, TypeError):
            base_price = 0

        if not model_code:
            continue

        existing = db.query(models.ServerPriceRef).filter(
            models.ServerPriceRef.model_code == model_code
        ).first()
        if existing:
            existing.category = category
            existing.manufacturer = manufacturer
            existing.model_display = model_display
            existing.spec_summary = spec_summary
            existing.base_price = base_price
            existing.updated_at = datetime.now()
        else:
            db.add(models.ServerPriceRef(
                category=category, manufacturer=manufacturer,
                model_code=model_code, model_display=model_display,
                spec_summary=spec_summary, base_price=base_price,
                updated_at=datetime.now(),
            ))
        count += 1

    db.commit()
    return RedirectResponse(f"/wm/price-refs?imported={count}", status_code=302)
